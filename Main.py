import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import re
from datetime import datetime, timedelta
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from func.create_pdf import create_summary_pdf_tables_only
from export_functions import ExportManager
from func.format import *
from func.matching import *
import time


class MediaReconciliationTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Media Reconciliation Tool")
        self.root.geometry("1280x1240")  
        self.root.configure(bg="#f0f0f0")

        self.export_manager = ExportManager(self)

        # Data storage
        self.lmrb_data = None
        self.tc_data = None
        self.schedule_data = None
        self.lmrb_filtered = None
        self.tc_filtered = None
        self.theme_mapping = []
        self.matched_data = []
        self.unmatched_tc_data = None
        self.matched_lmrb_indices = []
        self.selected_channel = None
        self.unmatched_tc_data = None
        self.unmatched_lmrb_data = None
        self.tc_data_original = None
        self.program_mismatched_data = None
        self.matched_indices = None
        self.filtered_lmrb_df = None

        # Schedule matching data
        self.schedule_summary_results = []
        self.selected_trailers = []
        self.manually_matched_schedule_items = []

        # Tkinter variables
        self.lmrb_file_var = tk.StringVar()
        self.tc_file_var = tk.StringVar()
        self.schedule_file_var = tk.StringVar()  # Schedule file variable
        self.channel_var = tk.StringVar()
        self.lmrb_theme_var = tk.StringVar()
        self.lmrb_dur_var = tk.StringVar()
        self.tc_theme_var = tk.StringVar()
        self.tc_dur_var = tk.StringVar()
        self.schedule_theme_var = tk.StringVar()
        self.schedule_dur_var = tk.StringVar()  # Schedule duration variable

        # Product filtering variables
        self.selected_products = []  # List of selected product names
        self.available_products = []  # List of all available products
        self.lmrb_product_filtered = None  # LMRB data filtered by products

        # New mapping type variable for 3-way mapping
        self.mapping_type_var = tk.StringVar(value="COMMERCIAL BENEFITS")

        # Manual matching filter variables
        self.tc_match_theme_var = tk.StringVar()
        self.tc_match_date_var = tk.StringVar()
        self.tc_match_program_var = tk.StringVar()
        self.lmrb_match_theme_var = tk.StringVar()
        self.lmrb_match_date_var = tk.StringVar()
        self.lmrb_match_program_var = tk.StringVar()
        self.lmrb_match_duration_var = tk.StringVar()

        # Matching options variables
        self.ignore_date_var = tk.BooleanVar(value=False)
        self.time_tolerance_var = tk.IntVar(value=30)

        # Variables for summary report
        self.report_month_var = tk.StringVar()
        self.estimate_no_var = tk.StringVar()
        self.supplier_invoice_var = tk.StringVar()
        self.po_no_var = tk.StringVar()
        self.invoice_no_var = tk.StringVar()

        # Progress and status variables
        self.progress_var = tk.DoubleVar()
        self.matching_status_var = tk.StringVar(value="Ready to match")
        self.total_lmrb_var = tk.StringVar(value="Total LMRB Records: 0")
        self.matched_count_var = tk.StringVar(value="Matched Records: 0")
        self.unmatched_tc_var = tk.StringVar(value="Unmatched TC Records: 0")
        self.match_rate_var = tk.StringVar(value="Match Rate: 0%")
        self.status_var = tk.StringVar(value="Ready")
        self.unmatched_lmrb_var = tk.StringVar(value="Unmatched LMRB Records: 0")  # Unmatched LMRB variable

        # Manual matching selection variables
        self.selected_tc_row = None
        self.selected_lmrb_row = None
        self.last_matched_tc = None
        self.last_matched_lmrb = None
        self.last_matched_full_lmrb = None

        # Create main frame
        self.main_frame = ttk.Frame(self.root, padding=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, pady=10)

        # Create tabs
        self.upload_tab = ttk.Frame(self.notebook, padding=10)
        self.mapping_tab = ttk.Frame(self.notebook, padding=10)
        self.matching_tab = ttk.Frame(self.notebook, padding=10)
        self.results_tab = ttk.Frame(self.notebook, padding=10)
        self.manual_matching_tab = ttk.Frame(self.notebook, padding=10)
        self.schedule_matching_tab = ttk.Frame(self.notebook, padding=10)
        self.exports_tab = ttk.Frame(self.notebook, padding=10)
        self.summary_report_tab = ttk.Frame(self.notebook, padding=10)

        # Add tabs to notebook
        self.notebook.add(self.upload_tab, text="Upload Data")
        self.notebook.add(self.mapping_tab, text="Theme Mapping")
        self.notebook.add(self.matching_tab, text="Matching")
        self.notebook.add(self.results_tab, text="Results")
        self.notebook.add(self.manual_matching_tab, text="Manual Matching")
        self.notebook.add(self.schedule_matching_tab, text="Schedule Matching")
        self.notebook.add(self.summary_report_tab, text="Summary Report")
        self.notebook.add(self.exports_tab, text="Exports")

        # Disable tabs initially
        self.notebook.tab(1, state="normal")  # Theme Mapping
        self.notebook.tab(2, state="normal")  # Matching
        self.notebook.tab(3, state="normal")  # Results
        self.notebook.tab(4, state="normal")  # Manual Matching
        self.notebook.tab(5, state="normal")  # Schedule Matching
        self.notebook.tab(6, state="normal")  # Summary Report
        self.notebook.tab(7, state="normal")  # Exports

        # Setup UI for each tab
        self.setup_upload_tab()
        self.setup_mapping_tab()
        self.setup_matching_tab()
        self.setup_results_tab()
        self.setup_manual_matching_tab()
        self.setup_schedule_matching_tab()
        self.setup_exports_tab() 
        self.setup_summary_report_tab() 


        # Status bar frame at the bottom
        self.status_frame = ttk.Frame(self.root)
        self.status_frame.pack(side=tk.BOTTOM, fill=tk.X)

        # Status bar label (left)
        self.status_bar = ttk.Label(self.status_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Reset Button (right)
        self.reset_button = ttk.Button(self.status_frame, text="Reset", command=self.reset_app)
        self.reset_button.pack(side=tk.RIGHT, padx=10, pady=2)

        # Apply custom style
        self.apply_style()

    def apply_style(self):
        # Configure ttk style
        style = ttk.Style()
        style.configure("TButton", font=("Arial", 10, "bold"), padding=5)
        style.configure("TLabel", font=("Arial", 10), padding=2)
        style.configure("TFrame", background="#f0f0f0")
        style.configure("Heading.TLabel", font=("Arial", 12, "bold"), padding=5)
        style.configure("Status.TLabel", font=("Arial", 9), padding=2)
        
        # Add this for a blue progress bar
        style.theme_use('clam')
        style.configure("Blue.Horizontal.TProgressbar", background="#17da38", troughcolor="#e0e0e0", thickness=18)
        # Configure treeview style
        style.configure("Treeview", font=("Arial", 9))
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"))

    def sync_data_across_tabs(self):
        """Central function to refresh all tabs and summary after any data change."""
        try:
            # Refresh schedule counts and LMRB counts
            self.refresh_schedule_counts()
            self._refresh_lmrb_counts()
            # Refresh results tab if you have a dedicated refresh function
            if hasattr(self, 'setup_results_tab'):
                self.setup_results_tab()
            # Refresh summary report tab
            if hasattr(self, 'sync_summary_report_data'):
                self.sync_summary_report_data()
            # Add other tab refreshes as needed
        except Exception as e:
            print(f"Error syncing tabs: {e}")


    def setup_exports_tab(self):
        """Setup the Exports tab with working export options."""
        # Title
        title_label = ttk.Label(self.exports_tab, text="Export Options", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=3, pady=10, sticky="w")

        # Description
        desc_label = ttk.Label(self.exports_tab, 
            text="Export detailed analysis and reports from the reconciliation process.",
            wraplength=800)
        desc_label.grid(row=1, column=0, columnspan=3, pady=5, sticky="w")

        # SECTION 1: PROGRAM ANALYSIS
        program_frame = ttk.LabelFrame(self.exports_tab, text="Program Analysis Exports", padding=10)
        program_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Button(program_frame, text="Schedule vs Actual Programs", 
                command=self.export_manager.export_schedule_vs_actual_programs,
                width=35).grid(row=0, column=0, padx=5, pady=5, sticky="w")

        ttk.Label(program_frame, text="Enhanced analysis: Program name matching + Time-based matching\nSee if ads aired in correct programs or moved to different time slots", 
                font=("Arial", 8), wraplength=400).grid(row=0, column=1, padx=10, sticky="w")

        ttk.Button(program_frame, text="Cross-Program Movement", 
                command=self.export_manager.export_cross_program_analysis,
                width=35).grid(row=1, column=0, padx=5, pady=5, sticky="w")

        ttk.Label(program_frame, text="Track which programs gained/lost ads from planned schedule\nIdentify program switching patterns", 
                font=("Arial", 8), wraplength=400).grid(row=1, column=1, padx=10, sticky="w")

        # SECTION 2: PERFORMANCE ANALYSIS
        performance_frame = ttk.LabelFrame(self.exports_tab, text="Performance Analysis", padding=10)
        performance_frame.grid(row=3, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Button(performance_frame, text="Theme Performance Analysis", 
                command=self.export_manager.export_theme_performance,
                width=35).grid(row=0, column=0, padx=5, pady=5, sticky="w")

        ttk.Label(performance_frame, text="Which themes have better matching rates and placement accuracy\nManual vs automatic matching by theme", 
                font=("Arial", 8), wraplength=400).grid(row=0, column=1, padx=10, sticky="w")

        ttk.Button(performance_frame, text="Ad Type Performance", 
                command=self.export_manager.export_ad_type_performance,
                width=35).grid(row=1, column=0, padx=5, pady=5, sticky="w")

        ttk.Label(performance_frame, text="Compare sponsorship vs commercial performance\nPlacement accuracy by advertisement type", 
                font=("Arial", 8), wraplength=400).grid(row=1, column=1, padx=10, sticky="w")

        # SECTION 3: COMPREHENSIVE REPORTS
        comprehensive_frame = ttk.LabelFrame(self.exports_tab, text="Comprehensive Reports", padding=10)
        comprehensive_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Button(comprehensive_frame, text="Comprehensive Report", 
                command=self.export_manager.export_comprehensive_report,
                width=35).grid(row=0, column=0, padx=5, pady=5, sticky="w")

        ttk.Label(comprehensive_frame, text="Complete reconciliation overview with all metrics\nExecutive summary with key insights", 
                font=("Arial", 8), wraplength=400).grid(row=0, column=1, padx=10, sticky="w")

        # SECTION 4: EXPORT STATUS
        status_frame = ttk.LabelFrame(self.exports_tab, text="Export Status", padding=10)
        status_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Label(status_frame, text="📊 Available Exports:", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=5)

        ttk.Label(status_frame, font=("Arial", 8), justify="left").grid(row=1, column=0, sticky="w", pady=5)

        # Make the grid expandable
        self.exports_tab.grid_columnconfigure(1, weight=1)
        for i in range(6):
            self.exports_tab.grid_rowconfigure(i, weight=0)


    def setup_upload_tab(self):
        # Title
        title_label = ttk.Label(self.upload_tab, text="Upload Data Files", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=3, pady=10, sticky="w")

        # LMRB File Upload
        lmrb_label = ttk.Label(self.upload_tab, text="LMRB Data File:")
        lmrb_label.grid(row=1, column=0, sticky="w", pady=5)

        lmrb_entry = ttk.Entry(self.upload_tab, textvariable=self.lmrb_file_var, width=50)
        lmrb_entry.grid(row=1, column=1, padx=5, pady=5)

        lmrb_button = ttk.Button(self.upload_tab, text="Browse", command=self.browse_lmrb_file)
        lmrb_button.grid(row=1, column=2, padx=5, pady=5)

        # TC File Upload
        tc_label = ttk.Label(self.upload_tab, text="TC Data File:")
        tc_label.grid(row=2, column=0, sticky="w", pady=5)

        tc_entry = ttk.Entry(self.upload_tab, textvariable=self.tc_file_var, width=50)
        tc_entry.grid(row=2, column=1, padx=5, pady=5)

        tc_button = ttk.Button(self.upload_tab, text="Browse", command=self.browse_tc_file)
        tc_button.grid(row=2, column=2, padx=5, pady=5)

        # Schedule File Upload
        schedule_label = ttk.Label(self.upload_tab, text="Schedule Data File:")
        schedule_label.grid(row=3, column=0, sticky="w", pady=5)

        schedule_entry = ttk.Entry(self.upload_tab, textvariable=self.schedule_file_var, width=50)
        schedule_entry.grid(row=3, column=1, padx=5, pady=5)

        schedule_button = ttk.Button(self.upload_tab, text="Browse", command=self.browse_schedule_file)
        schedule_button.grid(row=3, column=2, padx=5, pady=5)

        # Channel Selection
        channel_frame = ttk.LabelFrame(self.upload_tab, text="Channel Selection", padding=10)
        channel_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Label(channel_frame, text="Channel:").pack(side=tk.LEFT, padx=5)
        self.channel_combobox = ttk.Combobox(channel_frame, textvariable=self.channel_var, state="readonly", width=25)
        self.channel_combobox.pack(side=tk.LEFT, padx=5)
        self.channel_combobox.bind("<<ComboboxSelected>>", self.on_channel_selected)

        # ✅ NEW: Product Selection
        product_frame = ttk.LabelFrame(self.upload_tab, text="Product Filter (Optional - Select Multiple Products)", padding=10)
        product_frame.grid(row=5, column=0, columnspan=3, sticky="ew", pady=10)

        # Product selection controls
        product_controls_frame = ttk.Frame(product_frame)
        product_controls_frame.pack(fill=tk.X, pady=5)

        ttk.Label(product_controls_frame, text="Available Products:").pack(side=tk.LEFT, padx=5)

        # Dropdown for adding products
        self.product_var = tk.StringVar()
        self.product_combobox = ttk.Combobox(product_controls_frame, textvariable=self.product_var, 
                                            state="readonly", width=30)
        self.product_combobox.pack(side=tk.LEFT, padx=5)

        # Add product button
        add_product_button = ttk.Button(product_controls_frame, text="Add Product", 
                                    command=self.add_selected_product)
        add_product_button.pack(side=tk.LEFT, padx=5)

        # Clear all products button
        clear_products_button = ttk.Button(product_controls_frame, text="Clear All", 
                                        command=self.clear_selected_products)
        clear_products_button.pack(side=tk.LEFT, padx=5)

        # Select all products button
        select_all_button = ttk.Button(product_controls_frame, text="Select All", 
                                    command=self.select_all_products)
        select_all_button.pack(side=tk.LEFT, padx=5)

        # Selected products display
        selected_frame = ttk.Frame(product_frame)
        selected_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        ttk.Label(selected_frame, text="Selected Products:").pack(anchor=tk.W)

        # Listbox to show selected products
        self.selected_products_listbox = tk.Listbox(selected_frame, height=4, selectmode=tk.MULTIPLE)
        self.selected_products_listbox.pack(fill=tk.BOTH, expand=True, pady=2)

        # Remove selected products button
        remove_button = ttk.Button(selected_frame, text="Remove Selected", 
                                command=self.remove_selected_products)
        remove_button.pack(anchor=tk.W, pady=2)

        # Status label
        self.product_filter_status = ttk.Label(product_frame, text="No products selected - showing all products", 
                                            foreground="gray", font=("Arial", 9))
        self.product_filter_status.pack(anchor=tk.W, pady=2)

        # Reset Button next to Load Data
        reset_button = ttk.Button(self.upload_tab, text="Reset", command=self.reset_app)
        reset_button.grid(row=6, column=2, pady=20, sticky="ew")

        # Data Preview Frame
        preview_frame = ttk.LabelFrame(self.upload_tab, text="Data Preview", padding=10)
        preview_frame.grid(row=7, column=0, columnspan=3, sticky="nsew", pady=10)

        # Create notebook for data previews
        preview_notebook = ttk.Notebook(preview_frame)
        preview_notebook.pack(fill=tk.BOTH, expand=True)

        # LMRB Preview Tab
        lmrb_preview_frame = ttk.Frame(preview_notebook)
        preview_notebook.add(lmrb_preview_frame, text="LMRB Data")

        # LMRB Treeview
        self.lmrb_tree = ttk.Treeview(lmrb_preview_frame)
        lmrb_scrolly = ttk.Scrollbar(lmrb_preview_frame, orient="vertical", command=self.lmrb_tree.yview)
        lmrb_scrollx = ttk.Scrollbar(lmrb_preview_frame, orient="horizontal", command=self.lmrb_tree.xview)
        self.lmrb_tree.configure(yscrollcommand=lmrb_scrolly.set, xscrollcommand=lmrb_scrollx.set)

        lmrb_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.lmrb_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        lmrb_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # TC Preview Tab
        tc_preview_frame = ttk.Frame(preview_notebook)
        preview_notebook.add(tc_preview_frame, text="TC Data")

        # TC Treeview
        self.tc_tree = ttk.Treeview(tc_preview_frame)
        tc_scrolly = ttk.Scrollbar(tc_preview_frame, orient="vertical", command=self.tc_tree.yview)
        tc_scrollx = ttk.Scrollbar(tc_preview_frame, orient="horizontal", command=self.tc_tree.xview)
        self.tc_tree.configure(yscrollcommand=tc_scrolly.set, xscrollcommand=tc_scrollx.set)

        tc_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.tc_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        tc_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # Schedule Preview Tab
        schedule_preview_frame = ttk.Frame(preview_notebook)
        preview_notebook.add(schedule_preview_frame, text="Schedule Data")

        # Schedule Treeview
        self.schedule_tree = ttk.Treeview(schedule_preview_frame)
        schedule_scrolly = ttk.Scrollbar(schedule_preview_frame, orient="vertical", command=self.schedule_tree.yview)
        schedule_scrollx = ttk.Scrollbar(schedule_preview_frame, orient="horizontal", command=self.schedule_tree.xview)
        self.schedule_tree.configure(yscrollcommand=schedule_scrolly.set, xscrollcommand=schedule_scrollx.set)

        schedule_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.schedule_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        schedule_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # Make the grid expandable
        self.upload_tab.grid_rowconfigure(7, weight=1)
        self.upload_tab.grid_columnconfigure(1, weight=1)

    def setup_mapping_tab(self):
        # Title
        title_label = ttk.Label(self.mapping_tab, text="Theme Mapping", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=3, pady=10, sticky="w")

        # Instructions - Updated to show schedule as optional
        instructions_label = ttk.Label(self.mapping_tab, 
            text="Map themes across TC, LMRB, and Schedule data. Schedule mapping is optional. Leave TC fields empty for LMRB-Schedule only mappings (sponsorships).",
            wraplength=800)
        instructions_label.grid(row=1, column=0, columnspan=3, pady=5, sticky="w")

        # Theme Selection Frame
        theme_frame = ttk.Frame(self.mapping_tab)
        theme_frame.grid(row=2, column=0, columnspan=9, sticky="ew", pady=10)

        # TC Theme Selection (Optional - row 0)
        tc_theme_label = ttk.Label(theme_frame, text="TC Theme (Optional for sponsorships):")
        tc_theme_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

        self.tc_theme_combobox = ttk.Combobox(theme_frame, textvariable=self.tc_theme_var, width=60)
        self.tc_theme_combobox.grid(row=0, column=1, padx=5, pady=5)

        # TC Duration
        tc_dur_label = ttk.Label(theme_frame, text="Duration:")
        tc_dur_label.grid(row=0, column=2, sticky="w", padx=5, pady=5)

        self.tc_dur_combobox = ttk.Combobox(theme_frame, textvariable=self.tc_dur_var, width=10)
        self.tc_dur_combobox.grid(row=0, column=3, padx=5, pady=5)

        # LMRB Theme Selection (Required - row 1)
        lmrb_theme_label = ttk.Label(theme_frame, text="LMRB Theme (Required):")
        lmrb_theme_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)

        self.lmrb_theme_combobox = ttk.Combobox(theme_frame, textvariable=self.lmrb_theme_var, width=60)
        self.lmrb_theme_combobox.grid(row=1, column=1, padx=5, pady=5)

        # LMRB Duration
        lmrb_dur_label = ttk.Label(theme_frame, text="Duration:")
        lmrb_dur_label.grid(row=1, column=2, sticky="w", padx=5, pady=5)

        self.lmrb_dur_combobox = ttk.Combobox(theme_frame, textvariable=self.lmrb_dur_var, width=10)
        self.lmrb_dur_combobox.grid(row=1, column=3, padx=5, pady=5)

        # Schedule Theme Selection (Optional - row 2)
        schedule_theme_label = ttk.Label(theme_frame, text="Schedule Theme (Optional):")
        schedule_theme_label.grid(row=2, column=0, sticky="w", padx=5, pady=5)

        self.schedule_theme_combobox = ttk.Combobox(theme_frame, textvariable=self.schedule_theme_var, width=60)
        self.schedule_theme_combobox.grid(row=2, column=1, padx=5, pady=5)

        # Schedule Duration
        schedule_dur_label = ttk.Label(theme_frame, text="Duration:")
        schedule_dur_label.grid(row=2, column=2, sticky="w", padx=5, pady=5)

        self.schedule_dur_combobox = ttk.Combobox(theme_frame, textvariable=self.schedule_dur_var, width=10)
        self.schedule_dur_combobox.grid(row=2, column=3, padx=5, pady=5)

        # Mapping Type Selection
        mapping_type_label = ttk.Label(theme_frame, text="Mapping Type:")
        mapping_type_label.grid(row=3, column=0, sticky="w", padx=5, pady=5)

        self.mapping_type_var = tk.StringVar(value="COMMERCIAL BENEFITS")
        mapping_type_combo = ttk.Combobox(theme_frame, textvariable=self.mapping_type_var, 
                                        values=["COMMERCIAL BENEFITS", 
                                                "SPONSORSHIP BENEFITS"],
                                        state="readonly", width=30)
        mapping_type_combo.grid(row=3, column=1, padx=5, pady=5)

        # Add Mapping Button
        add_mapping_button = ttk.Button(theme_frame, text="Add Mapping", command=self.add_theme_mapping)
        add_mapping_button.grid(row=4, column=0, columnspan=4, pady=10)

        # Current Mappings Frame
        mappings_frame = ttk.LabelFrame(self.mapping_tab, text="Current Mappings", padding=10)
        mappings_frame.grid(row=3, column=0, columnspan=3, sticky="nsew", pady=10)

        # Mappings Treeview - Updated columns
        self.mappings_tree = ttk.Treeview(mappings_frame, columns=("mapping_type", "tc_theme", "tc_dur", "lmrb_theme", "lmrb_dur", "schedule_theme", "schedule_dur"))
        self.mappings_tree.heading("mapping_type", text="Type")
        self.mappings_tree.heading("tc_theme", text="TC Theme")
        self.mappings_tree.heading("tc_dur", text="TC Dur")
        self.mappings_tree.heading("lmrb_theme", text="LMRB Theme")
        self.mappings_tree.heading("lmrb_dur", text="LMRB Dur")
        self.mappings_tree.heading("schedule_theme", text="Schedule Theme")
        self.mappings_tree.heading("schedule_dur", text="Schedule Dur")

        self.mappings_tree.column("#0", width=50, stretch=tk.NO)
        self.mappings_tree.column("mapping_type", width=120)
        self.mappings_tree.column("tc_theme", width=150)
        self.mappings_tree.column("tc_dur", width=80)
        self.mappings_tree.column("lmrb_theme", width=150)
        self.mappings_tree.column("lmrb_dur", width=80)
        self.mappings_tree.column("schedule_theme", width=150)
        self.mappings_tree.column("schedule_dur", width=80)

        mappings_scrolly = ttk.Scrollbar(mappings_frame, orient="vertical", command=self.mappings_tree.yview)
        self.mappings_tree.configure(yscrollcommand=mappings_scrolly.set)

        mappings_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.mappings_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Delete Mapping Button
        delete_mapping_button = ttk.Button(self.mapping_tab, text="Delete Selected Mapping", command=self.delete_theme_mapping)
        delete_mapping_button.grid(row=4, column=0, pady=10)

        # Continue Button
        continue_button = ttk.Button(self.mapping_tab, text="Continue to Matching", command=self.continue_to_matching)
        continue_button.grid(row=4, column=2, pady=10, sticky="e")

        # Make the grid expandable
        self.mapping_tab.grid_rowconfigure(3, weight=1)
        self.mapping_tab.grid_columnconfigure(0, weight=1)
        self.mapping_tab.grid_columnconfigure(1, weight=1)
        self.mapping_tab.grid_columnconfigure(2, weight=1)

    def setup_schedule_matching_tab(self):
        """Setup count-based schedule matching tab - SIMPLIFIED UI"""
        # Title
        title_label = ttk.Label(self.schedule_matching_tab, text="Count-Based Schedule Reconciliation", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky="w")  # Changed to columnspan=2

        # Description
        desc_label = ttk.Label(self.schedule_matching_tab, 
            text="Compare scheduled vs actual ad counts by theme and duration. Fully fulfilled schedules are automatically hidden.",
            wraplength=800)
        desc_label.grid(row=1, column=0, columnspan=2, pady=5, sticky="w")  # Changed to columnspan=2

        # Control Frame
        control_frame = ttk.Frame(self.schedule_matching_tab)
        control_frame.grid(row=2, column=0, columnspan=2, sticky="ew", pady=10)  # Changed to columnspan=2

        # Run Analysis Button
        run_button = ttk.Button(control_frame, text="🔍 Load Schedule Data", command=self.load_schedule_counts)
        run_button.pack(side=tk.LEFT, padx=5)

        # Refresh Button
        refresh_button = ttk.Button(control_frame, text="🔄 Refresh", command=self.refresh_schedule_counts)
        refresh_button.pack(side=tk.LEFT, padx=5)

        # Export Button
        export_button = ttk.Button(control_frame, text="📊 Export Analysis", command=self.export_schedule_analysis)
        export_button.pack(side=tk.LEFT, padx=5)

        # === SECTION 1: SCHEDULE THEME COUNTS (ONLY PENDING/MISSED) ===
        schedule_count_frame = ttk.LabelFrame(self.schedule_matching_tab, text="📅 Pending Schedule Themes (Missing/Partially Fulfilled)", padding=10)
        schedule_count_frame.grid(row=3, column=0, sticky="nsew", pady=10, padx=(0,5))

        # Schedule Counts Tree - UPDATED to show only pending
        self.schedule_counts_tree = ttk.Treeview(schedule_count_frame,
            columns=("theme", "duration", "ad_type", "remaining_count", "status"),
            show='headings', selectmode='browse')

        self.schedule_counts_tree.heading("theme", text="Schedule Theme")
        self.schedule_counts_tree.heading("duration", text="Duration")
        self.schedule_counts_tree.heading("ad_type", text="Ad Type") 
        self.schedule_counts_tree.heading("remaining_count", text="Missing Count")
        self.schedule_counts_tree.heading("status", text="Status")

        self.schedule_counts_tree.column("theme", width=180)
        self.schedule_counts_tree.column("duration", width=80)
        self.schedule_counts_tree.column("ad_type", width=100)
        self.schedule_counts_tree.column("remaining_count", width=100)
        self.schedule_counts_tree.column("status", width=120)

        # Scrollbars
        schedule_scrolly = ttk.Scrollbar(schedule_count_frame, orient="vertical", command=self.schedule_counts_tree.yview)
        self.schedule_counts_tree.configure(yscrollcommand=schedule_scrolly.set)

        schedule_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.schedule_counts_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # === SECTION 2: UNMATCHED LMRB THEME COUNTS (WITH DURATION) ===
        lmrb_count_frame = ttk.LabelFrame(self.schedule_matching_tab, text="📺 Available LMRB Theme Counts", padding=10)
        lmrb_count_frame.grid(row=3, column=1, sticky="nsew", pady=10, padx=(5,0))

        # LMRB Counts Tree - same as before
        self.lmrb_counts_tree = ttk.Treeview(lmrb_count_frame,
            columns=("theme", "duration", "advertiser", "count", "sample_program"),
            show='headings', selectmode='browse')

        self.lmrb_counts_tree.heading("theme", text="LMRB Theme")
        self.lmrb_counts_tree.heading("duration", text="Duration")
        self.lmrb_counts_tree.heading("advertiser", text="Advertiser")
        self.lmrb_counts_tree.heading("count", text="Available Count")
        self.lmrb_counts_tree.heading("sample_program", text="Sample Program")

        self.lmrb_counts_tree.column("theme", width=150)
        self.lmrb_counts_tree.column("duration", width=80)
        self.lmrb_counts_tree.column("advertiser", width=120)
        self.lmrb_counts_tree.column("count", width=100)
        self.lmrb_counts_tree.column("sample_program", width=120)

        # Scrollbars
        lmrb_scrolly = ttk.Scrollbar(lmrb_count_frame, orient="vertical", command=self.lmrb_counts_tree.yview)
        self.lmrb_counts_tree.configure(yscrollcommand=lmrb_scrolly.set)

        lmrb_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.lmrb_counts_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # === SECTION 3: MANUAL MATCHING CONTROLS (UPDATED GRID) ===
        matching_frame = ttk.LabelFrame(self.schedule_matching_tab, text="🔗 Manual Theme Matching", padding=10)
        matching_frame.grid(row=4, column=0, columnspan=2, sticky="ew", pady=10)  # Changed to columnspan=2

        # Selection display
        selection_label = ttk.Label(matching_frame, text="Selected Themes:")
        selection_label.grid(row=0, column=0, sticky="w", pady=5)

        self.selected_schedule_var = tk.StringVar(value="No pending schedule theme selected")
        self.selected_lmrb_var = tk.StringVar(value="No LMRB theme selected")

        schedule_selection_label = ttk.Label(matching_frame, textvariable=self.selected_schedule_var, foreground="red")  # Changed to red for pending
        schedule_selection_label.grid(row=1, column=0, sticky="w", padx=20)

        lmrb_selection_label = ttk.Label(matching_frame, textvariable=self.selected_lmrb_var, foreground="green")
        lmrb_selection_label.grid(row=2, column=0, sticky="w", padx=20)

        # Action buttons
        button_frame = ttk.Frame(matching_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=10)

        # Match Selected Button
        self.match_themes_button = ttk.Button(button_frame, text="✅ Match & Fulfill Schedule", 
                                            command=self.match_and_consume_counts, state=tk.DISABLED)
        self.match_themes_button.pack(side=tk.LEFT, padx=5)

        # Clear Selection Button
        clear_button = ttk.Button(button_frame, text="🗑️ Clear Selection", command=self.clear_theme_selection)
        clear_button.pack(side=tk.LEFT, padx=5)

        # Bind selection events
        self.schedule_counts_tree.bind('<<TreeviewSelect>>', self.on_schedule_theme_select)
        self.lmrb_counts_tree.bind('<<TreeviewSelect>>', self.on_lmrb_theme_select)

        # Make grid expandable - UPDATED for 2 columns
        self.schedule_matching_tab.grid_rowconfigure(3, weight=1)
        self.schedule_matching_tab.grid_columnconfigure(0, weight=1)
        self.schedule_matching_tab.grid_columnconfigure(1, weight=1)


    def refresh_schedule_counts(self):
        """Refresh schedule counts to reflect latest matches and show only pending items"""
        try:
            if not hasattr(self, 'schedule_data') or self.schedule_data is None:
                messagebox.showinfo("Info", "Please load schedule data first using 'Load Schedule Data' button.")
                return
                
            print("\n🔄 Refreshing schedule counts...")
            
            # Reload the analysis with current match state
            self.load_schedule_counts()
            
            # Count how many themes are now fulfilled vs pending
            pending_count = len(self.schedule_count_data) if hasattr(self, 'schedule_count_data') else 0
            
            messagebox.showinfo("Schedule Refreshed", 
                f"✅ Schedule data refreshed!\n\n"
                f"🔴 Pending themes: {pending_count}\n"
                f"💡 Fulfilled themes are automatically hidden\n\n"
                f"🎯 Only themes needing attention are shown in the list.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error refreshing schedule counts: {str(e)}")
            import traceback
            traceback.print_exc()

    def export_schedule_analysis(self):
        """Export schedule analysis with consumption data to Excel"""
        try:
            if not hasattr(self, 'schedule_count_data') or not hasattr(self, 'lmrb_count_data'):
                messagebox.showerror("Error", "No schedule analysis data available. Please load data first.")
                return
                
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"schedule_analysis_consumption_{self.selected_channel or 'channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if filename:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Schedule counts with consumption data
                    if hasattr(self, 'schedule_count_data'):
                        schedule_df = pd.DataFrame(self.schedule_count_data)
                        # Remove internal keys for export
                        export_schedule_df = schedule_df.drop(['consumption_key'], axis=1, errors='ignore')
                        export_schedule_df.to_excel(writer, sheet_name='Schedule Consumption', index=False)
                    
                    # LMRB counts  
                    if hasattr(self, 'lmrb_count_data'):
                        lmrb_export_data = []
                        for item in self.lmrb_count_data:
                            export_item = {k: v for k, v in item.items() if k != 'group_data'}
                            lmrb_export_data.append(export_item)
                        pd.DataFrame(lmrb_export_data).to_excel(writer, sheet_name='LMRB Counts', index=False)
                    
                    # Consumption summary
                    if hasattr(self, 'schedule_consumption_tracker'):
                        consumption_data = []
                        for key, consumed in self.schedule_consumption_tracker.items():
                            # Parse the consumption key
                            parts = key.split('_')
                            if len(parts) >= 3:
                                theme = '_'.join(parts[:-2])
                                duration = parts[-2]
                                ad_type = parts[-1]
                            else:
                                theme = key
                                duration = "Unknown"
                                ad_type = "Unknown"
                            
                            consumption_data.append({
                                'Theme': theme,
                                'Duration': duration,
                                'Ad_Type': ad_type,
                                'Consumed_Count': consumed,
                                'Consumption_Key': key
                            })
                        
                        if consumption_data:
                            pd.DataFrame(consumption_data).to_excel(writer, sheet_name='Consumption Tracker', index=False)
                    
                    # Manual matches log
                    if hasattr(self, 'manually_matched_schedule_items'):
                        manual_matches_df = pd.DataFrame(self.manually_matched_schedule_items)
                        manual_matches_df.to_excel(writer, sheet_name='Manual Matches Log', index=False)
                    
                messagebox.showinfo("Success", f"Schedule analysis exported to {filename}")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting schedule analysis: {str(e)}")
            import traceback
            traceback.print_exc()

    def _analyze_unmatched_lmrb_with_duration(self):
        """FIXED: Analyze unmatched LMRB data considering consumption tracking"""
        try:
            if self.unmatched_lmrb_data is None or self.unmatched_lmrb_data.empty:
                print("No unmatched LMRB data available")
                return []

            lmrb_counts = []
            
            # Find columns
            theme_col = self._find_column(self.unmatched_lmrb_data, ['Advt_Theme', 'theme', 'aadvt'])
            dur_col = self._find_column(self.unmatched_lmrb_data, ['Dur', 'duration'])
            advertiser_col = self._find_column(self.unmatched_lmrb_data, ['advertiser', 'client'])
            program_col = self._find_column(self.unmatched_lmrb_data, ['Program', 'program'])
            
            if not theme_col or not dur_col:
                print(f"❌ Required LMRB columns not found - Theme: {theme_col}, Duration: {dur_col}")
                return []

            print(f"📋 Analyzing unmatched LMRB with columns - Theme: {theme_col}, Duration: {dur_col}")

            # Group by theme, duration, and advertiser
            if advertiser_col:
                grouped = self.unmatched_lmrb_data.groupby([theme_col, dur_col, advertiser_col])
            else:
                grouped = self.unmatched_lmrb_data.groupby([theme_col, dur_col])

            for group_key, group_data in grouped:
                try:
                    if isinstance(group_key, tuple) and len(group_key) >= 2:
                        theme = group_key[0]
                        duration = group_key[1]
                        advertiser = group_key[2] if len(group_key) > 2 else "Unknown"
                    else:
                        theme = group_key
                        duration = "Unknown"
                        advertiser = "Unknown"

                    original_count = len(group_data)
                    sample_program = group_data[program_col].iloc[0] if program_col else "Unknown"

                    # Calculate consumed count for this LMRB theme
                    consumption_key = f"{theme}_{duration}_{advertiser}"
                    consumed_count = 0
                    
                    if hasattr(self, 'lmrb_consumption_tracker'):
                        consumed_count = self.lmrb_consumption_tracker.get(consumption_key, 0)

                    # Calculate remaining count
                    remaining_count = max(0, original_count - consumed_count)
                    
                    # Only add if there are remaining items
                    if remaining_count > 0:
                        # Get remaining group data (subtract consumed items)
                        remaining_group_data = group_data.tail(remaining_count) if consumed_count > 0 else group_data

                        lmrb_counts.append({
                            'theme': theme,
                            'duration': duration,
                            'advertiser': advertiser,
                            'count': remaining_count,
                            'original_count': original_count,
                            'consumed_count': consumed_count,
                            'sample_program': sample_program,
                            'group_data': remaining_group_data,  # Only remaining data
                            'consumption_key': consumption_key
                        })
                        
                        if consumed_count > 0:
                            print(f"📊 LMRB theme {theme} ({advertiser}): {original_count} original - {consumed_count} consumed = {remaining_count} remaining")
                    else:
                        print(f"✅ LMRB theme {theme} ({advertiser}) fully consumed: {original_count}/{original_count}")

                except Exception as e:
                    print(f"Error processing LMRB group: {e}")
                    continue

            # Sort by theme and duration
            lmrb_counts.sort(key=lambda x: (x['theme'], str(x['duration'])))
            print(f"📊 Found {len(lmrb_counts)} LMRB themes with remaining items")
            return lmrb_counts

        except Exception as e:
            print(f"Error analyzing LMRB with duration and consumption: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _populate_consumption_summary_tree(self, summary_data):
        """Populate the consumption summary tree"""
        try:
            # Clear existing items
            for item in self.consumption_summary_tree.get_children():
                self.consumption_summary_tree.delete(item)
            
            # Add new items
            for summary_item in summary_data:
                self.consumption_summary_tree.insert('', 'end', values=(
                    summary_item['metric'],
                    summary_item['value']
                ))
                    
        except Exception as e:
            print(f"Error populating consumption summary tree: {e}")

    def _create_consumption_summary(self, schedule_counts, lmrb_counts):
        """Create consumption summary statistics"""
        try:
            total_schedule_entries = len(schedule_counts)
            total_lmrb_entries = len(lmrb_counts)
            
            total_original = sum(item['original_count'] for item in schedule_counts)
            total_remaining = sum(item['remaining_count'] for item in schedule_counts)
            total_consumed = sum(item['consumed_count'] for item in schedule_counts)
            
            fully_consumed = len([item for item in schedule_counts if item['remaining_count'] == 0])
            partially_consumed = len([item for item in schedule_counts if 0 < item['remaining_count'] < item['original_count']])
            available = len([item for item in schedule_counts if item['consumed_count'] == 0])

            return [
                {'metric': 'Schedule Themes', 'value': total_schedule_entries},
                {'metric': 'LMRB Themes', 'value': total_lmrb_entries},
                {'metric': 'Total Original Count', 'value': total_original},
                {'metric': 'Total Remaining Count', 'value': total_remaining},
                {'metric': 'Total Consumed Count', 'value': total_consumed},
                {'metric': 'Fully Consumed Themes', 'value': fully_consumed},
                {'metric': 'Partially Consumed', 'value': partially_consumed},
                {'metric': 'Available Themes', 'value': available},
                {'metric': 'Consumption Rate', 'value': f'{total_consumed/max(total_original,1)*100:.1f}%'}
            ]
        except Exception as e:
            print(f"Error creating consumption summary: {e}")
            return [{'metric': 'Error', 'value': 'Could not calculate'}]

    def get_planned_count_with_consumption(self, theme, duration):
        """
        UPDATED: Get planned count accounting for consumed schedule entries
        """
        if self.schedule_data is None or self.schedule_data.empty:
            return 0

        # Find schedule theme and duration columns
        theme_col = self.find_schedule_column(['theme', 'aadvt', 'advt_theme'])
        dur_col = self.find_schedule_column(['dur', 'duration'])

        if not theme_col or not dur_col:
            return 0

        try:
            # Get original planned count
            filtered = self.schedule_data[
                (self.schedule_data[theme_col].astype(str).str.strip() == str(theme).strip()) & 
                (self.schedule_data[dur_col].astype(str).str.strip() == str(duration).strip())
            ]
            
            original_count = len(filtered)
            
            # Calculate consumed count using consumption tracker
            consumed_count = 0
            if hasattr(self, 'schedule_consumption_tracker'):
                # Try different possible consumption keys
                possible_keys = [
                    f"{theme}_{duration}_Unknown",
                    f"{theme}_{duration}_Commercial", 
                    f"{theme}_{duration}_Sponsorship"
                ]
                
                for key in possible_keys:
                    consumed_count += self.schedule_consumption_tracker.get(key, 0)
            
            # Also check manually matched items (backward compatibility)
            if hasattr(self, 'manually_matched_schedule_items'):
                for match_record in self.manually_matched_schedule_items:
                    if (match_record.get('schedule_theme') == theme and 
                        str(match_record.get('schedule_duration', '')) == str(duration)):
                        consumed_count += match_record.get('schedule_count_consumed', 0)
            
            # Return remaining count (original - consumed)
            remaining_count = max(0, original_count - consumed_count)
            
            print(f"📊 Planned count for {theme} ({duration}s): {original_count} original - {consumed_count} consumed = {remaining_count} remaining")
            return remaining_count
            
        except Exception as e:
            print(f"Error calculating planned count with consumption for {theme}: {e}")
            return 0

    def match_and_consume_counts(self):
        """FIXED: Properly update LMRB counts after consumption and track schedule ad type."""
        try:
            if not hasattr(self, 'selected_schedule_theme') or not hasattr(self, 'selected_lmrb_theme'):
                messagebox.showerror("Error", "Please select both schedule and LMRB themes")
                return

            schedule_theme = self.selected_schedule_theme['theme']
            schedule_duration = self.selected_schedule_theme['duration']
            schedule_remaining = self.selected_schedule_theme['remaining_count']
            schedule_consumption_key = self.selected_schedule_theme['consumption_key']
            
            lmrb_theme = self.selected_lmrb_theme['theme']
            lmrb_duration = self.selected_lmrb_theme['duration']
            lmrb_advertiser = self.selected_lmrb_theme['advertiser']
            lmrb_count = self.selected_lmrb_theme['count']
            lmrb_group_data = self.selected_lmrb_theme['group_data']

            consumption_amount = min(schedule_remaining, lmrb_count)
            
            if consumption_amount <= 0:
                messagebox.showwarning("Warning", "No items available to consume!")
                return

            # FIXED: Get the schedule ad type for proper categorization
            schedule_ad_type = self._get_schedule_ad_type(schedule_theme)

            result = messagebox.askyesno("Confirm Dual Consumption Match", 
                f"🔄 DUAL CONSUMPTION MATCHING\n\n"
                f"📅 SCHEDULE: {schedule_theme} ({schedule_duration}s) - Will consume: {consumption_amount}\n"
                f"📺 LMRB: {lmrb_theme} ({lmrb_advertiser}) - Will consume: {consumption_amount}\n"
                f"🏷️ Ad Type: {schedule_ad_type}\n\n"
                f"After match:\n"
                f"Schedule remaining: {schedule_remaining - consumption_amount}\n"
                f"LMRB remaining: {lmrb_count - consumption_amount}\n\n"
                f"Continue?")

            if not result:
                return

            # Initialize trackers
            if not hasattr(self, 'schedule_consumption_tracker'):
                self.schedule_consumption_tracker = {}
            if not hasattr(self, 'lmrb_consumption_tracker'):
                self.lmrb_consumption_tracker = {}

            # Update consumption tracking
            schedule_current_consumed = self.schedule_consumption_tracker.get(schedule_consumption_key, 0)
            self.schedule_consumption_tracker[schedule_consumption_key] = schedule_current_consumed + consumption_amount

            lmrb_consumption_key = f"{lmrb_theme}_{lmrb_duration}_{lmrb_advertiser}"
            lmrb_current_consumed = self.lmrb_consumption_tracker.get(lmrb_consumption_key, 0)
            self.lmrb_consumption_tracker[lmrb_consumption_key] = lmrb_current_consumed + consumption_amount

            # Add LMRB records to matched data
            lmrb_records_to_add = lmrb_group_data.head(consumption_amount)
            
            for _, record in lmrb_records_to_add.iterrows():
                matched_record = record.to_dict()
                matched_record['Match_Type'] = 'Dual-Consumption Manual Match'
                matched_record['Match_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                matched_record['Original_Schedule_Theme'] = schedule_theme
                matched_record['Original_LMRB_Theme'] = lmrb_theme
                
                # FIXED: Add schedule ad type for proper categorization
                matched_record['Schedule_Ad_Type'] = schedule_ad_type
                matched_record['Schedule_Duration'] = schedule_duration
                
                # Add to matched data properly
                if isinstance(self.matched_data, list):
                    self.matched_data.append(matched_record)
                else:
                    if self.matched_data is None:
                        self.matched_data = []
                    self.matched_data.append(matched_record)

            # Remove consumed LMRB entries from unmatched data
            indices_to_remove = lmrb_records_to_add.index.tolist()
            self.unmatched_lmrb_data = self.unmatched_lmrb_data.drop(indices_to_remove).reset_index(drop=True)

            # Update displays
            print("🔄 Updating displays after consumption...")
            self.load_schedule_counts()
            self._refresh_lmrb_counts_after_consumption()
            self.sync_data_across_tabs()
            self.clear_theme_selection()

            remaining_after_match = schedule_remaining - consumption_amount
            fulfillment_message = f"\n🎉 '{schedule_theme}' remaining: {remaining_after_match}"
            if remaining_after_match == 0:
                fulfillment_message = f"\n🎉 '{schedule_theme}' is now FULLY FULFILLED!"

            messagebox.showinfo("Consumption Success", 
                f"✅ Consumption completed!\n\n"
                f"📊 SCHEDULE: {schedule_theme} consumed {consumption_amount} (Ad Type: {schedule_ad_type})\n"
                f"📊 LMRB: {lmrb_theme} ({lmrb_advertiser}) consumed {consumption_amount}\n"
                f"📊 LMRB remaining: {lmrb_count - consumption_amount}\n"
                f"{fulfillment_message}\n\n"
                f"🔄 Displays updated!")

        except Exception as e:
            messagebox.showerror("Error", f"Error in consumption matching: {str(e)}")
            import traceback
            traceback.print_exc()

    def _get_schedule_ad_type(self, schedule_theme):
        """Get ad type from theme mapping for this schedule theme."""
        try:
            # First try to find exact theme mapping
            for mapping in self.theme_mapping:
                if mapping.get('schedule_theme') == schedule_theme:
                    mapping_type = mapping.get('mapping_type', 'COMMERCIAL BENEFITS')
                    print(f"  🏷️ Found ad type '{mapping_type}' for schedule theme '{schedule_theme}'")
                    return mapping_type
            
            # If no exact mapping found, check if it's in schedule data with ad type column
            if hasattr(self, 'schedule_data') and self.schedule_data is not None:
                schedule_theme_col = self._find_schedule_column(['theme', 'aadvt', 'advt_theme'])
                schedule_adtype_col = self._find_schedule_column(['advertisement_type', 'ad_type'])
                
                if schedule_theme_col and schedule_adtype_col:
                    matching_rows = self.schedule_data[
                        self.schedule_data[schedule_theme_col].astype(str).str.strip() == str(schedule_theme).strip()
                    ]
                    
                    if not matching_rows.empty:
                        ad_type_value = matching_rows[schedule_adtype_col].iloc[0]
                        if pd.notna(ad_type_value):
                            ad_type_str = str(ad_type_value).upper()
                            if 'SPONSOR' in ad_type_str:
                                print(f"  🏷️ Found schedule ad type 'SPONSORSHIP' for theme '{schedule_theme}'")
                                return 'SPONSORSHIP BENEFITS'
                            elif 'COMMERCIAL' in ad_type_str:
                                print(f"  🏷️ Found schedule ad type 'COMMERCIAL' for theme '{schedule_theme}'")
                                return 'COMMERCIAL BENEFITS'
            
            print(f"  ⚠️ No ad type found for schedule theme '{schedule_theme}', defaulting to COMMERCIAL")
            return 'COMMERCIAL BENEFITS'
            
        except Exception as e:
            print(f"Error getting schedule ad type for {schedule_theme}: {e}")
            return 'COMMERCIAL BENEFITS'

    def _refresh_lmrb_counts_after_consumption(self):
        """FIXED: Force refresh LMRB counts immediately after consumption."""
        try:
            if not hasattr(self, 'lmrb_count_data') or not self.lmrb_count_data:
                print("⚠️ No LMRB count data to refresh")
                return

            print("🔄 Force refreshing LMRB counts after consumption...")

            if not hasattr(self, 'lmrb_consumption_tracker'):
                self.lmrb_consumption_tracker = {}

            updated_lmrb_counts = []
            
            for lmrb_item in self.lmrb_count_data:
                try:
                    theme = lmrb_item['theme']
                    duration = lmrb_item['duration']
                    advertiser = lmrb_item['advertiser']
                    original_count = lmrb_item.get('original_count', lmrb_item['count'])
                    
                    # Get consumed amount
                    consumption_key = f"{theme}_{duration}_{advertiser}"
                    consumed_amount = self.lmrb_consumption_tracker.get(consumption_key, 0)
                    
                    # Calculate remaining
                    remaining_count = max(0, original_count - consumed_amount)
                    
                    # FIXED: Only add if remaining > 0
                    if remaining_count > 0:
                        updated_item = lmrb_item.copy()
                        updated_item['count'] = remaining_count
                        updated_item['original_count'] = original_count
                        updated_item['consumed_count'] = consumed_amount
                        
                        # FIXED: Update group_data to reflect remaining items
                        if 'group_data' in lmrb_item and consumed_amount > 0:
                            original_group_data = lmrb_item['group_data']
                            # Remove consumed items from the beginning
                            remaining_group_data = original_group_data.tail(remaining_count).reset_index(drop=True)
                            updated_item['group_data'] = remaining_group_data
                        
                        updated_lmrb_counts.append(updated_item)
                        print(f"   📊 {theme} ({advertiser}): {original_count} → {remaining_count}")
                    else:
                        print(f"   ✅ {theme} ({advertiser}): FULLY CONSUMED")
                            
                except Exception as e:
                    print(f"❌ Error processing LMRB item: {e}")
                    continue

            # FIXED: Update stored data and display immediately
            self.lmrb_count_data = updated_lmrb_counts
            self._populate_lmrb_counts_tree(updated_lmrb_counts)
            
            print(f"✅ LMRB counts refreshed: {len(updated_lmrb_counts)} themes remaining")

        except Exception as e:
            print(f"❌ Error in force refresh: {e}")
            import traceback
            traceback.print_exc()

    def _refresh_lmrb_counts(self):
        """Refresh LMRB counts display after consumption"""
        try:
            if hasattr(self, 'lmrb_count_data'):
                # Recalculate LMRB counts with consumption
                updated_lmrb_counts = []
                
                for lmrb_item in self.lmrb_count_data:
                    theme = lmrb_item['theme']
                    duration = lmrb_item['duration']
                    advertiser = lmrb_item['advertiser']
                    original_count = lmrb_item['count']
                    
                    # Calculate consumed amount
                    consumption_key = f"{theme}_{duration}_{advertiser}"
                    consumed_amount = self.lmrb_consumption_tracker.get(consumption_key, 0)
                    remaining_count = max(0, original_count - consumed_amount)
                    
                    # Only add if there are remaining items
                    if remaining_count > 0:
                        updated_item = lmrb_item.copy()
                        updated_item['count'] = remaining_count
                        updated_item['original_count'] = original_count
                        updated_item['consumed_count'] = consumed_amount
                        updated_lmrb_counts.append(updated_item)
                
                # Update the tree display
                self._populate_lmrb_counts_tree(updated_lmrb_counts)
                self.lmrb_count_data = updated_lmrb_counts
                
                print(f"📊 LMRB counts refreshed: {len(updated_lmrb_counts)} themes with remaining items")

        except Exception as e:
            print(f"Error refreshing LMRB counts: {e}")

    def load_schedule_counts(self):
        """ENHANCED: Load schedule counts showing only themes with missing items"""
        try:
            if self.schedule_data is None:
                messagebox.showerror("Error", "No schedule data available")
                return

            print("\n" + "="*60)
            print("🔍 ENHANCED: Loading Schedule Counts (Pending Items Only)")
            print("="*60)

            # Initialize consumption trackers if not exist
            if not hasattr(self, 'schedule_consumption_tracker'):
                self.schedule_consumption_tracker = {}
            if not hasattr(self, 'lmrb_consumption_tracker'):
                self.lmrb_consumption_tracker = {}

            # Analyze pending schedule items (only those with missing counts)
            print("📊 Analyzing pending schedule themes...")
            pending_schedule_counts = self._analyze_schedule_with_duration()
            self._populate_schedule_counts_tree(pending_schedule_counts)

            # Analyze available LMRB counts
            print("📊 Analyzing available LMRB themes...")
            lmrb_counts = self._analyze_unmatched_lmrb_with_duration()
            self._populate_lmrb_counts_tree(lmrb_counts)

            # Store for later use
            self.schedule_count_data = pending_schedule_counts
            self.lmrb_count_data = lmrb_counts

            # Calculate summary stats
            total_pending_themes = len(pending_schedule_counts)
            total_lmrb_themes = len(lmrb_counts)
            total_missing_items = sum(item['remaining_count'] for item in pending_schedule_counts)
            total_available_items = sum(item['count'] for item in lmrb_counts)

            # Show results
            if pending_schedule_counts:
                messagebox.showinfo("Schedule Analysis Complete", 
                    f"✅ ENHANCED schedule analysis completed!\n\n"
                    f"🔴 PENDING THEMES:\n"
                    f"• Themes with missing items: {total_pending_themes}\n"
                    f"• Total missing items: {total_missing_items}\n\n"
                    f"📺 AVAILABLE LMRB:\n"
                    f"• Available themes: {total_lmrb_themes}\n"
                    f"• Available items: {total_available_items}\n\n"
                    f"💡 Fulfilled themes are hidden from the pending list!\n"
                    f"🎯 Only themes needing attention are shown.")
            else:
                messagebox.showinfo("🎉 All Schedules Fulfilled!", 
                    f"🎉 Congratulations! All schedule themes are fulfilled!\n\n"
                    f"📊 SUMMARY:\n"
                    f"• No pending themes remaining\n"
                    f"• Available LMRB themes: {total_lmrb_themes}\n"
                    f"• Available LMRB items: {total_available_items}\n\n"
                    f"✅ Schedule reconciliation is complete!")

            print("="*60)
            print("🎯 ENHANCED Schedule Analysis Results:")
            print(f"   🔴 Pending themes: {total_pending_themes}")
            print(f"   📉 Missing items: {total_missing_items}")
            print(f"   📺 Available LMRB: {total_lmrb_themes}")
            print("="*60)

        except Exception as e:
            messagebox.showerror("Error", f"Error loading enhanced schedule data: {str(e)}")
            import traceback
            traceback.print_exc()

    def _analyze_schedule_with_duration(self):
        """
        FIXED: Analyze schedule data using summary report logic to show only truly pending themes
        """
        try:
            print("🔍 Analyzing schedule with duration using summary report logic...")
            
            # Update schedule mapped themes first
            self.update_schedule_mapped_themes()
            
            # Create grouped schedule DataFrame
            grouped_schedule = self.create_grouped_schedule_df()
            if grouped_schedule is None:
                print("❌ No grouped schedule data available")
                return []
            
            # Get current matched data for accurate counting
            matched_lmrb_data = self.get_current_matched_data_df()
            
            pending_schedule_themes = []
            fulfilled_count = 0
            
            # Find required columns in grouped schedule
            schedule_theme_col = self._find_schedule_column(['theme', 'aadvt', 'advt_theme'])
            schedule_dur_col = self._find_schedule_column(['dur', 'duration'])
            schedule_adtype_col = self._find_schedule_column(['advertisement_type', 'ad_type'])
            
            if not schedule_adtype_col:
                schedule_adtype_col = 'temp_ad_type'
            
            print(f"📊 Processing {len(grouped_schedule)} grouped schedule combinations...")
            
            for idx, group in grouped_schedule.iterrows():
                try:
                    # Extract data from grouped schedule
                    schedule_theme = group[schedule_theme_col]
                    schedule_duration = str(group[schedule_dur_col])
                    planned_count = group['planned_count']
                    mapped_theme = group['mapped_theme']
                    has_mapping = group['has_mapping']
                    
                    # Get ad type
                    ad_type = group[schedule_adtype_col] if schedule_adtype_col in group else 'COMMERCIAL BENEFITS'
                    
                    print(f"\n🔍 Checking: '{schedule_theme}' ({schedule_duration}s) - Planned: {planned_count}")
                    
                    # Calculate total aired count using the same logic as summary report
                    total_aired_count = 0
                    
                    # 1. Automatic matches from LMRB data
                    if mapped_theme and has_mapping:
                        automatic_count = self._get_aired_count_by_mapped_theme(
                            mapped_theme, schedule_duration, matched_lmrb_data
                        )
                        total_aired_count += automatic_count
                        print(f"   📺 Automatic matches: {automatic_count}")
                    
                    # 2. Manual matches
                    manual_count = self._get_manual_match_count(schedule_theme, schedule_duration)
                    total_aired_count += manual_count
                    if manual_count > 0:
                        print(f"   🖐️ Manual matches: {manual_count}")
                    
                    # 3. Consumption matches
                    consumption_count = self._get_consumption_match_count(schedule_theme, schedule_duration)
                    total_aired_count += consumption_count
                    if consumption_count > 0:
                        print(f"   🔄 Consumption matches: {consumption_count}")
                    
                    print(f"   🎯 Total aired: {total_aired_count}")
                    
                    # Calculate missing count
                    missing_count = max(0, planned_count - total_aired_count)
                    
                    # ONLY ADD TO PENDING LIST IF THERE ARE MISSING ITEMS
                    if missing_count > 0:
                        # Determine status
                        if total_aired_count > 0:
                            status = f"🟡 PARTIAL ({total_aired_count}/{planned_count} done, {missing_count} missing)"
                        else:
                            status = f"🔴 PENDING ({missing_count} missing)"
                        
                        # Create consumption key for tracking
                        consumption_key = f"{schedule_theme}_{schedule_duration}_{ad_type}"
                        
                        pending_item = {
                            'theme': schedule_theme,
                            'duration': schedule_duration,
                            'ad_type': ad_type,
                            'original_count': planned_count,
                            'remaining_count': missing_count,  # Only missing items
                            'aired_count': total_aired_count,
                            'status': status,
                            'consumption_key': consumption_key,
                            'has_mapping': has_mapping,
                            'mapped_theme': mapped_theme
                        }
                        
                        pending_schedule_themes.append(pending_item)
                        print(f"   ➕ Added to pending list - Missing: {missing_count}")
                    else:
                        # Fully fulfilled - don't add to list
                        fulfilled_count += 1
                        print(f"   ✅ FULFILLED - Hidden from pending list ({total_aired_count}/{planned_count})")
                    
                except Exception as e:
                    print(f"❌ Error processing schedule group {idx}: {e}")
                    continue
            
            print(f"\n📊 Schedule Analysis Results:")
            print(f"   🔴 Pending themes: {len(pending_schedule_themes)}")
            print(f"   ✅ Fulfilled themes: {fulfilled_count} (hidden)")
            print(f"   📋 Total themes: {len(grouped_schedule)}")
            
            # Sort pending themes by missing count (highest first)
            pending_schedule_themes.sort(key=lambda x: x['remaining_count'], reverse=True)
            
            return pending_schedule_themes
            
        except Exception as e:
            print(f"❌ Error analyzing schedule with duration: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _calculate_automatic_matches_count(self, schedule_theme, duration):
        """Calculate count of automatic matches for this schedule theme"""
        try:
            # Find LMRB theme mapped to this schedule theme
            lmrb_theme = None
            for mapping in self.theme_mapping:
                if mapping.get('schedule_theme') == schedule_theme:
                    lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme'))
                    break
            
            if not lmrb_theme:
                return 0
            
            # Count automatic matches in matched data
            matched_df = self.get_current_matched_data_df()
            if matched_df.empty:
                return 0
            
            # Find LMRB theme column
            lmrb_theme_col = self._find_column(matched_df, ['Advt_Theme', 'theme', 'Product Details'])
            lmrb_dur_col = self._find_column(matched_df, ['Dur', 'duration'])
            
            if not lmrb_theme_col or not lmrb_dur_col:
                return 0
            
            # Count automatic matches (exclude manual matches)
            automatic_matches = matched_df[
                (matched_df[lmrb_theme_col].astype(str).str.strip() == str(lmrb_theme).strip()) & 
                (matched_df[lmrb_dur_col].astype(str).str.strip() == str(duration).strip()) &
                (~matched_df.get('Match_Type', pd.Series(dtype='object')).isin(['Manual', 'Count-Based Manual Match', 'Dual-Consumption Manual Match']))
            ]
            
            count = len(automatic_matches)
            if count > 0:
                print(f"  📊 Automatic matches for {schedule_theme} → {lmrb_theme}: {count}")
            
            return count
            
        except Exception as e:
            print(f"Error calculating automatic matches for {schedule_theme}: {e}")
            return 0

    def refresh_count_data(self):
        """Refresh count analysis data."""
        try:
            if hasattr(self, 'schedule_count_data') and hasattr(self, 'lmrb_count_data'):
                self.run_count_analysis()
                messagebox.showinfo("Refreshed", "Count data refreshed successfully!")
            else:
                messagebox.showinfo("Info", "Please run count analysis first.")
        except Exception as e:
            messagebox.showerror("Error", f"Error refreshing count data: {str(e)}")

    def export_count_analysis(self):
        """Export count analysis to Excel."""
        try:
            if not hasattr(self, 'schedule_count_data') or not hasattr(self, 'lmrb_count_data'):
                messagebox.showerror("Error", "No count data available. Please run analysis first.")
                return
                
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=f"count_analysis_{self.selected_channel or 'channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if filename:
                with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                    # Schedule counts
                    if self.schedule_count_data:
                        pd.DataFrame(self.schedule_count_data).to_excel(writer, sheet_name='Schedule Counts', index=False)
                    
                    # LMRB counts  
                    if self.lmrb_count_data:
                        # Remove group_data column for export
                        export_lmrb_data = []
                        for item in self.lmrb_count_data:
                            export_item = {k: v for k, v in item.items() if k != 'group_data'}
                            export_lmrb_data.append(export_item)
                        pd.DataFrame(export_lmrb_data).to_excel(writer, sheet_name='LMRB Counts', index=False)
                
                messagebox.showinfo("Success", f"Count analysis exported to {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting count analysis: {str(e)}")


    def _populate_schedule_counts_tree(self, schedule_counts):
        """Populate schedule counts tree - ONLY pending/missed themes with enhanced info"""
        try:
            # Clear existing items
            for item in self.schedule_counts_tree.get_children():
                self.schedule_counts_tree.delete(item)
            
            print(f"🌳 Populating schedule counts tree with {len(schedule_counts)} pending items...")
            
            if not schedule_counts:
                # Insert a congratulatory message if no pending themes
                self.schedule_counts_tree.insert('', 'end', values=(
                    "🎉 All schedule themes fulfilled!", 
                    "", 
                    "", 
                    "0", 
                    "✅ ALL COMPLETE"
                ))
                print("   🎉 All themes fulfilled - showing completion message")
                return
            
            # Add only pending items with enhanced information
            for idx, count_data in enumerate(schedule_counts):
                try:
                    # Enhanced status with mapping info
                    status = count_data['status']
                    if not count_data.get('has_mapping', False):
                        status = f"🔸 NO MAPPING | {status}"
                    
                    # Color coding for different priority levels
                    if count_data['remaining_count'] >= 5:
                        priority_icon = "🔴 HIGH"  # High priority
                    elif count_data['remaining_count'] >= 2:
                        priority_icon = "🟡 MED"   # Medium priority  
                    else:
                        priority_icon = "🟢 LOW"   # Low priority
                    
                    values = (
                        f"{priority_icon} {count_data['theme']}",
                        count_data['duration'],
                        count_data['ad_type'], 
                        count_data['remaining_count'],  # Only show missing count
                        status
                    )
                    
                    item_id = self.schedule_counts_tree.insert('', 'end', values=values)
                    
                    print(f"   [{idx+1}] {count_data['theme']} ({count_data['duration']}s) - Missing: {count_data['remaining_count']}")
                    
                except Exception as e:
                    print(f"❌ Error adding schedule item {idx}: {e}")
                    continue
                        
            print(f"✅ Schedule counts tree populated with {len(schedule_counts)} pending themes")
            
        except Exception as e:
            print(f"❌ Error populating schedule counts tree: {e}")
            import traceback
            traceback.print_exc()

    def _populate_lmrb_counts_tree(self, lmrb_counts):
        """Populate the LMRB counts tree with duration info"""
        try:
            # Clear existing items
            for item in self.lmrb_counts_tree.get_children():
                self.lmrb_counts_tree.delete(item)
            
            # Add new items with duration info
            for count_data in lmrb_counts:
                self.lmrb_counts_tree.insert('', 'end', values=(
                    count_data['theme'],
                    count_data['duration'],
                    count_data['advertiser'],
                    count_data['count'],
                    count_data['sample_program']
                ))
                    
        except Exception as e:
            print(f"Error populating LMRB counts tree: {e}")


    def _populate_count_summary_tree(self, summary_data):
        """Populate the count summary tree."""
        try:
            # Clear existing items
            for item in self.count_summary_tree.get_children():
                self.count_summary_tree.delete(item)
            
            # Add new items
            for summary_item in summary_data:
                self.count_summary_tree.insert('', 'end', values=(
                    summary_item['metric'],
                    summary_item['value']
                ))
                
        except Exception as e:
            print(f"Error populating count summary tree: {e}")


    def on_lmrb_theme_select(self, event):
        """Handle LMRB theme selection."""
        try:
            selected = self.lmrb_counts_tree.selection()
            if selected:
                item = selected[0]
                values = self.lmrb_counts_tree.item(item)['values']
                # Extract actual theme name by removing priority indicators
                raw_theme = values[0]
                # Remove priority indicators like "🔴 HIGH ", "🟡 MED ", "🟢 LOW "
                import re
                theme = re.sub(r'^[🔴🟡🟢]\s*(HIGH|MED|LOW)\s+', '', raw_theme)
                advertiser = values[1]
                count = values[2]
                
                # Find the full data
                if hasattr(self, 'lmrb_count_data'):
                    matching_data = next((item for item in self.lmrb_count_data 
                                        if item['theme'] == theme and item['advertiser'] == advertiser), None)
                    
                    if matching_data:
                        self.selected_lmrb_theme = matching_data
                        self.selected_lmrb_var.set(f"📺 LMRB: {theme} ({advertiser}) - {count} ads")
                        self._check_enable_match_button()
        except Exception as e:
            print(f"Error in LMRB theme selection: {e}")



    def clear_theme_selection(self):
        """Clear theme selections."""
        try:
            self.selected_schedule_var.set("No schedule theme selected")
            self.selected_lmrb_var.set("No LMRB theme selected")
            
            if hasattr(self, 'selected_schedule_theme'):
                delattr(self, 'selected_schedule_theme')
            if hasattr(self, 'selected_lmrb_theme'):
                delattr(self, 'selected_lmrb_theme')
            
            if hasattr(self, 'match_themes_button'):
                self.match_themes_button.config(state=tk.DISABLED)
            
            # Clear tree selections
            if hasattr(self, 'schedule_counts_tree'):
                self.schedule_counts_tree.selection_remove(self.schedule_counts_tree.selection())
            if hasattr(self, 'lmrb_counts_tree'):
                self.lmrb_counts_tree.selection_remove(self.lmrb_counts_tree.selection())
        except Exception as e:
            print(f"Error clearing theme selection: {e}")


    def on_schedule_theme_select(self, event):
        """Handle schedule theme selection - only pending themes"""
        try:
            selected = self.schedule_counts_tree.selection()
            if selected:
                item = selected[0]
                values = self.schedule_counts_tree.item(item)['values']
                
                # Check if it's the "all fulfilled" message
                if "All schedule themes fulfilled" in str(values[0]):
                    self.selected_schedule_var.set("🎉 All schedules are fulfilled!")
                    return
                    
                # Extract actual theme name by removing priority indicators
                raw_theme = values[0]
                # Remove priority indicators like "🔴 HIGH ", "🟡 MED ", "🟢 LOW "
                import re
                theme = re.sub(r'^[🔴🟡🟢]\s*(HIGH|MED|LOW)\s+', '', raw_theme)
                duration = values[1]
                ad_type = values[2]
                remaining_count = values[3]
                
                # Find the full data
                if hasattr(self, 'schedule_count_data'):
                    matching_data = next((item for item in self.schedule_count_data 
                                        if item['theme'] == theme and 
                                        str(item['duration']) == str(duration) and
                                        item['ad_type'] == ad_type), None)
                    
                    if matching_data:
                        self.selected_schedule_theme = matching_data
                        self.selected_schedule_var.set(f"📅 Pending: {theme} ({duration}s) - Need: {remaining_count}")
                        self._check_enable_match_button()
        except Exception as e:
            print(f"Error in schedule theme selection: {e}")

    def on_lmrb_theme_select(self, event):
        """Handle LMRB theme selection with duration info"""
        try:
            selected = self.lmrb_counts_tree.selection()
            if selected:
                item = selected[0]
                values = self.lmrb_counts_tree.item(item)['values']
                # Extract actual theme name by removing priority indicators
                raw_theme = values[0]
                # Remove priority indicators like "🔴 HIGH ", "🟡 MED ", "🟢 LOW "
                import re
                theme = re.sub(r'^[🔴🟡🟢]\s*(HIGH|MED|LOW)\s+', '', raw_theme)
                duration = values[1]
                advertiser = values[2]
                count = values[3]
                
                # Find the full data
                if hasattr(self, 'lmrb_count_data'):
                    matching_data = next((item for item in self.lmrb_count_data 
                                        if item['theme'] == theme and 
                                        str(item['duration']) == str(duration) and
                                        item['advertiser'] == advertiser), None)
                    
                    if matching_data:
                        self.selected_lmrb_theme = matching_data
                        self.selected_lmrb_var.set(f"📺 LMRB: {theme} ({duration}s) - Available: {count}")
                        self._check_enable_match_button()
        except Exception as e:
            print(f"Error in LMRB theme selection: {e}")


    def _check_enable_match_button(self):
        """Enable match button if both themes are selected."""
        if hasattr(self, 'selected_schedule_theme') and hasattr(self, 'selected_lmrb_theme'):
            self.match_themes_button.config(state=tk.NORMAL)
        else:
            self.match_themes_button.config(state=tk.DISABLED)

    def match_selected_themes(self):
        """Match selected schedule and LMRB themes and add to matched data."""
        try:
            if not hasattr(self, 'selected_schedule_theme') or not hasattr(self, 'selected_lmrb_theme'):
                messagebox.showerror("Error", "Please select both schedule and LMRB themes")
                return

            schedule_theme = self.selected_schedule_theme['theme']
            schedule_count = self.selected_schedule_theme['count']
            lmrb_theme = self.selected_lmrb_theme['theme']
            lmrb_count = self.selected_lmrb_theme['count']
            lmrb_group_data = self.selected_lmrb_theme['group_data']

            # Confirm the match
            result = messagebox.askyesno("Confirm Match", 
                f"Match these themes?\n\n"
                f"📅 Schedule: {schedule_theme} ({schedule_count} ads)\n"
                f"📺 LMRB: {lmrb_theme} ({lmrb_count} ads)\n\n"
                f"Difference: {lmrb_count - schedule_count} ads\n"
                f"This will add {lmrb_count} records to matched data.")

            if result:
                # NEW: Track the manual match for schedule counting
                schedule_match_record = {
                    'schedule_theme': schedule_theme,
                    'schedule_count_used': min(schedule_count, lmrb_count),  # How many schedule items were "consumed"
                    'lmrb_theme': lmrb_theme,
                    'lmrb_count': lmrb_count,
                    'match_timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                
                # Add to tracking list
                if not hasattr(self, 'manually_matched_schedule_items'):
                    self.manually_matched_schedule_items = []
                self.manually_matched_schedule_items.append(schedule_match_record)
                
                print(f"Tracked manual schedule match: {schedule_theme} used {min(schedule_count, lmrb_count)} items")
                
                # Add all LMRB records from this group to matched data
                added_count = 0
                for _, record in lmrb_group_data.iterrows():
                    matched_record = record.to_dict()
                    matched_record['Match_Type'] = 'Count-Based Manual Match'
                    matched_record['Match_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    matched_record['Schedule_Theme_Matched'] = schedule_theme
                    matched_record['Count_Difference'] = lmrb_count - schedule_count
                    
                    # Add to matched data
                    if isinstance(self.matched_data, list):
                        self.matched_data.append(matched_record)
                    else:
                        self.matched_data = pd.concat([self.matched_data, pd.DataFrame([matched_record])], ignore_index=True)
                    
                    added_count += 1

                # Remove from unmatched LMRB data
                indices_to_remove = lmrb_group_data.index.tolist()
                self.unmatched_lmrb_data = self.unmatched_lmrb_data.drop(indices_to_remove).reset_index(drop=True)

                # Refresh all displays
                self.sync_data_across_tabs()
                self.run_count_analysis()  # Refresh the analysis
                self.clear_theme_selection()

                # Update summary report
                self.sync_summary_report_data()

                messagebox.showinfo("Success", 
                    f"✅ Successfully matched!\n\n"
                    f"Added {added_count} records to matched data\n"
                    f"All tabs have been synchronized.")

        except Exception as e:
            messagebox.showerror("Error", f"Error matching themes: {str(e)}")


    def run_count_analysis(self):
        """Run simple count-based analysis."""
        try:
            if self.schedule_data is None:
                messagebox.showerror("Error", "No schedule data available")
                return

            print("🔍 Running Count Analysis...")

            # Analyze schedule counts
            schedule_counts = self._analyze_schedule_counts()
            self._populate_schedule_counts_tree(schedule_counts)

            # Analyze unmatched LMRB counts
            lmrb_counts = self._analyze_unmatched_lmrb_counts()
            self._populate_lmrb_counts_tree(lmrb_counts)

            # Create summary
            summary_data = self._create_count_summary(schedule_counts, lmrb_counts)
            self._populate_count_summary_tree(summary_data)

            # Store for later use
            self.schedule_count_data = schedule_counts
            self.lmrb_count_data = lmrb_counts

            messagebox.showinfo("Analysis Complete", 
                f"✅ Count Analysis Complete!\n\n"
                f"📅 Schedule themes: {len(schedule_counts)}\n"
                f"📺 Unmatched LMRB themes: {len(lmrb_counts)}\n\n"
                f"💡 Select themes from each side to match them manually.")

        except Exception as e:
            messagebox.showerror("Error", f"Error in count analysis: {str(e)}")

    def _find_column(self, df, possible_names):
        """Find column by possible names (case insensitive)"""
        if df is None or df.empty:
            return None
            
        for name in possible_names:
            # Exact match first
            if name in df.columns:
                return name
            # Case insensitive search
            for col in df.columns:
                if name.lower() in col.lower():
                    return col
        return None

    def _is_theme_already_matched(self, schedule_theme):
        """Check if schedule theme is already in theme mappings."""
        return any(mapping.get('schedule_theme') == schedule_theme for mapping in self.theme_mapping)

    def _analyze_schedule_counts(self):
        """Analyze schedule data and group by theme."""
        try:
            schedule_counts = []
            
            # Find columns
            theme_col = self._find_column(self.schedule_data, ['theme', 'aadvt', 'advt_theme'])
            adtype_col = self._find_column(self.schedule_data, ['advertisement_type', 'ad_type'])
            
            if not theme_col:
                return []

            # Group by theme and ad type
            if adtype_col:
                grouped = self.schedule_data.groupby([theme_col, adtype_col]).size().reset_index(name='count')
            else:
                grouped = self.schedule_data.groupby([theme_col]).size().reset_index(name='count')
                grouped[adtype_col] = 'Unknown'

            for _, row in grouped.iterrows():
                theme = row[theme_col]
                ad_type = row[adtype_col] if adtype_col else 'Unknown'
                count = row['count']
                
                # Check if this theme is already matched
                is_matched = self._is_theme_already_matched(theme)
                status = "✅ Matched" if is_matched else "⏳ Pending"
                
                schedule_counts.append({
                    'theme': theme,
                    'ad_type': ad_type,
                    'count': count,
                    'status': status
                })

            return schedule_counts

        except Exception as e:
            print(f"Error analyzing schedule counts: {e}")
            return []

    def _analyze_unmatched_lmrb_counts(self):
        """Analyze unmatched LMRB data and group by theme."""
        try:
            if self.unmatched_lmrb_data is None or self.unmatched_lmrb_data.empty:
                return []

            lmrb_counts = []
            
            # Find columns
            theme_col = self._find_column(self.unmatched_lmrb_data, ['Advt_Theme', 'theme', 'aadvt'])
            advertiser_col = self._find_column(self.unmatched_lmrb_data, ['advertiser', 'client'])
            program_col = self._find_column(self.unmatched_lmrb_data, ['Program', 'program'])
            
            if not theme_col:
                return []

            # Group by theme (and advertiser if available)
            if advertiser_col:
                grouped = self.unmatched_lmrb_data.groupby([theme_col, advertiser_col])
            else:
                grouped = self.unmatched_lmrb_data.groupby([theme_col])

            for group_key, group_data in grouped:
                if isinstance(group_key, tuple):
                    theme, advertiser = group_key
                else:
                    theme = group_key
                    advertiser = "Unknown"

                count = len(group_data)
                sample_program = group_data[program_col].iloc[0] if program_col else "Unknown"

                lmrb_counts.append({
                    'theme': theme,
                    'advertiser': advertiser,
                    'count': count,
                    'sample_program': sample_program,
                    'group_data': group_data  # Store for matching
                })

            # Sort by count (highest first)
            lmrb_counts.sort(key=lambda x: x['count'], reverse=True)
            return lmrb_counts

        except Exception as e:
            print(f"Error analyzing LMRB counts: {e}")
            return []


    def _create_count_summary(self, schedule_counts, lmrb_counts):
        """Create summary statistics."""
        total_schedule_themes = len(schedule_counts)
        total_lmrb_themes = len(lmrb_counts)
        total_schedule_count = sum(item['count'] for item in schedule_counts)
        total_lmrb_count = sum(item['count'] for item in lmrb_counts)

        matched_themes = len([item for item in schedule_counts if item['status'] == "✅ Matched"])
        pending_themes = total_schedule_themes - matched_themes

        return [
            {'metric': 'Schedule Themes', 'value': total_schedule_themes},
            {'metric': 'LMRB Themes', 'value': total_lmrb_themes},
            {'metric': 'Total Scheduled Ads', 'value': total_schedule_count},
            {'metric': 'Total Unmatched Ads', 'value': total_lmrb_count},
            {'metric': 'Matched Themes', 'value': matched_themes},
            {'metric': 'Pending Themes', 'value': pending_themes}
        ]

    def get_current_matched_data_df(self):
        """FIXED: Get the current matched data as a DataFrame"""
        try:
            # FIXED: Proper DataFrame checking
            if self.matched_data is None:
                return pd.DataFrame()
            
            if isinstance(self.matched_data, list):
                if len(self.matched_data) == 0:  # FIXED: Use len() instead of boolean
                    return pd.DataFrame()
                df = pd.DataFrame(self.matched_data)
            else:
                if self.matched_data.empty:  # FIXED: Use .empty for DataFrame
                    return pd.DataFrame()
                df = self.matched_data.copy()
            
            print(f"✅ Current matched data: {len(df)} records")
            if 'Match_Type' in df.columns:
                match_counts = df['Match_Type'].value_counts()
                print(f"📊 Match types: {match_counts.to_dict()}")
            
            return df
            
        except Exception as e:
            print(f"❌ Error getting matched data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def sync_data_across_tabs(self):
        """FIXED: Immediate and efficient data synchronization across all tabs."""
        try:
            print("🔄 IMMEDIATE SYNC: Starting data synchronization...")
            sync_start = time.time()
            
            # FIXED: Update statistics first (fastest operation)
            self.update_statistics_immediate()
            
            # FIXED: Update results tab immediately if initialized
            if hasattr(self, 'matched_tree'):
                self.update_results_tab()
            
            # FIXED: Update schedule matching displays if initialized
            if hasattr(self, 'schedule_counts_tree'):
                # Only refresh if we have data to avoid unnecessary processing
                if hasattr(self, 'schedule_count_data') and self.schedule_count_data:
                    self._populate_schedule_counts_tree(self.schedule_count_data)
                if hasattr(self, 'lmrb_count_data') and self.lmrb_count_data:
                    self._populate_lmrb_counts_tree(self.lmrb_count_data)
            
            # FIXED: Update summary report if initialized and has schedule data
            if (hasattr(self, 'commercial_tree') and 
                hasattr(self, 'schedule_data') and self.schedule_data is not None):
                self.sync_summary_report_data()
            
            sync_time = time.time() - sync_start
            print(f"✅ IMMEDIATE SYNC COMPLETE: {sync_time:.3f}s")
            
            # Update status bar
            self.status_var.set(f"Data synchronized ({sync_time:.2f}s)")
            
        except Exception as e:
            print(f"❌ Sync error: {e}")
            import traceback
            traceback.print_exc()

    def update_statistics_immediate(self):
        """FIXED: Immediate statistics update without lag."""
        try:
            # FIXED: Fast counting
            total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0
            total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0
            
            # FIXED: Handle different matched_data formats efficiently
            if isinstance(self.matched_data, list):
                total_matches = len([m for m in self.matched_data if m]) if self.matched_data else 0
            elif isinstance(self.matched_data, pd.DataFrame):
                total_matches = len(self.matched_data) if not self.matched_data.empty else 0
            else:
                total_matches = 0
            
            total_unmatched_tc = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            total_unmatched_lmrb = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            match_rate = (total_matches / total_tc * 100) if total_tc > 0 else 0

            # FIXED: Update all variables immediately
            self.total_lmrb_var.set(f"Total LMRB Records: {total_lmrb:,}")
            self.matched_count_var.set(f"Matched Records: {total_matches:,}")
            self.unmatched_tc_var.set(f"Unmatched TC Records: {total_unmatched_tc:,}")
            self.unmatched_lmrb_var.set(f"Unmatched LMRB Records: {total_unmatched_lmrb:,}")
            self.match_rate_var.set(f"Match Rate: {match_rate:.1f}%")
            
            # FIXED: Force immediate UI update
            self.root.update_idletasks()
            
        except Exception as e:
            print(f"❌ Statistics update error: {e}")

    def refresh_schedule_data(self):
        """Refresh schedule matching data to include latest matches."""
        try:
            # Synchronize all data
            self.sync_data_across_tabs()
            
            # Clear schedule results
            self.schedule_summary_results = []
            
            # Clear tree
            for item in self.schedule_summary_tree.get_children():
                self.schedule_summary_tree.delete(item)
            
            # Show current data status
            current_matches = len(self.matched_data) if self.matched_data else 0
            messagebox.showinfo("Data Refreshed", 
                            f"Schedule data refreshed!\n"
                            f"Current matched records: {current_matches}\n"
                            f"Please run 'Run Schedule Matching' to see updated results.")
            
        except Exception as e:
            messagebox.showerror("Refresh Error", f"Error refreshing data: {str(e)}")

    def run_schedule_matching(self):
        """Run schedule matching using the most current matched data including manual matches."""
        try:
            if self.schedule_data is None or self.schedule_data.empty:
                messagebox.showerror("Error", "No schedule data available")
                return
                
            # CRITICAL FIX: Ensure we're using the most current matched data
            if not self.matched_data:
                messagebox.showerror("Error", "No matched LMRB data available. Please run TC-LMRB matching first.")
                return
            
            print(f"\n=== SCHEDULE MATCHING DEBUG ===")
            print(f"Current matched data count: {len(self.matched_data)}")
            print(f"Matched data type: {type(self.matched_data)}")
            
            # Convert matched_data to DataFrame ensuring we get the latest data
            if isinstance(self.matched_data, list):
                if not self.matched_data:
                    messagebox.showerror("Error", "Matched data list is empty")
                    return
                matched_lmrb_df = pd.DataFrame(self.matched_data).copy()
                print(f"Converted list to DataFrame: {len(matched_lmrb_df)} rows")
            else:
                matched_lmrb_df = self.matched_data.copy()
                print(f"Using existing DataFrame: {len(matched_lmrb_df)} rows")
            
            # Debug: Show sample of matched data
            if len(matched_lmrb_df) > 0:
                print("Sample of matched data columns:", matched_lmrb_df.columns.tolist()[:10])
                if 'Match_Type' in matched_lmrb_df.columns:
                    match_types = matched_lmrb_df['Match_Type'].value_counts()
                    print("Match types in current data:", match_types.to_dict())
            
            # Get schedule columns
            schedule_theme_col = next((col for col in self.schedule_data.columns 
                                    if 'theme' in col.lower() or 'aadvt' in col.lower()), None)
            schedule_dur_col = next((col for col in self.schedule_data.columns 
                                if 'dur' in col.lower()), None)
            schedule_adtype_col = next((col for col in self.schedule_data.columns 
                                    if 'advertisement_type' in col.lower() or 'ad_type' in col.lower()), None)
            schedule_program_col = next((col for col in self.schedule_data.columns 
                                    if 'program' in col.lower()), None)
            
            if not schedule_theme_col or not schedule_dur_col:
                messagebox.showerror("Error", "Could not find theme or duration columns in schedule data")
                return
            
            # Get LMRB columns
            lmrb_theme_col = next((col for col in matched_lmrb_df.columns 
                                if 'theme' in col.lower() or 'aadvt' in col.lower()), None)
            lmrb_dur_col = next((col for col in matched_lmrb_df.columns 
                                if 'dur' in col.lower()), None)
            
            if not lmrb_theme_col or not lmrb_dur_col:
                messagebox.showerror("Error", "Could not find theme or duration columns in matched LMRB data")
                return
            
            # Create theme mapping dictionary for quick lookup
            schedule_to_lmrb_mapping = {}
            mapping_types = {}  # Store whether each mapping is sponsorship or commercial
            
            for mapping in self.theme_mapping:
                schedule_theme = mapping.get('schedule_theme')
                schedule_dur = mapping.get('schedule_duration', '')
                lmrb_theme = mapping.get('media_watch_theme')
                lmrb_dur = mapping.get('lmrb_duration', '')
                mapping_type = mapping.get('mapping_type', 'COMMERCIAL BENEFITS')
                
                if schedule_theme and lmrb_theme:
                    key = (schedule_theme, str(schedule_dur))
                    schedule_to_lmrb_mapping[key] = (lmrb_theme, str(lmrb_dur))
                    mapping_types[key] = mapping_type
            
            # Separate schedule data by advertisement type
            schedule_sponsorship = pd.DataFrame()
            schedule_commercial = pd.DataFrame()
            
            if schedule_adtype_col and schedule_adtype_col in self.schedule_data.columns:
                # Split by Advertisement_Type column
                schedule_sponsorship = self.schedule_data[
                    self.schedule_data[schedule_adtype_col].str.lower().str.contains('sponsor', na=False)
                ].copy()
                schedule_commercial = self.schedule_data[
                    self.schedule_data[schedule_adtype_col].str.lower().str.contains('commercial', na=False)
                ].copy()
                
                print(f"Found Advertisement_Type column: {schedule_adtype_col}")
                print(f"Sponsorship entries: {len(schedule_sponsorship)}")
                print(f"Commercial entries: {len(schedule_commercial)}")
            else:
                # Fallback: use mapping types to separate
                print("No Advertisement_Type column found, using mapping types")
                sponsorship_themes = [k[0] for k, v in mapping_types.items() if 'SPONSORSHIP' in v]
                commercial_themes = [k[0] for k, v in mapping_types.items() if 'COMMERCIAL' in v]
                
                if sponsorship_themes:
                    schedule_sponsorship = self.schedule_data[
                        self.schedule_data[schedule_theme_col].isin(sponsorship_themes)
                    ].copy()
                if commercial_themes:
                    schedule_commercial = self.schedule_data[
                        self.schedule_data[schedule_theme_col].isin(commercial_themes)
                    ].copy()
            
            summary_results = []
            
            # ==========================================
            # HANDLE SPONSORSHIP MATCHING (Count-based only)
            # ==========================================
            if not schedule_sponsorship.empty:
                print("\nProcessing Sponsorship matching...")
                
                # Group sponsorships by theme and duration only (ignore programs)
                sponsorship_grouped = schedule_sponsorship.groupby([schedule_theme_col, schedule_dur_col]).agg({
                    schedule_program_col: lambda x: ', '.join(x.unique()) if schedule_program_col else 'Multiple Programs',
                    schedule_theme_col: 'count'  # This gives us the count
                }).rename(columns={schedule_theme_col: 'planned_count'}).reset_index()
                
                for _, spon_row in sponsorship_grouped.iterrows():
                    schedule_theme = spon_row[schedule_theme_col]
                    schedule_duration = str(spon_row[schedule_dur_col])
                    planned_count = spon_row['planned_count']
                    programs_involved = spon_row[schedule_program_col] if schedule_program_col else 'Multiple Programs'
                    
                    # Find corresponding LMRB data using mapping
                    aired_count = 0
                    mapping_key = (schedule_theme, schedule_duration)
                    
                    if mapping_key in schedule_to_lmrb_mapping:
                        lmrb_theme, lmrb_duration = schedule_to_lmrb_mapping[mapping_key]
                        
                        # For sponsorships: Count ALL matching entries regardless of program
                        matching_aired = matched_lmrb_df[
                            (matched_lmrb_df[lmrb_theme_col] == lmrb_theme) & 
                            (matched_lmrb_df[lmrb_dur_col].astype(str) == lmrb_duration)
                        ]
                        aired_count = len(matching_aired)
                        
                        print(f"Sponsorship: {schedule_theme} ({schedule_duration}s)")
                        print(f"  Planned: {planned_count}, Aired: {aired_count}")
                        print(f"  Programs involved: {programs_involved}")
                    
                    # Calculate differences
                    extra = max(0, aired_count - planned_count)
                    missed = max(0, planned_count - aired_count)
                    
                    # Calculate 30sec duration
                    try:
                        duration_numeric = float(schedule_duration)
                        duration_30sec = planned_count * duration_numeric / 30
                    except:
                        duration_30sec = 0
                    
                    summary_results.append({
                        'advertisement_type': 'SPONSORSHIP',
                        'schedule_theme': schedule_theme,
                        'duration': schedule_duration,
                        'planned_count': planned_count,
                        'aired_count': aired_count,
                        'extra': extra,
                        'missed': missed,
                        'duration_30sec': f"{duration_30sec:.2f}",
                        'programs_involved': programs_involved,
                        'matching_logic': 'Count-based (cross-program)'
                    })
            
            # ==========================================
            # HANDLE COMMERCIAL MATCHING (Count-based)
            # ==========================================
            if not schedule_commercial.empty:
                print("\nProcessing Commercial matching...")
                
                # Group commercials by theme and duration
                commercial_grouped = schedule_commercial.groupby([schedule_theme_col, schedule_dur_col]).agg({
                    schedule_program_col: lambda x: ', '.join(x.unique()) if schedule_program_col else 'Multiple Programs',
                    schedule_theme_col: 'count'
                }).rename(columns={schedule_theme_col: 'planned_count'}).reset_index()
                
                for _, comm_row in commercial_grouped.iterrows():
                    schedule_theme = comm_row[schedule_theme_col]
                    schedule_duration = str(comm_row[schedule_dur_col])
                    planned_count = comm_row['planned_count']
                    programs_involved = comm_row[schedule_program_col] if schedule_program_col else 'Multiple Programs'
                    
                    # Find corresponding LMRB data using mapping
                    aired_count = 0
                    mapping_key = (schedule_theme, schedule_duration)
                    
                    if mapping_key in schedule_to_lmrb_mapping:
                        lmrb_theme, lmrb_duration = schedule_to_lmrb_mapping[mapping_key]
                        
                        # For commercials: Also use count-based approach
                        matching_aired = matched_lmrb_df[
                            (matched_lmrb_df[lmrb_theme_col] == lmrb_theme) & 
                            (matched_lmrb_df[lmrb_dur_col].astype(str) == lmrb_duration)
                        ]
                        aired_count = len(matching_aired)
                        
                        print(f"Commercial: {schedule_theme} ({schedule_duration}s)")
                        print(f"  Planned: {planned_count}, Aired: {aired_count}")
                        print(f"  Programs involved: {programs_involved}")
                    
                    # Calculate differences
                    extra = max(0, aired_count - planned_count)
                    missed = max(0, planned_count - aired_count)
                    
                    # Calculate 30sec duration
                    try:
                        duration_numeric = float(schedule_duration)
                        duration_30sec = planned_count * duration_numeric / 30
                    except:
                        duration_30sec = 0
                    
                    summary_results.append({
                        'advertisement_type': 'COMMERCIAL',
                        'schedule_theme': schedule_theme,
                        'duration': schedule_duration,
                        'planned_count': planned_count,
                        'aired_count': aired_count,
                        'extra': extra,
                        'missed': missed,
                        'duration_30sec': f"{duration_30sec:.2f}",
                        'programs_involved': programs_involved,
                        'matching_logic': 'Count-based (cross-program)'
                    })
            
            # Store results
            self.schedule_summary_results = summary_results
            
            # Update treeview with improved display
            self.update_schedule_summary_tree_improved()
            
            # Populate trailer selection tree with unmatched LMRB data
            self.populate_trailer_selection_tree()
            
            # Enable schedule matching tab
            self.notebook.tab(5, state="normal")
            
            # Count manual vs automatic matches
            manual_matches = 0
            automatic_matches = 0
            if isinstance(self.matched_data, list):
                for match in self.matched_data:
                    if match.get('Match_Type') == 'Manual':
                        manual_matches += 1
                    else:
                        automatic_matches += 1
            
            # Show detailed results
            total_sponsorships = len([r for r in summary_results if r['advertisement_type'] == 'SPONSORSHIP'])
            total_commercials = len([r for r in summary_results if r['advertisement_type'] == 'COMMERCIAL'])
            
            completion_message = (
                f"Schedule matching completed using LIVE data!\n\n"
                f"📊 Data Source Summary:\n"
                f"• Total matched records used: {len(matched_lmrb_df)}\n"
                f"• Automatic matches: {automatic_matches}\n"
                f"• Manual matches: {manual_matches}\n\n"
                f"📋 Schedule Summary:\n"
                f"• Sponsorship entries: {total_sponsorships}\n"
                f"• Commercial entries: {total_commercials}\n"
                f"• Total combinations: {len(summary_results)}\n\n"
                f"💡 Sponsorship Logic:\n"
                f"• Counts total tags across all programs\n"
                f"• Ignores specific program distribution\n\n"
                f"✅ All tabs are now synchronized!"
            )
            
            messagebox.showinfo("Schedule Matching Complete", completion_message)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error in schedule matching: {str(e)}")
            import traceback
            traceback.print_exc()

    def update_schedule_summary_tree_improved(self):
        """Update the schedule summary treeview with improved columns including advertisement type."""
        try:
            # Clear existing data
            for item in self.schedule_summary_tree.get_children():
                self.schedule_summary_tree.delete(item)

            # Update treeview columns to include advertisement type
            current_columns = self.schedule_summary_tree["columns"]
            if "advertisement_type" not in current_columns:
                new_columns = ("advertisement_type", "schedule_theme", "duration", "planned_count", 
                            "aired_count", "extra", "missed", "duration_30sec", "programs_involved")
                self.schedule_summary_tree["columns"] = new_columns
                
                # Configure new column headings
                self.schedule_summary_tree.heading("advertisement_type", text="Ad Type")
                self.schedule_summary_tree.heading("schedule_theme", text="Schedule Theme")
                self.schedule_summary_tree.heading("duration", text="Duration")
                self.schedule_summary_tree.heading("planned_count", text="Planned")
                self.schedule_summary_tree.heading("aired_count", text="Aired")
                self.schedule_summary_tree.heading("extra", text="Extra")
                self.schedule_summary_tree.heading("missed", text="Missed")
                self.schedule_summary_tree.heading("duration_30sec", text="30sec Duration")
                self.schedule_summary_tree.heading("programs_involved", text="Programs")
                
                # Configure column widths
                self.schedule_summary_tree.column("advertisement_type", width=100)
                self.schedule_summary_tree.column("schedule_theme", width=200)
                self.schedule_summary_tree.column("duration", width=80)
                self.schedule_summary_tree.column("planned_count", width=80)
                self.schedule_summary_tree.column("aired_count", width=80)
                self.schedule_summary_tree.column("extra", width=60)
                self.schedule_summary_tree.column("missed", width=60)
                self.schedule_summary_tree.column("duration_30sec", width=100)
                self.schedule_summary_tree.column("programs_involved", width=200)

            if hasattr(self, 'schedule_summary_results'):
                for result in self.schedule_summary_results:
                    # Color code based on status and advertisement type
                    tags = []
                    ad_type = result.get('advertisement_type', 'UNKNOWN')
                    
                    if result['missed'] > 0:
                        tags.append('missed')
                    elif result['extra'] > 0:
                        tags.append('extra')
                    else:
                        tags.append('matched')
                    
                    # Add advertisement type tag
                    if ad_type == 'SPONSORSHIP':
                        tags.append('sponsorship')
                    elif ad_type == 'COMMERCIAL':
                        tags.append('commercial')
                    
                    self.schedule_summary_tree.insert(
                        "", "end",
                        values=(
                            result.get('advertisement_type', 'UNKNOWN'),
                            result['schedule_theme'],
                            result['duration'],
                            result['planned_count'],
                            result['aired_count'],
                            result['extra'],
                            result['missed'],
                            result['duration_30sec'],
                            result.get('programs_involved', 'N/A')
                        ),
                        tags=tags
                    )
                
                # Configure tags for color coding
                self.schedule_summary_tree.tag_configure('missed', background='#ffcccc')
                self.schedule_summary_tree.tag_configure('extra', background='#ccffcc')
                self.schedule_summary_tree.tag_configure('matched', background='#f0f0f0')
                self.schedule_summary_tree.tag_configure('sponsorship', foreground='#0066cc')
                self.schedule_summary_tree.tag_configure('commercial', foreground='#cc6600')
        
        except Exception as e:
            print(f"Error updating schedule summary tree: {e}")
            import traceback
            traceback.print_exc()

    def populate_trailer_selection_tree(self):
        """Populate the trailer selection tree with unmatched LMRB data."""
        # Clear existing data
        for item in self.trailer_selection_tree.get_children():
            self.trailer_selection_tree.delete(item)

        if self.unmatched_lmrb_data is not None and not self.unmatched_lmrb_data.empty:
            # Configure columns
            columns = list(self.unmatched_lmrb_data.columns)
            self.trailer_selection_tree["columns"] = columns
            
            # Configure column headings
            for col in columns:
                self.trailer_selection_tree.column(col, width=100)
                self.trailer_selection_tree.heading(col, text=col)
            
            # Add data rows
            for i, row in self.unmatched_lmrb_data.iterrows():
                values = [str(row[col]) if not pd.isna(row[col]) else "" for col in columns]
                self.trailer_selection_tree.insert("", "end", text=str(i), values=values)

    def add_selected_trailers(self):
        """Add selected trailer records to the final matched results with synchronization."""
        try:
            selected_items = self.trailer_selection_tree.selection()
            if not selected_items:
                messagebox.showwarning("Warning", "Please select trailer records to add")
                return
            
            added_count = 0
            for item in selected_items:
                # Get the row index
                row_index = int(self.trailer_selection_tree.item(item)['text'])
                
                # Get the row data
                trailer_row = self.unmatched_lmrb_data.iloc[row_index].to_dict()
                trailer_row['Match_Type'] = 'Trailer/Sponsorship'
                trailer_row['Match_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Add to matched data
                if isinstance(self.matched_data, list):
                    self.matched_data.append(trailer_row)
                else:
                    self.matched_data = pd.concat([self.matched_data, pd.DataFrame([trailer_row])], ignore_index=True)
                
                added_count += 1
            
            # Remove selected items from unmatched data
            selected_indices = [int(self.trailer_selection_tree.item(item)['text']) for item in selected_items]
            self.unmatched_lmrb_data = self.unmatched_lmrb_data.drop(selected_indices).reset_index(drop=True)
            
            # CRITICAL: Synchronize all tabs after adding trailers
            self.refresh_all_dependent_tabs()
            self.sync_summary_report_data()
            
            # Update displays
            self.populate_trailer_selection_tree()
            
            messagebox.showinfo("Success", f"Added {added_count} trailer/sponsorship records to final results. All tabs synchronized.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error adding trailers: {str(e)}")

    def export_schedule_summary(self):
        """Export schedule summary to Excel file."""
        try:
            if not hasattr(self, 'schedule_summary_results') or not self.schedule_summary_results:
                messagebox.showerror("Error", "No schedule summary data available")
                return
            
            default_name = f"schedule_summary_{self.selected_channel or 'channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                # Create DataFrame from summary results
                summary_df = pd.DataFrame(self.schedule_summary_results)
                
                # Export to Excel
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    summary_df.to_excel(writer, sheet_name='Schedule Summary', index=False)
                    
                    # Add formatting
                    worksheet = writer.sheets['Schedule Summary']
                    
                    # Color code rows based on status
                    from openpyxl.styles import PatternFill
                    red_fill = PatternFill(start_color='ffcccc', end_color='ffcccc', fill_type='solid')
                    green_fill = PatternFill(start_color='ccffcc', end_color='ccffcc', fill_type='solid')
                    
                    for row_num, result in enumerate(self.schedule_summary_results, start=2):
                        if result['missed'] > 0:
                            for col in range(1, 8):
                                worksheet.cell(row=row_num, column=col).fill = red_fill
                        elif result['extra'] > 0:
                            for col in range(1, 8):
                                worksheet.cell(row=row_num, column=col).fill = green_fill
                
                messagebox.showinfo("Success", f"Schedule summary exported to {save_path}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting schedule summary: {str(e)}")

    def setup_matching_tab(self):
        # Title
        title_label = ttk.Label(self.matching_tab, text="Data Matching", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky="w")
        
        # Options Frame
        options_frame = ttk.LabelFrame(self.matching_tab, text="Matching Options", padding=10)
        options_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=10)
        
        # Ignore Date Option
        
        ignore_date_check = ttk.Checkbutton(options_frame, text="Ignore Date (Match Only by Theme, Program, and Duration)", 
                                            variable=self.ignore_date_var)
        ignore_date_check.grid(row=0, column=0, sticky="w", padx=5, pady=5)
        
        # Time Tolerance Option
        time_label = ttk.Label(options_frame, text="Time Matching Tolerance (seconds):")
        time_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)
        
        
        time_spinbox = ttk.Spinbox(options_frame, from_=5, to=60, increment=5, textvariable=self.time_tolerance_var, width=5)
        time_spinbox.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        # Run Matching Button
        run_button = ttk.Button(self.matching_tab, text="Run Matching Process", command=self.run_matching)
        run_button.grid(row=2, column=0, columnspan=2, pady=20)
        
        # Progress Bar

        self.progress_bar = ttk.Progressbar(
            self.matching_tab,
            variable=self.progress_var,
            maximum=100,
            style="Blue.Horizontal.TProgressbar"  # Use your custom style
        )
        self.progress_bar.grid(row=3, column=0, columnspan=2, sticky="ew", pady=5)
        
        # Status Label
        self.matching_status_var = tk.StringVar(value="Ready to match")
        status_label = ttk.Label(self.matching_tab, textvariable=self.matching_status_var)
        status_label.grid(row=4, column=0, columnspan=2, pady=5)
        
        # Continue Button
        continue_button = ttk.Button(self.matching_tab, text="Continue to Results", command=self.continue_to_results)
        continue_button.grid(row=5, column=1, pady=10, sticky="e")
        
        # Make the grid expandable
        self.matching_tab.grid_columnconfigure(0, weight=1)
        self.matching_tab.grid_columnconfigure(1, weight=1)

    def setup_results_tab(self):
        # Title
        title_label = ttk.Label(self.results_tab, text="Matching Results", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky="w")

        # Results Summary Frame
        summary_frame = ttk.LabelFrame(self.results_tab, text="Summary", padding=10)
        summary_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=10)

        # Summary Labels
        total_lmrb_label = ttk.Label(summary_frame, textvariable=self.total_lmrb_var)
        total_lmrb_label.grid(row=0, column=0, sticky="w", padx=5, pady=2)

        matched_count_label = ttk.Label(summary_frame, textvariable=self.matched_count_var)
        matched_count_label.grid(row=1, column=0, sticky="w", padx=5, pady=2)

        unmatched_tc_label = ttk.Label(summary_frame, textvariable=self.unmatched_tc_var)
        unmatched_tc_label.grid(row=2, column=0, sticky="w", padx=5, pady=2)

        # ADD: Unmatched LMRB label
        unmatched_lmrb_label = ttk.Label(summary_frame, textvariable=self.unmatched_lmrb_var)
        unmatched_lmrb_label.grid(row=3, column=0, sticky="w", padx=5, pady=2)

        self.match_rate_var = tk.StringVar(value="Match Rate: 0%")
        match_rate_label = ttk.Label(summary_frame, textvariable=self.match_rate_var)
        match_rate_label.grid(row=4, column=0, sticky="w", padx=5, pady=2)  # Updated row position

        # Results Notebook
        results_notebook = ttk.Notebook(self.results_tab)
        results_notebook.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=10)

        # Matched Data Tab
        matched_frame = ttk.Frame(results_notebook)
        results_notebook.add(matched_frame, text="Matched Data")

        # Matched Treeview
        self.matched_tree = ttk.Treeview(matched_frame)
        matched_scrolly = ttk.Scrollbar(matched_frame, orient="vertical", command=self.matched_tree.yview)
        matched_scrollx = ttk.Scrollbar(matched_frame, orient="horizontal", command=self.matched_tree.xview)
        self.matched_tree.configure(yscrollcommand=matched_scrolly.set, xscrollcommand=matched_scrollx.set)

        matched_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.matched_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        matched_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # Unmatched TC Tab
        unmatched_tc_frame = ttk.Frame(results_notebook)
        results_notebook.add(unmatched_tc_frame, text="Unmatched TC Data")

        # Unmatched TC Treeview
        self.unmatched_tree = ttk.Treeview(unmatched_tc_frame)
        unmatched_scrolly = ttk.Scrollbar(unmatched_tc_frame, orient="vertical", command=self.unmatched_tree.yview)
        unmatched_scrollx = ttk.Scrollbar(unmatched_tc_frame, orient="horizontal", command=self.unmatched_tree.xview)
        self.unmatched_tree.configure(yscrollcommand=unmatched_scrolly.set, xscrollcommand=unmatched_scrollx.set)

        unmatched_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.unmatched_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        unmatched_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # ADD NEW TAB: Unmatched LMRB Tab
        unmatched_lmrb_frame = ttk.Frame(results_notebook)
        results_notebook.add(unmatched_lmrb_frame, text="Unmatched LMRB Data")

        # Unmatched LMRB Treeview
        self.unmatched_lmrb_tree = ttk.Treeview(unmatched_lmrb_frame)
        unmatched_lmrb_scrolly = ttk.Scrollbar(unmatched_lmrb_frame, orient="vertical", command=self.unmatched_lmrb_tree.yview)
        unmatched_lmrb_scrollx = ttk.Scrollbar(unmatched_lmrb_frame, orient="horizontal", command=self.unmatched_lmrb_tree.xview)
        self.unmatched_lmrb_tree.configure(yscrollcommand=unmatched_lmrb_scrolly.set, xscrollcommand=unmatched_lmrb_scrollx.set)

        unmatched_lmrb_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.unmatched_lmrb_tree.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        unmatched_lmrb_scrollx.pack(side=tk.BOTTOM, fill=tk.X)

        # Export Buttons Frame - SINGLE ROW
        export_frame = ttk.Frame(self.results_tab)
        export_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=10)

        # Export Matched Data Button
        export_matched_button = ttk.Button(export_frame, text="Export Matched Data", command=self.export_matched_data)
        export_matched_button.pack(side=tk.LEFT, padx=5)

        # Export Unmatched TC Button
        export_unmatched_tc_button = ttk.Button(export_frame, text="Export Unmatched TC Data", command=self.export_unmatched_data)
        export_unmatched_tc_button.pack(side=tk.LEFT, padx=5)

        # NEW BUTTON: Export Unmatched LMRB Button - POSITIONED AFTER TC BUTTON
        export_unmatched_lmrb_button = ttk.Button(export_frame, text="Export Unmatched LMRB Data", command=self.export_unmatched_lmrb_data)
        export_unmatched_lmrb_button.pack(side=tk.LEFT, padx=5)

        # Export Detailed Excel Report Button
        export_report_button = ttk.Button(export_frame, text="Export Detailed Excel Report", command=self.export_detailed_excel_report)
        export_report_button.pack(side=tk.LEFT, padx=5)

        # Export Highlighted LMRB Button
        export_highlighted_lmrb_button = ttk.Button(export_frame, text="Export Highlighted LMRB Data", command=self.export_highlighted_lmrb)
        export_highlighted_lmrb_button.pack(side=tk.LEFT, padx=5)

        # Generate PDF Report Button
        generate_pdf_button = ttk.Button(export_frame, text="Generate PDF Report", command=self.generate_pdf_report)
        generate_pdf_button.pack(side=tk.LEFT, padx=5)

        # Make the grid expandable
        self.results_tab.grid_rowconfigure(2, weight=1)
        self.results_tab.grid_columnconfigure(0, weight=1)
        self.results_tab.grid_columnconfigure(1, weight=1)

    def browse_lmrb_file(self):
        """Browse and load LMRB file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file_path:
                # Load the LMRB file
                self.lmrb_data = pd.read_excel(file_path)
                self.lmrb_file_var.set(file_path)
                
                # Auto-detect required columns
                theme_col = next(
                    (col for col in self.lmrb_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()),
                    None
                )
                
                if not theme_col:
                    messagebox.showerror("Error", "No theme column found in LMRB data")
                    return False
                    
                if "Channel" not in self.lmrb_data.columns:
                    messagebox.showerror("Error", "Channel column not found in LMRB data")
                    return False
                
                # Update channel selection
                channels = sorted(self.lmrb_data["Channel"].unique())
                self.channel_combobox["values"] = channels
                
                if channels:
                    self.channel_combobox.current(0)
                    self.selected_channel = channels[0]
                    self.lmrb_filtered = self.filter_by_channel(self.lmrb_data, channels[0])
                
                # Update LMRB preview
                self.update_treeview(self.lmrb_tree, self.lmrb_data.head(100))
                
                # Enable TC file button
                if hasattr(self, 'tc_button'):
                    self.tc_button['state'] = 'normal'
                
                self.status_var.set("LMRB file loaded successfully")
                return True
                
        except Exception as e:
            messagebox.showerror("Error", f"Error loading LMRB file: {str(e)}")
            return False

    def browse_tc_file(self):
        """Browse and load TC file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file_path:
                # Load the TC file
                self.tc_data = pd.read_excel(file_path)
                
                # Clean dates right after loading
                self.tc_data = self.clean_dataframe_dates(self.tc_data)
                self.tc_data_original = self.tc_data.copy()  # Store cleaned copy
                
                self.tc_file_var.set(file_path)

                # Auto-detect columns
                theme_col = next(
                    (col for col in self.tc_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()),
                    None
                )
                
                if not theme_col:
                    messagebox.showerror("Error", "No theme column found in TC data")
                    return False
                
                # Update TC preview
                self.update_treeview(self.tc_tree, self.tc_data.head(100))
                
                # Enable theme mapping tab if LMRB data is also loaded
                if self.lmrb_filtered is not None:
                    self.notebook.tab(1, state="normal")
                    #self.notebook.select(1)  # Switch to theme mapping tab
                    self.populate_theme_dropdowns()
                
                self.status_var.set("TC file loaded successfully")
                return True
                
        except Exception as e:
            messagebox.showerror("Error", f"Error loading TC file: {str(e)}")
            return False

    def load_lmrb_data(self, file_path):
        """Load LMRB data from the selected file."""
        try:
            self.status_var.set("Loading LMRB data...")
            self.root.update_idletasks()
            
            # Load LMRB data
            self.lmrb_data = self.read_file(file_path)
            
            # Check if Channel column exists
            if "Channel" not in self.lmrb_data.columns:
                messagebox.showerror("Error", "Channel column not found in LMRB data")
                return
            
            # Populate channel dropdown
            channels = sorted(self.lmrb_data["Channel"].unique())
            self.channel_combobox["values"] = channels
            
            if channels:
                self.channel_combobox.current(0)
            
            # Update LMRB treeview
            self.update_treeview(self.lmrb_tree, self.lmrb_data.head(100))
            
            self.status_var.set("LMRB data loaded successfully. Please select a channel.")
            messagebox.showinfo("Success", "LMRB data loaded successfully. Please select a channel.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error loading LMRB data: {str(e)}")
            self.status_var.set("Error loading LMRB data")

    def browse_schedule_file(self):
        """Browse and load Schedule file"""
        try:
            file_path = filedialog.askopenfilename(
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if file_path:
                # Load the Schedule file
                self.schedule_data = pd.read_excel(file_path)
                
                # Clean dates right after loading
                self.schedule_data = self.clean_dataframe_dates(self.schedule_data)
                self.schedule_data_original = self.schedule_data.copy()  # Store cleaned copy

                self.schedule_file_var.set(file_path)

                # Auto-detect columns
                theme_col = next(
                    (col for col in self.schedule_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()),
                    None
                )
    

                if not theme_col:
                    messagebox.showerror("Error", "No theme column found in Schedule data")
                    return False

                # Update Schedule preview
                self.update_treeview(self.schedule_tree, self.schedule_data.head(100))

                # Enable theme mapping tab if LMRB data is also loaded
                if self.lmrb_filtered is not None:
                    self.notebook.tab(1, state="normal")
                    self.notebook.select(1)  # Switch to theme mapping tab
                    self.populate_theme_dropdowns()

                self.status_var.set("Schedule file loaded successfully")
                return True
                
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Schedule file: {str(e)}")
            return False


    def load_data(self):
        lmrb_file = self.lmrb_file_var.get()
        tc_file = self.tc_file_var.get()
        schedule_file = self.schedule_file_var.get()  # New variable for schedule file
        
        if not lmrb_file:
            messagebox.showerror("Error", "Please select LMRB data file")
            return
        
        if not tc_file:
            messagebox.showerror("Error", "Please select TC data file")
            return
        
        if not schedule_file:
            messagebox.showerror("Error", "Please select Schedule data file")
            return

        try:
            self.status_var.set("Loading LMRB data...")
            self.root.update_idletasks()
            
            # Load LMRB data
            self.lmrb_data = self.read_file(lmrb_file)
            
            # Check if Channel column exists
            if "Channel" not in self.lmrb_data.columns:
                messagebox.showerror("Error", "Channel column not found in LMRB data")
                return
            
            # Populate channel dropdown
            channels = sorted(self.lmrb_data["Channel"].unique())
            self.channel_combobox["values"] = channels
            
            if channels:
                self.channel_combobox.current(0)
            
            # Load TC data
            self.status_var.set("Loading TC data...")
            self.root.update_idletasks()
            self.tc_data = self.read_file(tc_file)
            self.tc_data_original = self.tc_data.copy()  # <-- Store original
            
            # Update LMRB treeview
            self.update_treeview(self.lmrb_tree, self.lmrb_data.head(100))
            
            # Update TC treeview
            self.update_treeview(self.tc_tree, self.tc_data.head(100))

            #Update Schedule treeview
            self.update_treeview(self.schedule_tree, self.schedule_data.head(100))
            
            self.status_var.set("Data loaded successfully. Please select a channel.")
            messagebox.showinfo("Success", "Data loaded successfully. Please select a channel.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error loading data: {str(e)}")
            self.status_var.set("Error loading data")

    def read_file(self, file_path):
        """Read an uploaded file into a pandas DataFrame."""
        try:
            file_extension = file_path.split('.')[-1].lower()
            
            if file_extension == 'xlsx':
                df = pd.read_excel(file_path, engine='openpyxl')
            elif file_extension == 'xls':
                df = pd.read_excel(file_path, engine='xlrd')
            elif file_extension == 'csv':
                df = pd.read_csv(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            df = df.reset_index(drop=True)
            df = clean_column_names(df)
            
            return df
            
        except Exception as e:
            raise Exception(f"Error reading file: {str(e)}")    

    def update_treeview(self, tree, df):
        """Update a treeview with dataframe content"""
        # Clear existing data
        for item in tree.get_children():
            tree.delete(item)
        
        # Configure columns
        tree["columns"] = list(df.columns)
        
        # Configure column headings
        tree.column("#0", width=50, stretch=tk.NO)
        for col in df.columns:
            tree.column(col, width=100)
            tree.heading(col, text=col)
        
        # Add data rows
        for i, row in df.iterrows():
            values = [row[col] for col in df.columns]
            tree.insert("", "end", text=str(i), values=values)

    def on_channel_selected(self, event=None):
        channel = self.channel_var.get()
        if channel and self.lmrb_data is not None:
            self.selected_channel = channel
            self.lmrb_filtered = self.filter_by_channel(self.lmrb_data, channel)
            
            # ✅ NEW: Clear and repopulate products after channel selection
            self.selected_products.clear()
            self.populate_product_dropdown()
            
            # ✅ NEW: Apply product filter (will show all since none selected)
            self.apply_product_filter()
            
            # Enable mapping tab
            self.notebook.tab(1, state="normal")
            
            self.status_var.set(f"Selected channel: {channel}. Select products to filter (optional) or proceed to mapping.")

    def populate_product_dropdown(self):
        """Populate product dropdown from LMRB data after channel selection."""
        try:
            if self.lmrb_filtered is None or self.lmrb_filtered.empty:
                return
            
            # Find product column (try different possible names)
            product_cols = ['Product Details', 'Product', 'Advt_Theme', 'Theme', 'Product_Details']
            product_col = None
            
            for col in product_cols:
                if col in self.lmrb_filtered.columns:
                    product_col = col
                    break
            
            if not product_col:
                print("⚠️ No product column found in LMRB data")
                self.product_combobox["values"] = []
                return
            
            # Get unique products, sorted
            unique_products = sorted(self.lmrb_filtered[product_col].dropna().unique())
            self.available_products = unique_products
            
            # Update dropdown
            self.product_combobox["values"] = unique_products
            
            print(f"✅ Found {len(unique_products)} unique products in LMRB data")
            
            # Update status
            self.product_filter_status.config(
                text=f"Found {len(unique_products)} products. Select products to filter or leave empty for all.",
                foreground="blue"
            )
            
        except Exception as e:
            print(f"Error populating product dropdown: {e}")
            messagebox.showerror("Error", f"Error loading products: {str(e)}")

    def add_selected_product(self):
        """Add selected product from dropdown to selected list."""
        try:
            selected_product = self.product_var.get()
            if not selected_product:
                messagebox.showwarning("Warning", "Please select a product from the dropdown")
                return
            
            if selected_product in self.selected_products:
                messagebox.showinfo("Info", f"Product '{selected_product}' is already selected")
                return
            
            # Add to selected list
            self.selected_products.append(selected_product)
            
            # Update listbox
            self.update_selected_products_display()
            
            # Apply filter
            self.apply_product_filter()
            
            # Clear dropdown selection
            self.product_var.set("")
            
            print(f"✅ Added product: {selected_product}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error adding product: {str(e)}")

    def remove_selected_products(self):
        """Remove selected products from the selected list."""
        try:
            selected_indices = self.selected_products_listbox.curselection()
            if not selected_indices:
                messagebox.showwarning("Warning", "Please select products to remove from the list")
                return
            
            # Remove selected products (in reverse order to maintain indices)
            for index in reversed(selected_indices):
                product_name = self.selected_products[index]
                self.selected_products.remove(product_name)
                print(f"🗑️ Removed product: {product_name}")
            
            # Update display
            self.update_selected_products_display()
            
            # Apply filter
            self.apply_product_filter()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error removing products: {str(e)}")

    def clear_selected_products(self):
        """Clear all selected products."""
        try:
            if not self.selected_products:
                messagebox.showinfo("Info", "No products selected to clear")
                return
            
            result = messagebox.askyesno("Confirm Clear", 
                f"Clear all {len(self.selected_products)} selected products?\n\n"
                "This will show all products in the data.")
            
            if result:
                self.selected_products.clear()
                self.update_selected_products_display()
                self.apply_product_filter()
                print("🗑️ Cleared all selected products")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error clearing products: {str(e)}")

    def select_all_products(self):
        """Select all available products."""
        try:
            if not self.available_products:
                messagebox.showwarning("Warning", "No products available to select")
                return
            
            result = messagebox.askyesno("Confirm Select All", 
                f"Select all {len(self.available_products)} available products?\n\n"
                "This will include all products in matching.")
            
            if result:
                self.selected_products = self.available_products.copy()
                self.update_selected_products_display()
                self.apply_product_filter()
                print(f"✅ Selected all {len(self.available_products)} products")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error selecting all products: {str(e)}")

    def update_selected_products_display(self):
        """Update the selected products listbox display."""
        try:
            # Clear listbox
            self.selected_products_listbox.delete(0, tk.END)
            
            # Add selected products
            for product in self.selected_products:
                self.selected_products_listbox.insert(tk.END, product)
            
            # Update status
            if self.selected_products:
                status_text = f"✅ {len(self.selected_products)} products selected - showing filtered data"
                status_color = "green"
            else:
                status_text = "No products selected - showing all products"
                status_color = "gray"
            
            self.product_filter_status.config(text=status_text, foreground=status_color)
            
        except Exception as e:
            print(f"Error updating products display: {e}")

    def apply_product_filter(self):
        """Apply product filter to LMRB data and update theme dropdowns."""
        try:
            if self.lmrb_filtered is None or self.lmrb_filtered.empty:
                print("No LMRB data to filter")
                return
            
            # Find product column
            product_cols = ['Product Details', 'Product', 'Advt_Theme', 'Theme', 'Product_Details']
            product_col = None
            
            for col in product_cols:
                if col in self.lmrb_filtered.columns:
                    product_col = col
                    break
            
            if not product_col:
                print("⚠️ No product column found for filtering")
                return
            
            # Apply product filter
            if self.selected_products:
                # Filter by selected products
                self.lmrb_product_filtered = self.lmrb_filtered[
                    self.lmrb_filtered[product_col].isin(self.selected_products)
                ].reset_index(drop=True)
                
                print(f"🔍 Product filter applied: {len(self.lmrb_filtered)} → {len(self.lmrb_product_filtered)} records")
                print(f"   Selected products: {', '.join(self.selected_products[:3])}{'...' if len(self.selected_products) > 3 else ''}")
            else:
                # No products selected - show all
                self.lmrb_product_filtered = self.lmrb_filtered.copy()
                print("📋 No product filter - showing all products")
            
            # Update theme dropdowns with filtered data
            self.populate_theme_dropdowns()
            
            # Update LMRB preview with filtered data
            self.update_treeview(self.lmrb_tree, self.lmrb_product_filtered.head(100))
            
            # Update status
            filter_summary = f"Channel: {self.selected_channel}, Products: {len(self.selected_products) if self.selected_products else 'All'}"
            self.status_var.set(f"LMRB filtered - {filter_summary}")
            
        except Exception as e:
            print(f"Error applying product filter: {e}")
            messagebox.showerror("Error", f"Error applying product filter: {str(e)}")

    def filter_by_channel(self, df, channel):
        """Filter data by selected channel."""
        if "Channel" not in df.columns:
            messagebox.error("Channel column not found in the dataframe.")
            return df
        return df[df["Channel"] == channel].reset_index(drop=True)

    def filter_lmrb_by_channel(self, channel_window):
            """
            Filter LMRB data by selected channel
            """
            try:
                selected_channel = self.channel_var.get()
                if not selected_channel:
                    messagebox.showerror("Error", "Please select a channel")
                    return
                
                self.selected_channel = selected_channel
                self.filtered_lmrb_df = self.lmrb_df[self.lmrb_df['Channel'] == selected_channel].copy()
                
                # Close the channel selection window
                channel_window.destroy()
                
                # Enable TC file button if it exists
                if hasattr(self, 'tc_button'):
                    self.tc_button['state'] = 'normal'
                
                messagebox.showinfo("Success", f"Data filtered for channel: {selected_channel}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error filtering channel: {str(e)}")
                return False
            
    def clean_dataframe_dates(self, df):
        """Clean date columns in a dataframe to remove time components."""
        if df is None or df.empty:
            return df

        df_cleaned = df.copy()

        # Find all date columns
        date_columns = []
        for col in df_cleaned.columns:
            if 'date' in col.lower():
                date_columns.append(col)

        # Clean each date column
        for date_col in date_columns:
            df_cleaned[date_col] = df_cleaned[date_col].apply(self.format_date_display)

        return df_cleaned

    def format_date_display(self, date_value):
        """Format date for display without time component."""
        if pd.isna(date_value):
            return ""

        date_str = str(date_value).strip()

        # Remove time component if present
        if ' 00:00:00' in date_str:
            date_str = date_str.replace(' 00:00:00', '')

        return date_str

    def populate_theme_dropdowns(self):
        """Populate theme and duration dropdowns using product-filtered data"""
        # ✅ CHANGE: Use product-filtered LMRB data instead of just channel-filtered
        lmrb_data_to_use = self.lmrb_product_filtered if self.lmrb_product_filtered is not None else self.lmrb_filtered

        if lmrb_data_to_use is not None:
            # Find theme column in LMRB data
            lmrb_theme_col = next(
                (col for col in lmrb_data_to_use.columns if 'theme' in col.lower() or 'aadvt' in col.lower()), 
                None
            )
            
            if lmrb_theme_col:
                # Get unique themes from filtered data
                lmrb_themes = sorted(lmrb_data_to_use[lmrb_theme_col].dropna().unique())
                self.lmrb_theme_combobox["values"] = lmrb_themes
                print(f"📋 LMRB themes updated: {len(lmrb_themes)} themes from filtered data")
                
                # Get durations if available
                lmrb_dur_col = next(
                    (col for col in lmrb_data_to_use.columns if 'dur' in col.lower()), 
                    None
                )
                if lmrb_dur_col:
                    lmrb_durs = sorted(lmrb_data_to_use[lmrb_dur_col].dropna().unique())
                    self.lmrb_dur_combobox["values"] = lmrb_durs

        # Populate TC themes if TC data is available
        if self.tc_data is not None:
            # Find theme column in TC data
            tc_theme_col = next(
                (col for col in self.tc_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()),
                None
            )
            
            if tc_theme_col:
                # Get unique themes
                tc_themes = sorted(self.tc_data[tc_theme_col].dropna().unique())
                self.tc_theme_combobox["values"] = tc_themes
                
                # Get durations if available
                tc_dur_col = next(
                    (col for col in self.tc_data.columns if 'dur' in col.lower()),
                    None
                )
                if tc_dur_col:
                    tc_durs = sorted(self.tc_data[tc_dur_col].dropna().unique())
                    self.tc_dur_combobox["values"] = tc_durs
        else:
            # Clear TC dropdowns if no TC data
            self.tc_theme_combobox["values"] = []
            self.tc_dur_combobox["values"] = []

        # Populate Schedule themes if schedule data is available
        if self.schedule_data is not None:
            # Find theme column in Schedule data
            schedule_theme_col = next(
                (col for col in self.schedule_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()),
                None
            )

            if schedule_theme_col:
                # Get unique themes
                schedule_themes = sorted(self.schedule_data[schedule_theme_col].dropna().unique())
                self.schedule_theme_combobox["values"] = schedule_themes

                # Get durations if available
                schedule_dur_col = next(
                    (col for col in self.schedule_data.columns if 'dur' in col.lower()),
                    None
                )
                if schedule_dur_col:
                    schedule_durs = sorted(self.schedule_data[schedule_dur_col].dropna().unique())
                    self.schedule_dur_combobox["values"] = schedule_durs
        else:
            # Clear schedule dropdowns if no schedule data
            self.schedule_theme_combobox["values"] = []
            self.schedule_dur_combobox["values"] = []


    def setup_gui(self):
            """
            Set up the main GUI window
            """
            self.root = tk.Tk()
            self.root.title("Media Reconciliation System")
            self.root.geometry("1280x1240")
            
            # Create and configure the main frame
            main_frame = ttk.Frame(self.root, padding="10")
            main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
            
            # LMRB file button
            ttk.Button(
                main_frame,
                text="Load LMRB File",
                command=self.browse_lmrb_file
            ).grid(row=0, column=0, pady=10, padx=10, sticky=tk.W+tk.E)
            
            # TC file button (initially disabled)
            self.tc_button = ttk.Button(
                main_frame,
                text="Load TC File",
                command=self.browse_tc_file
            )
            self.tc_button.grid(row=1, column=0, pady=10, padx=10, sticky=tk.W+tk.E)

                        
            # Schedule file button (initially disabled)
            self.schedule_button = ttk.Button(
                main_frame,
                text="Load Schedule File",
                command=self.browse_schedule_file,
                state='disabled'
            )
            self.schedule_button.grid(row=1, column=0, pady=10, padx=10, sticky=tk.W+tk.E)

            # Excel report button (initially disabled)
            self.excel_button = ttk.Button(
                main_frame,
                text="Generate Excel Report",
                command=self.generate_excel_report,
                state='disabled'
            )
            self.excel_button.grid(row=2, column=0, pady=10, padx=10, sticky=tk.W+tk.E)
            
            # PDF report button (initially disabled)
            self.pdf_button = ttk.Button(
                main_frame,
                text="Generate PDF Report",
                command=self.generate_pdf_report,
                state='disabled'
            )
            self.pdf_button.grid(row=3, column=0, pady=10, padx=10, sticky=tk.W+tk.E)
            
            # Status label
            self.status_label = ttk.Label(main_frame, text="Ready")
            self.status_label.grid(row=4, column=0, pady=10, padx=10)
            
            return self.root

    def add_theme_mapping(self):
        """Add a new theme mapping with support for 3-way mapping and multiple TC mappings per LMRB theme"""
        try:
            lmrb_theme = self.lmrb_theme_var.get()
            lmrb_dur = self.lmrb_dur_var.get()
            schedule_theme = self.schedule_theme_var.get()
            schedule_dur = self.schedule_dur_var.get()
            tc_theme = self.tc_theme_var.get()
            tc_dur = self.tc_dur_var.get()
            mapping_type = self.mapping_type_var.get()
            
            # Validate required fields - LMRB theme is always required
            if not lmrb_theme:
                messagebox.showerror("Error", "Please select LMRB theme (required)")
                return
            
            # For full mapping, TC theme is required
            if mapping_type == "COMMERCIAL BENEFITS" and not tc_theme:
                messagebox.showerror("Error", "TC theme is required for Commercial Benefits mapping")
                return
            
            # For sponsorship mapping, schedule theme is required
            if mapping_type == "SPONSORSHIP BENEFITS" and not schedule_theme:
                messagebox.showerror("Error", "Schedule theme is required for Sponsorship Benefits mapping")
                return
            
            # ✅ NEW: Check for duplicate LMRB+TC combination and show warning (but allow)
            existing_count = 0
            existing_tc_themes = []
            for mapping in self.theme_mapping:
                existing_lmrb = mapping.get('lmrb_theme', mapping.get('media_watch_theme', ''))
                existing_tc = mapping.get('tc_theme', '')
                
                if existing_lmrb == lmrb_theme:
                    existing_count += 1
                    if existing_tc and existing_tc not in existing_tc_themes:
                        existing_tc_themes.append(existing_tc)
            
            if existing_count > 0 and tc_theme:
                # Check if this exact TC theme already exists for this LMRB
                if tc_theme in existing_tc_themes:
                    messagebox.showerror("Error", 
                        f"Exact mapping already exists:\n"
                        f"LMRB '{lmrb_theme}' → TC '{tc_theme}'\n\n"
                        f"Please choose a different TC theme or modify the existing mapping.")
                    return
                
                result = messagebox.askyesno("Multiple Mapping Detected", 
                    f"⚠️ LMRB theme '{lmrb_theme}' already has {existing_count} mapping(s):\n"
                    f"Existing TC themes: {', '.join(existing_tc_themes)}\n\n"
                    f"Adding TC theme '{tc_theme}' will enable sequential matching:\n"
                    f"• 1st attempt: Try existing mappings\n"
                    f"• 2nd attempt: Try new mapping → '{tc_theme}'\n\n"
                    f"This helps when one LMRB theme matches multiple TC themes.\n\n"
                    f"Continue adding this mapping?")
                if not result:
                    return
            
            # Determine mapping type display
            type_display = "COMMERCIAL BENEFITS" if mapping_type == "COMMERCIAL BENEFITS" else "SPONSORSHIP BENEFITS"
            
            # Add to theme mapping list
            mapping_entry = {
                'mapping_type': mapping_type,
                'media_watch_theme': lmrb_theme,
                'lmrb_theme': lmrb_theme,  # Keep both for compatibility
                'lmrb_duration': lmrb_dur or "",
                'tc_theme': tc_theme or "",
                'tc_duration': tc_dur or "",
                'schedule_theme': schedule_theme or "",
                'schedule_duration': schedule_dur or ""
            }
            self.theme_mapping.append(mapping_entry)
            
            # ✅ NEW: Update mappings tree with enhanced display
            self.update_mappings_tree_display()
            
            # NEW: Update schedule data with mapped_theme column
            self.update_schedule_mapped_themes()
            
            # Clear selections
            self.lmrb_theme_var.set('')
            self.lmrb_dur_var.set('')
            self.tc_theme_var.set('')
            self.tc_dur_var.set('')
            self.schedule_theme_var.set('')
            self.schedule_dur_var.set('')
            
            # ✅ NEW: Show enhanced status message
            total_lmrb_mappings = len([m for m in self.theme_mapping 
                                    if m.get('lmrb_theme', m.get('media_watch_theme', '')) == lmrb_theme])
            
            status_msg = f"Added {type_display} mapping: {lmrb_theme}"
            if total_lmrb_mappings > 1:
                status_msg += f" (now has {total_lmrb_mappings} TC mappings - will try sequentially)"
            
            self.status_var.set(status_msg)
            print(f"✅ {status_msg}")

        except Exception as e:
            messagebox.showerror("Error", f"Error adding theme mapping: {str(e)}")
            import traceback
            traceback.print_exc()

    def update_mappings_tree_display(self):
        """Update the mappings tree with multiple mapping indicators"""
        try:
            # Clear existing tree
            for item in self.mappings_tree.get_children():
                self.mappings_tree.delete(item)
            
            # Count mappings per LMRB theme
            lmrb_counts = {}
            for mapping in self.theme_mapping:
                lmrb_theme = mapping.get('lmrb_theme', mapping.get('media_watch_theme', ''))
                if lmrb_theme:
                    lmrb_counts[lmrb_theme] = lmrb_counts.get(lmrb_theme, 0) + 1
            
            # Populate tree with indicators
            for i, mapping in enumerate(self.theme_mapping):
                lmrb_theme = mapping.get('lmrb_theme', mapping.get('media_watch_theme', ''))
                tc_theme = mapping.get('tc_theme', '')
                count = lmrb_counts.get(lmrb_theme, 1)
                
                # Add indicator for multiple mappings
                if count > 1 and tc_theme:
                    display_lmrb = f"🔀 {lmrb_theme} (#{count})"
                elif tc_theme:
                    display_lmrb = lmrb_theme
                else:
                    display_lmrb = f"📋 {lmrb_theme} (Schedule only)"
                
                self.mappings_tree.insert("", "end", values=(
                    mapping.get('mapping_type', ''),
                    mapping.get('tc_theme', ''),
                    mapping.get('tc_duration', ''),
                    display_lmrb,  # Enhanced display
                    mapping.get('lmrb_duration', ''),
                    mapping.get('schedule_theme', ''),
                    mapping.get('schedule_duration', '')
                ))
                
        except Exception as e:
            print(f"Error updating mappings tree display: {e}")
            # Fallback to simple display
            for item in self.mappings_tree.get_children():
                self.mappings_tree.delete(item)
            
            for mapping in self.theme_mapping:
                self.mappings_tree.insert("", "end", values=(
                    mapping.get('mapping_type', ''),
                    mapping.get('tc_theme', ''),
                    mapping.get('tc_duration', ''),
                    mapping.get('lmrb_theme', mapping.get('media_watch_theme', '')),
                    mapping.get('lmrb_duration', ''),
                    mapping.get('schedule_theme', ''),
                    mapping.get('schedule_duration', '')
                ))

    def update_schedule_mapped_themes(self):
        """Create/update mapped_theme column in schedule data based on current theme mappings"""
        try:
            if self.schedule_data is None or self.schedule_data.empty:
                print("No schedule data available to update")
                return
                
            print(f"🔄 Updating schedule mapped themes...")
            
            # Initialize mapped_theme column if it doesn't exist
            if 'mapped_theme' not in self.schedule_data.columns:
                self.schedule_data['mapped_theme'] = ''
                print("   Created new mapped_theme column")
            else:
                # Clear existing mapped themes
                self.schedule_data['mapped_theme'] = ''
                print("   Cleared existing mapped_theme column")
                
            # Find schedule theme and duration columns
            schedule_theme_col = self._find_schedule_column(['theme', 'aadvt', 'advt_theme'])
            schedule_dur_col = self._find_schedule_column(['dur', 'duration'])
            
            if not schedule_theme_col:
                print("❌ Could not find schedule theme column")
                return
                
            print(f"   Using columns - Theme: {schedule_theme_col}, Duration: {schedule_dur_col}")
                
            # Create mapping lookup: (schedule_theme, duration) -> lmrb_theme
            mapping_lookup = {}
            for mapping in self.theme_mapping:
                schedule_theme = mapping.get('schedule_theme', '')
                schedule_duration = mapping.get('schedule_duration', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                
                if schedule_theme and lmrb_theme:
                    key = (schedule_theme, schedule_duration)
                    mapping_lookup[key] = lmrb_theme
                    print(f"   Mapping: '{schedule_theme}' ({schedule_duration}s) → '{lmrb_theme}'")
                    
            print(f"   Created mapping lookup with {len(mapping_lookup)} entries")
            
            # Update mapped_theme column
            updated_count = 0
            for index, row in self.schedule_data.iterrows():
                schedule_theme = row[schedule_theme_col]
                schedule_duration = str(row[schedule_dur_col]) if schedule_dur_col else ''
                
                # Look up the mapped theme
                lookup_key = (schedule_theme, schedule_duration)
                mapped_theme = mapping_lookup.get(lookup_key, '')
                
                # Update the mapped_theme column
                self.schedule_data.at[index, 'mapped_theme'] = mapped_theme
                
                if mapped_theme:
                    updated_count += 1
                    
            print(f"✅ Updated {updated_count} schedule records with mapped themes")
            
            # Debug: Show some examples
            if updated_count > 0:
                sample = self.schedule_data[self.schedule_data['mapped_theme'] != ''].head(3)
                print("   Sample mapped entries:")
                for idx, row in sample.iterrows():
                    print(f"     '{row[schedule_theme_col]}' ({row[schedule_dur_col]}s) → '{row['mapped_theme']}'")
            
        except Exception as e:
            print(f"❌ Error updating schedule mapped themes: {e}")
            import traceback
            traceback.print_exc()

    def create_grouped_schedule_df(self):
        """Create a grouped schedule DataFrame with mapped themes for easier counting"""
        try:
            if self.schedule_data is None or self.schedule_data.empty:
                print("No schedule data available for grouping")
                return None
                
            print("📊 Creating grouped schedule DataFrame...")
                
            # Ensure mapped_theme column exists and is up to date
            if 'mapped_theme' not in self.schedule_data.columns:
                print("   Mapped theme column missing, updating...")
                self.update_schedule_mapped_themes()
                
            # Find required columns
            schedule_theme_col = self._find_schedule_column(['theme', 'aadvt', 'advt_theme'])
            schedule_dur_col = self._find_schedule_column(['dur', 'duration'])
            schedule_adtype_col = self._find_schedule_column(['advertisement_type', 'ad_type'])
            
            if not schedule_theme_col or not schedule_dur_col:
                print(f"❌ Required columns not found - Theme: {schedule_theme_col}, Duration: {schedule_dur_col}")
                return None
                
            print(f"   Using columns - Theme: {schedule_theme_col}, Duration: {schedule_dur_col}, AdType: {schedule_adtype_col}")
                
            # Prepare grouping columns
            group_cols = [schedule_theme_col, schedule_dur_col, 'mapped_theme']
            if schedule_adtype_col:
                group_cols.append(schedule_adtype_col)
            else:
                # Add default ad type if column doesn't exist
                self.schedule_data['temp_ad_type'] = 'COMMERCIAL BENEFITS'
                group_cols.append('temp_ad_type')
                schedule_adtype_col = 'temp_ad_type'
                
            print(f"   Grouping by: {group_cols}")
                
            # Group and count
            grouped_schedule = self.schedule_data.groupby(group_cols).size().reset_index(name='planned_count')
            
            # Add helpful columns
            grouped_schedule['has_mapping'] = grouped_schedule['mapped_theme'] != ''
            grouped_schedule['unique_key'] = (
                grouped_schedule[schedule_theme_col].astype(str) + "_" + 
                grouped_schedule[schedule_dur_col].astype(str)
            )
            
            print(f"✅ Created grouped schedule with {len(grouped_schedule)} unique combinations")
            
            # Debug: Show grouping results
            mapped_count = len(grouped_schedule[grouped_schedule['has_mapping']])
            unmapped_count = len(grouped_schedule[~grouped_schedule['has_mapping']])
            print(f"   Mapped combinations: {mapped_count}")
            print(f"   Unmapped combinations: {unmapped_count}")
            
            # Show sample grouped data
            if len(grouped_schedule) > 0:
                print("   Sample grouped entries:")
                sample = grouped_schedule.head(5)
                for idx, row in sample.iterrows():
                    mapped_status = "✅" if row['has_mapping'] else "❌"
                    print(f"     {mapped_status} '{row[schedule_theme_col]}' ({row[schedule_dur_col]}s) → '{row['mapped_theme']}' | Planned: {row['planned_count']}")
            
            # Clean up temporary column if we created it
            if 'temp_ad_type' in self.schedule_data.columns:
                self.schedule_data.drop('temp_ad_type', axis=1, inplace=True)
                
            return grouped_schedule
            
        except Exception as e:
            print(f"❌ Error creating grouped schedule DataFrame: {e}")
            import traceback
            traceback.print_exc()
            return None

    def delete_theme_mapping(self):
        """Delete selected theme mapping and update schedule mapped themes"""
        try:
            selected = self.mappings_tree.selection()
            if not selected:
                messagebox.showerror("Error", "Please select a mapping to delete")
                return
            
            print(f"🗑️ Deleting {len(selected)} theme mapping(s)...")
            
            for item in selected:
                values = self.mappings_tree.item(item)['values']
                print(f"   Deleting: {values[3]} (LMRB) ← {values[5]} (Schedule)")
                
                # Remove from theme mapping list
                self.theme_mapping = [
                    m for m in self.theme_mapping 
                    if not ((m.get('lmrb_theme', m.get('media_watch_theme')) == values[3]) and 
                        (m.get('tc_theme', '') == values[1]) and
                        (m.get('schedule_theme', '') == values[5]))
                ]
                self.mappings_tree.delete(item)
            
            # Update schedule mapped themes after deletion
            print("🔄 Updating schedule mapped themes after deletion...")
            self.update_schedule_mapped_themes()
            
            self.status_var.set(f"Deleted {len(selected)} mapping(s) and updated schedule themes")
            print("✅ Mapping(s) deleted and schedule themes updated")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error deleting theme mapping: {str(e)}")
            import traceback
            traceback.print_exc()

    def continue_to_matching(self):
        """Proceed to matching tab"""
        try:
            if not self.theme_mapping:
                messagebox.showerror("Error", "Please add at least one theme mapping before continuing")
                return
            
            # Check if we have at least TC and LMRB mappings (schedule is optional)
            has_valid_mapping = False
            for mapping in self.theme_mapping:
                tc_theme = mapping.get('tc_theme', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                
                # At least one mapping should have both TC and LMRB themes
                if tc_theme.strip() and lmrb_theme.strip():
                    has_valid_mapping = True
                    break
            
            if not has_valid_mapping:
                messagebox.showerror("Error", "Please add at least one mapping with both TC and LMRB themes")
                return
            
            # Enable matching tab
            self.notebook.tab(2, state="normal")
            # Switch to matching tab
            self.notebook.select(2)
            self.status_var.set("Ready to start matching process")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error continuing to matching: {str(e)}")

    def run_matching(self):
        """FIXED: Use product-filtered LMRB data for matching."""
        # ✅ CHANGE: Use product-filtered data instead of channel-filtered only
        lmrb_data_for_matching = self.lmrb_product_filtered if self.lmrb_product_filtered is not None else self.lmrb_filtered

        if not all([lmrb_data_for_matching is not None, self.tc_data is not None]):
            messagebox.showerror("Error", "Please load both LMRB and TC data first")
            return

        if not self.theme_mapping:
            messagebox.showerror("Error", "Please create theme mappings before running the matching process")
            return

        # FIXED: Properly handle existing matches
        existing_matches = []
        preserve_existing = False

        if hasattr(self, 'matched_data') and self.matched_data:
            # Count existing matches properly
            if isinstance(self.matched_data, list):
                existing_count = len([m for m in self.matched_data if m])  # Count non-empty matches
            elif isinstance(self.matched_data, pd.DataFrame):
                existing_count = len(self.matched_data)
            else:
                existing_count = 0
                
            if existing_count > 0:
                # Show filtering info in confirmation dialog
                if self.selected_products:
                    product_info = f"Using {len(self.selected_products)} selected products:\n{', '.join(self.selected_products[:5])}{'...' if len(self.selected_products) > 5 else ''}"
                else:
                    product_info = "Using all products (no product filter applied)"
                
                result = messagebox.askyesno("Existing Matches Found", 
                    f"Found {existing_count} existing matches.\n\n"
                    f"📊 Current Filter Settings:\n"
                    f"• Channel: {self.selected_channel}\n"
                    f"• LMRB records to match: {len(lmrb_data_for_matching):,}\n"
                    f"• {product_info}\n\n"
                    f"YES = Keep existing + add new matches\n"
                    f"NO = Replace with new matches only")
                
                if result:  # User wants to preserve
                    preserve_existing = True
                    if isinstance(self.matched_data, list):
                        existing_matches = [m for m in self.matched_data if m]  # Filter out empty matches
                    elif isinstance(self.matched_data, pd.DataFrame):
                        existing_matches = self.matched_data.to_dict('records')
            else:
                # Show filtering info for new matching
                if self.selected_products:
                    product_info = f"Using {len(self.selected_products)} selected products:\n{', '.join(self.selected_products[:5])}{'...' if len(self.selected_products) > 5 else ''}"
                else:
                    product_info = "Using all products (no product filter applied)"
                
                result = messagebox.askyesno("Start Matching", 
                    f"Start matching process?\n\n"
                    f"📊 Data Summary:\n"
                    f"• Channel: {self.selected_channel}\n"
                    f"• LMRB records: {len(lmrb_data_for_matching):,}\n"
                    f"• TC records: {len(self.tc_data):,}\n"
                    f"• Theme mappings: {len(self.theme_mapping)}\n\n"
                    f"🔍 Product Filter:\n"
                    f"• {product_info}\n\n"
                    f"Continue with matching?")
                
                if not result:
                    return

        self.matching_status_var.set("Running matching process...")
        self.progress_var.set(0)
        self.root.update_idletasks()

        # Auto-detect columns and standardize data (keeping original logic)
        lmrb_theme_col = next((col for col in lmrb_data_for_matching.columns if 'theme' in col.lower() or 'advt_theme' in col.lower() or 'aadvt' in col.lower() or 'Product Details' in col), None)
        lmrb_program_col = next((col for col in lmrb_data_for_matching.columns if 'program' in col.lower()), None)
        lmrb_time_col = next((col for col in lmrb_data_for_matching.columns if 'time' in col.lower() or 'advt_time' in col.lower() or 'air' in col.lower()), None)
        lmrb_date_col = next((col for col in lmrb_data_for_matching.columns if 'date' in col.lower()), None)

        self.matching_status_var.set("Standardizing LMRB data...")
        self.progress_var.set(20)
        self.root.update_idletasks()
        # ✅ CRITICAL: Use the filtered LMRB data for standardization
        self.lmrb_filtered = standardize_dataframe(
            lmrb_data_for_matching.copy().reset_index(drop=True),
            theme_col=lmrb_theme_col,
            program_col=lmrb_program_col,
            time_col=lmrb_time_col,
            date_col=lmrb_date_col
        )

        # TC standardization remains the same
        tc_theme_col = next((col for col in self.tc_data.columns if 'theme' in col.lower() or 'aadvt' in col.lower()), None)
        tc_program_col = next((col for col in self.tc_data.columns if 'program' in col.lower()), None)
        tc_time_col = next((col for col in self.tc_data.columns if 'time' in col.lower() or 'spot' in col.lower() or 'air' in col.lower()), None)
        tc_date_col = next((col for col in self.tc_data.columns if 'date' in col.lower()), None)

        self.matching_status_var.set("Standardizing TC data...")
        self.progress_var.set(40)
        self.root.update_idletasks()
        self.tc_data = standardize_dataframe(
            self.tc_data.copy().reset_index(drop=True),
            theme_col=tc_theme_col,
            program_col=tc_program_col,
            time_col=tc_time_col,
            date_col=tc_date_col
        )

        # Perform matching
        self.matching_status_var.set("Performing matching...")
        self.progress_var.set(60)
        self.root.update_idletasks()

        matched_df, matched_indices, program_mismatched_df, matched_tc_indices = match_media_watch_with_tc(
            self.lmrb_filtered,
            self.tc_data,
            self.theme_mapping,
            ignore_date=self.ignore_date_var.get(),
            time_tolerance=self.time_tolerance_var.get()
        )

        # FIXED: Properly combine existing and new matches
        self.matching_status_var.set("Combining matches...")
        self.progress_var.set(80)
        self.root.update_idletasks()

        # Convert new matches to list format
        if isinstance(matched_df, pd.DataFrame):
            new_matches = matched_df.to_dict('records')
        else:
            new_matches = matched_df if isinstance(matched_df, list) else []

        # Add timestamps to new matches
        for match in new_matches:
            match['Match_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if 'Match_Type' not in match:
                match['Match_Type'] = 'Automatic'

        # FIXED: Combine matches properly
        if preserve_existing and existing_matches:
            # Combine existing and new matches
            final_matches = existing_matches + new_matches
            print(f"✅ Combined: {len(existing_matches)} existing + {len(new_matches)} new = {len(final_matches)} total")
        else:
            final_matches = new_matches

        # Process and store results
        self.process_matching_results(final_matches, matched_indices, program_mismatched_df, matched_tc_indices)

        # Update status
        final_count = len(final_matches)
        product_summary = f" ({len(self.selected_products)} products)" if self.selected_products else " (all products)"
        self.matching_status_var.set(f"Matching completed! Found {final_count} matches{product_summary}")
        self.status_var.set("Matching process completed")
        self.progress_var.set(100)

        # Enable results tab
        self.notebook.tab(3, state="normal")
        self.notebook.select(3)


    def continue_to_results(self):
        """Switch to results tab"""
        if self.matched_data is None:
            messagebox.showerror("Error", "Please run the matching process first")
            return
        
        self.notebook.select(3)
        self.status_var.set("Viewing results")

    def export_matched_data(self):
        """Export matched data to Excel file"""
        try:
            if self.matched_data is None:
                messagebox.showerror("Error", "No matched data available")
                return
            
            default_name = f"matched_data_{self.selected_channel or 'channel'}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if save_path:
                # Create Excel file with matched data
                matched_df = pd.DataFrame(self.matched_data)
                matched_df.to_excel(save_path, index=False)
                messagebox.showinfo("Success", f"Matched data exported to {save_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting matched data: {str(e)}")
            
    def export_unmatched_data(self):
        """Export unmatched TC data to Excel file - ONLY DISPLAY COLUMNS"""
        try:
            if self.unmatched_tc_data is None:
                messagebox.showerror("Error", "No unmatched data available")
                return
            
            default_name = f"Unmatched_TC_Data_{self.selected_channel or 'channel'}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )

            if save_path:
                # Define display columns only
                display_columns = ['Date', 'Program', 'Advt_time', 'Advt_theme', 'Duration']
                
                # Filter to only export display columns that exist
                available_display_columns = [col for col in display_columns if col in self.unmatched_tc_data.columns]
                
                if available_display_columns:
                    # Export only display columns
                    export_df = self.unmatched_tc_data[available_display_columns]
                    print(f"Exporting unmatched TC data with columns: {available_display_columns}")
                else:
                    # Fallback if display columns don't exist
                    export_df = self.unmatched_tc_data
                    print("Warning: Display columns not found, exporting all columns")
                
                # Add metadata sheet with summary information
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Write main data (display columns only)
                    export_df.to_excel(writer, sheet_name='Unmatched TC Data', index=False)
                    
                    # Create summary sheet
                    summary_data = {
                        'Metric': [
                            'Total TC Records',
                            'Matched Records',
                            'Unmatched Records',
                            'Columns Displayed',
                            'Export Date',
                            'Channel'
                        ],
                        'Value': [
                            len(self.tc_data_original) if self.tc_data_original is not None else 'N/A',
                            len(self.matched_data) if self.matched_data is not None else 0,
                            len(self.unmatched_tc_data),
                            ', '.join(available_display_columns),
                            pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S'),
                            self.selected_channel or 'Unknown'
                        ]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                    
                    # Add note about backend columns
                    notes_data = {
                        'Note': [
                            'Display Columns Only',
                            'Backend Processing Columns',
                            'Explanation'
                        ],
                        'Details': [
                            f"This export shows only user-facing columns: {', '.join(display_columns)}",
                            "Backend columns (Advt_Theme, Normalized_Theme, Advt_Time, Time_Seconds, Dd, Mn, Yr) are kept for processing",
                            "This provides a cleaner view while maintaining full functionality"
                        ]
                    }
                    pd.DataFrame(notes_data).to_excel(writer, sheet_name='Export Notes', index=False)
                
                messagebox.showinfo("Success", 
                    f"Unmatched TC data exported to {save_path}\n"
                    f"Total unmatched records: {len(self.unmatched_tc_data)}\n"
                    f"Columns exported: {', '.join(available_display_columns)}")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting unmatched data: {str(e)}")
            print(f"Export error details: {str(e)}")

    def generate_pdf_report(self):
        """Generate PDF summary report"""
        try:
            if not self.matched_data:
                messagebox.showerror("Error", "No matching results available")
                return
                
            # Import and call the PDF creation function
            from func.create_pdf import create_summary_pdf_tables_only
            filename = create_summary_pdf_tables_only(self)
            
            if filename:
                messagebox.showinfo("Success", f"PDF report generated at {filename}")
                
                # Try to open the PDF
                try:
                    import os
                    import platform
                    import subprocess
                    
                    if platform.system() == 'Darwin':       # macOS
                        subprocess.call(('open', filename))
                    elif platform.system() == 'Windows':    # Windows
                        os.startfile(filename)
                    else:                                   # Linux variants
                        subprocess.call(('xdg-open', filename))
                except:
                    pass
            else:
                messagebox.showerror("Error", "Failed to create PDF report")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error generating PDF report: {str(e)}")
            import traceback
            traceback.print_exc()


    def export_detailed_excel_report(self):
        """Export a detailed Excel report with all sheets - Fixed Version"""
        try:
            # Check if we have the required data
            if not self.matched_data:
                messagebox.showerror("Error", "No matched data available for export")
                return
                
            # Get save path
            default_name = f"detailed_report_{self.selected_channel or 'channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not save_path:
                return
                
            print("Starting detailed Excel export...")  # Debug
            
            # Prepare data with safe conversion
            matched_df = self.safe_convert_to_dataframe(self.matched_data, "Matched Data")
            unmatched_tc_df = self.safe_convert_to_dataframe(self.unmatched_tc_data, "Unmatched TC Data")
            unmatched_lmrb_df = self.safe_convert_to_dataframe(self.unmatched_lmrb_data, "Unmatched LMRB Data")
            program_mismatched_df = self.safe_convert_to_dataframe(
                getattr(self, 'program_mismatched_data', None), "Program Mismatched Data"
            )
            lmrb_filtered_df = self.safe_convert_to_dataframe(self.lmrb_filtered, "LMRB Filtered Data")
            
            # Create the Excel report using our safe method
            self.create_detailed_excel_report_safe(
                save_path, 
                matched_df, 
                unmatched_tc_df, 
                unmatched_lmrb_df, 
                program_mismatched_df, 
                lmrb_filtered_df
            )
            
            messagebox.showinfo("Success", f"Detailed Excel report exported successfully!\n\nFile: {save_path}")
            
        except Exception as e:
            error_msg = f"Error exporting detailed Excel report: {str(e)}"
            print(f"Export error: {error_msg}")  # Debug
            import traceback
            traceback.print_exc()
            messagebox.showerror("Export Error", error_msg)

    def safe_convert_to_dataframe(self, data, data_name):
        """Safely convert data to DataFrame with error handling"""
        try:
            if data is None:
                print(f"{data_name}: No data available")
                return pd.DataFrame()
                
            if isinstance(data, pd.DataFrame):
                print(f"{data_name}: Already DataFrame with {len(data)} rows")
                return data.copy()
                
            if isinstance(data, list):
                if not data:
                    print(f"{data_name}: Empty list")
                    return pd.DataFrame()
                print(f"{data_name}: Converting list with {len(data)} items to DataFrame")
                return pd.DataFrame(data)
                
            if isinstance(data, dict):
                print(f"{data_name}: Converting dict to DataFrame")
                return pd.DataFrame([data])
                
            print(f"{data_name}: Unknown data type {type(data)}, returning empty DataFrame")
            return pd.DataFrame()
            
        except Exception as e:
            print(f"Error converting {data_name} to DataFrame: {e}")
            return pd.DataFrame()

    def create_detailed_excel_report_safe(self, save_path, matched_df, unmatched_tc_df, 
                                        unmatched_lmrb_df, program_mismatched_df, lmrb_filtered_df):
        """Create detailed Excel report with multiple sheets - Safe Version"""
        try:
            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                
                # 1. SUMMARY SHEET
                print("Creating Summary sheet...")
                self.create_summary_sheet(writer, matched_df, unmatched_tc_df, unmatched_lmrb_df, program_mismatched_df)
                
                # 2. MATCHED DATA SHEET
                if not matched_df.empty:
                    print(f"Creating Matched Data sheet with {len(matched_df)} rows...")
                    matched_df.to_excel(writer, sheet_name='Matched Data', index=False)
                else:
                    print("No matched data to export")
                    pd.DataFrame({'Message': ['No matched data available']}).to_excel(
                        writer, sheet_name='Matched Data', index=False
                    )
                
                # 3. UNMATCHED TC DATA SHEET
                if not unmatched_tc_df.empty:
                    print(f"Creating Unmatched TC Data sheet with {len(unmatched_tc_df)} rows...")
                    unmatched_tc_df.to_excel(writer, sheet_name='Unmatched TC Data', index=False)
                else:
                    print("No unmatched TC data to export")
                    pd.DataFrame({'Message': ['No unmatched TC data available']}).to_excel(
                        writer, sheet_name='Unmatched TC Data', index=False
                    )
                
                # 4. UNMATCHED LMRB DATA SHEET
                if not unmatched_lmrb_df.empty:
                    print(f"Creating Unmatched LMRB Data sheet with {len(unmatched_lmrb_df)} rows...")
                    unmatched_lmrb_df.to_excel(writer, sheet_name='Unmatched LMRB Data', index=False)
                else:
                    print("No unmatched LMRB data to export")
                    pd.DataFrame({'Message': ['No unmatched LMRB data available']}).to_excel(
                        writer, sheet_name='Unmatched LMRB Data', index=False
                    )
                
                # 5. PROGRAM MISMATCHED DATA SHEET
                if not program_mismatched_df.empty:
                    print(f"Creating Program Mismatched Data sheet with {len(program_mismatched_df)} rows...")
                    program_mismatched_df.to_excel(writer, sheet_name='Program Mismatches', index=False)
                else:
                    print("No program mismatched data to export")
                    pd.DataFrame({'Message': ['No program mismatched data available']}).to_excel(
                        writer, sheet_name='Program Mismatches', index=False
                    )
                
                # 6. THEME MAPPINGS SHEET
                print("Creating Theme Mappings sheet...")
                self.create_theme_mappings_sheet(writer)
                
                # 7. ALL LMRB DATA SHEET (if available and not too large)
                if not lmrb_filtered_df.empty and len(lmrb_filtered_df) < 10000:  # Limit size
                    print(f"Creating All LMRB Data sheet with {len(lmrb_filtered_df)} rows...")
                    lmrb_filtered_df.to_excel(writer, sheet_name='All LMRB Data', index=False)
                else:
                    print("LMRB data too large or empty, skipping...")
                    pd.DataFrame({'Message': ['LMRB data too large for export or not available']}).to_excel(
                        writer, sheet_name='All LMRB Data', index=False
                    )
                
            print("Excel export completed successfully!")
            
        except Exception as e:
            print(f"Error creating Excel report: {e}")
            import traceback
            traceback.print_exc()
            raise

    def create_summary_sheet(self, writer, matched_df, unmatched_tc_df, unmatched_lmrb_df, program_mismatched_df):
        """Create summary sheet with statistics"""
        try:
            # Calculate statistics
            total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0
            total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0
            total_matched = len(matched_df)
            total_unmatched_tc = len(unmatched_tc_df)
            total_unmatched_lmrb = len(unmatched_lmrb_df)
            total_program_mismatched = len(program_mismatched_df)
            
            match_rate = (total_matched / total_tc * 100) if total_tc > 0 else 0
            
            # Create summary data
            summary_data = {
                'Metric': [
                    'Export Date',
                    'Export Time', 
                    'Channel',
                    'Total LMRB Records',
                    'Total TC Records',
                    'Successfully Matched',
                    'Unmatched TC Records',
                    'Unmatched LMRB Records',
                    'Program Mismatches',
                    'Overall Match Rate (%)',
                    'Theme Mappings Used'
                ],
                'Value': [
                    datetime.now().strftime('%Y-%m-%d'),
                    datetime.now().strftime('%H:%M:%S'),
                    self.selected_channel or 'Unknown',
                    total_lmrb,
                    total_tc,
                    total_matched,
                    total_unmatched_tc,
                    total_unmatched_lmrb,
                    total_program_mismatched,
                    f"{match_rate:.2f}%",
                    len(self.theme_mapping)
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Add matching options used
            options_data = {
                'Setting': [
                    'Ignore Date Matching',
                    'Time Tolerance (seconds)',
                    'Manual Matches Included',
                    'Schedule Data Available'
                ],
                'Value': [
                    'Yes' if self.ignore_date_var.get() else 'No',
                    self.time_tolerance_var.get(),
                    'Yes' if any(match.get('Match_Type') == 'Manual' for match in (self.matched_data or [])) else 'No',
                    'Yes' if hasattr(self, 'schedule_data') and self.schedule_data is not None else 'No'
                ]
            }
            
            # Add some spacing and then the options
            spacing_df = pd.DataFrame({'': ['', 'MATCHING OPTIONS:', '']})
            options_df = pd.DataFrame(options_data)
            
            # Combine and write to same sheet (you might need to adjust this based on openpyxl version)
            with pd.ExcelWriter(writer.path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer2:
                spacing_df.to_excel(writer2, sheet_name='Summary', startrow=len(summary_df) + 2, index=False, header=False)
                options_df.to_excel(writer2, sheet_name='Summary', startrow=len(summary_df) + 5, index=False)
                
        except Exception as e:
            print(f"Error creating summary sheet: {e}")
            # Create basic summary if detailed one fails
            basic_summary = pd.DataFrame({
                'Metric': ['Export Date', 'Total Matches', 'Channel'],
                'Value': [datetime.now().strftime('%Y-%m-%d %H:%M:%S'), len(matched_df), self.selected_channel or 'Unknown']
            })
            basic_summary.to_excel(writer, sheet_name='Summary', index=False)

    def create_theme_mappings_sheet(self, writer):
        """Create theme mappings sheet"""
        try:
            if self.theme_mapping:
                mappings_data = []
                for i, mapping in enumerate(self.theme_mapping, 1):
                    mappings_data.append({
                        'Mapping #': i,
                        'Type': mapping.get('mapping_type', 'Unknown'),
                        'LMRB Theme': mapping.get('media_watch_theme', mapping.get('lmrb_theme', '')),
                        'LMRB Duration': mapping.get('lmrb_duration', ''),
                        'TC Theme': mapping.get('tc_theme', ''),
                        'TC Duration': mapping.get('tc_duration', mapping.get('tc_dur', '')),
                        'Schedule Theme': mapping.get('schedule_theme', ''),
                        'Schedule Duration': mapping.get('schedule_duration', '')
                    })
                
                mappings_df = pd.DataFrame(mappings_data)
                mappings_df.to_excel(writer, sheet_name='Theme Mappings', index=False)
            else:
                pd.DataFrame({'Message': ['No theme mappings available']}).to_excel(
                    writer, sheet_name='Theme Mappings', index=False
                )
                
        except Exception as e:
            print(f"Error creating theme mappings sheet: {e}")
            pd.DataFrame({'Error': [f'Could not create theme mappings: {str(e)}']}).to_excel(
                writer, sheet_name='Theme Mappings', index=False
            )
            
    def export_highlighted_lmrb(self):
        """Export LMRB data with matched rows highlighted."""
        try:
            from func.highlightV1 import highlight_matched_rows_excel  # Use your app.py logic
            if self.lmrb_filtered is None or self.matched_data is None:
                messagebox.showerror("Error", "No LMRB or matched data available")
                return
            
            default_name = f"highlighted_lmrb_{self.selected_channel or 'channel'}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            if save_path:
                matched_lmrb = pd.DataFrame(self.matched_data)
                media_watch_filtered = pd.DataFrame(self.lmrb_filtered)
                wb = highlight_matched_rows_excel(media_watch_filtered, matched_lmrb)
                wb.save(save_path)
                messagebox.showinfo("Success", f"Highlighted LMRB data exported to {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting highlighted LMRB data: {str(e)}")

    def export_unmatched_lmrb_data(self):
        """Export unmatched LMRB data to Excel file"""
        try:
            if self.unmatched_lmrb_data is None or self.unmatched_lmrb_data.empty:
                messagebox.showerror("Error", "No unmatched LMRB data available")
                return
            
            default_name = f"Unmatched_LMRB_Data_{self.selected_channel or 'channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )

            if save_path:
                # Create Excel file with unmatched LMRB data
                unmatched_lmrb_df = pd.DataFrame(self.unmatched_lmrb_data)
                
                # Add metadata sheet with summary information
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Write main unmatched LMRB data
                    unmatched_lmrb_df.to_excel(writer, sheet_name='Unmatched LMRB Data', index=False)
                    
                    # Create summary sheet
                    total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0
                    total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0
                    matched_count = len(self.matched_data) if self.matched_data is not None else 0
                    unmatched_lmrb_count = len(self.unmatched_lmrb_data)
                    match_rate = (matched_count / total_tc * 100) if total_tc > 0 else 0

                    summary_data = {
                        'Metric': [
                            'Channel',
                            'Total LMRB Records',
                            'Matched LMRB Records',
                            'Unmatched LMRB Records',
                            'Match Rate (%)',
                            'Export Date',
                            'Export Time'
                        ],
                        'Value': [
                            self.selected_channel or 'Unknown',
                            total_lmrb,
                            matched_count,
                            unmatched_lmrb_count,
                            f"{match_rate:.2f}%",
                            pd.Timestamp.now().strftime('%Y-%m-%d'),
                            pd.Timestamp.now().strftime('%H:%M:%S')
                        ]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                    
                    # If theme mapping exists, add it to the export
                    if self.theme_mapping:
                        theme_mapping_data = []
                        for mapping in self.theme_mapping:
                            theme_mapping_data.append({
                                'LMRB Theme': mapping.get('media_watch_theme', mapping.get('lmrb_theme', '')),
                                'LMRB Duration': mapping.get('lmrb_duration', ''),
                                'TC Theme': mapping.get('tc_theme', ''),
                                'TC Duration': mapping.get('tc_duration', mapping.get('tc_dur', ''))
                            })
                        
                        if theme_mapping_data:
                            pd.DataFrame(theme_mapping_data).to_excel(writer, sheet_name='Theme Mappings', index=False)
                
                messagebox.showinfo("Success", 
                    f"Unmatched LMRB data exported successfully!\n\n"
                    f"File: {save_path}\n"
                    f"Total unmatched LMRB records: {len(self.unmatched_lmrb_data)}\n"
                    f"Channel: {self.selected_channel or 'Unknown'}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting unmatched LMRB data: {str(e)}")
            print(f"Export error details: {str(e)}")
            import traceback
            traceback.print_exc()


    def process_matching_results(self, matched_df, matched_indices, program_mismatched_df, matched_tc_indices):
        """Process and store matching results."""
        self.progress_var.set(80)
        self.matching_status_var.set("Processing results...")
        self.root.update_idletasks()

        try:
            # Store matched data
            self.matched_data = matched_df
            self.matched_indices = matched_indices
            self.program_mismatched_data = program_mismatched_df

            # Find unmatched TC entries using standardized data
            if hasattr(self, 'tc_data') and self.tc_data is not None:
                # Use the standardized TC data instead of original
                self.unmatched_tc_data = self.tc_data.drop(matched_tc_indices).reset_index(drop=True)
                # Clean dates in unmatched TC data
                self.unmatched_tc_data = self.clean_dataframe_dates(self.unmatched_tc_data)
            else:
                self.unmatched_tc_data = None

            # Find unmatched LMRB entries
            if hasattr(self, 'lmrb_filtered') and self.lmrb_filtered is not None:
                # Create a list of indices to drop
                unmatched_indices = list(set(range(len(self.lmrb_filtered))) - set(matched_indices))
                
                # Drop matched indices to get unmatched data
                self.unmatched_lmrb_data = self.lmrb_filtered.iloc[unmatched_indices].reset_index(drop=True)
                # Clean dates in unmatched LMRB data  
                self.unmatched_lmrb_data = self.clean_dataframe_dates(self.unmatched_lmrb_data)
            else:
                self.unmatched_lmrb_data = None

            if (hasattr(self, 'schedule_data') and 
                self.schedule_data is not None and 
                not self.schedule_data.empty and
                self.theme_mapping):
                
                # Enable the summary report tab (adjust tab number based on your setup)
                try:
                    # Find the summary report tab index
                    for i in range(self.notebook.index("end")):
                        if self.notebook.tab(i, "text") == "Summary Report":
                            self.notebook.tab(i, state="normal")
                            print("Summary Report tab enabled")
                            break
                except:
                    pass
                
                # Sync summary report data
                self.sync_summary_report_data()


            # Ensure Date column exists in unmatched data
            def create_date_column(df):
                """Create a Date column if it doesn't exist."""
                if df is not None and not df.empty:
                    # Check if Date column already exists
                    if 'Date' not in df.columns:
                        # Try to create Date column from Dd, Mn, Yr
                        if all(col in df.columns for col in ['Dd', 'Mn', 'Yr']):
                            try:
                                df['Date'] = df.apply(
                                    lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                                    axis=1
                                )
                            except Exception as e:
                                print(f"Error creating Date column: {e}")
                                # Fallback to using existing columns if possible
                                if 'DATE' in df.columns:
                                    df['Date'] = df['DATE']
                                else:
                                    # If no date columns are available, add a placeholder
                                    df['Date'] = 'Unknown'
                return df

            # Create Date columns for unmatched data
            self.unmatched_tc_data = create_date_column(self.unmatched_tc_data)
            self.unmatched_lmrb_data = create_date_column(self.unmatched_lmrb_data)

            # Update progress
            self.progress_var.set(90)
            self.matching_status_var.set("Updating results display...")
            self.root.update_idletasks()

            # Update the results tab
            self.update_results_tab()
            
            # Enable and switch to results tab
            self.notebook.tab(3, state="normal")
            self.notebook.select(3)

            # ENABLE THE MANUAL MATCHING TAB (always available after matching)
            self.notebook.tab(4, state="normal")  # Enable Manual Matching tab
            print("Manual matching tab enabled")

            # ENABLE SCHEDULE MATCHING TAB ONLY if schedule data is available AND has schedule mappings
            schedule_tab_enabled = False
            if (hasattr(self, 'schedule_data') and 
                self.schedule_data is not None and 
                not self.schedule_data.empty):
                
                # Check if we have any schedule mappings
                has_schedule_mappings = any(
                    mapping.get('schedule_theme', '').strip() 
                    for mapping in self.theme_mapping
                )
                
                if has_schedule_mappings:
                    self.notebook.tab(5, state="normal")  # Enable Schedule Matching tab
                    schedule_tab_enabled = True
                    print("Schedule matching tab enabled - schedule data and mappings are available")
                else:
                    print("Schedule matching tab not enabled - no schedule mappings found")
                    print("Tip: Add schedule themes in Theme Mapping to enable schedule matching")
            else:
                print("Schedule matching tab not enabled - no schedule data available")
                print("Note: Schedule matching is optional. You can continue without it.")

            # Populate manual matching data
            try:
                self.populate_manual_matching_data()
                print("Manual matching data populated successfully")
            except Exception as e:
                print(f"Error populating manual matching data: {e}")
                # Don't fail the entire process if manual matching data fails
                messagebox.showwarning("Warning", 
                    f"Manual matching data could not be populated: {e}\n"
                    "You can still view results and export data.")

            # Update progress to complete
            self.progress_var.set(100)
            
            # Calculate and display final statistics
            total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0
            total_matches = len(self.matched_data) if self.matched_data is not None else 0
            total_unmatched_tc = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            total_unmatched_lmrb = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            total_program_mismatches = len(program_mismatched_df) if program_mismatched_df is not None else 0
            total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0

            # Calculate match rate
            match_rate = (total_matches / total_tc * 100) if total_tc > 0 else 0

            # Update status message
            self.matching_status_var.set(
                f"Matching completed! Found {total_matches} matches, "
                f"{total_program_mismatches} program mismatches, "
                f"{total_unmatched_tc} unmatched TC, "
                f"{total_unmatched_lmrb} unmatched LMRB"
            )

            # Log final statistics
            print("\n" + "="*60)
            print("MATCHING RESULTS SUMMARY")
            print("="*60)
            print(f"Channel: {self.selected_channel or 'Unknown'}")
            print(f"Total LMRB records processed: {total_lmrb}")
            print(f"Successfully matched: {total_matches}")
            print(f"Program mismatches: {total_program_mismatches}")
            print(f"Unmatched TC records: {total_unmatched_tc}")
            print(f"Unmatched LMRB records: {total_unmatched_lmrb}")
            print(f"Match rate: {match_rate:.2f}%")
            print(f"Schedule matching available: {'Yes' if schedule_tab_enabled else 'No'}")
            print("="*60)

            # Build completion message with detailed statistics
            completion_message = (
                f"Matching process completed successfully!\n\n"
                f"📊 Results Summary:\n"
                f"• Channel: {self.selected_channel or 'Unknown'}\n"
                f"• Total LMRB records: {total_lmrb:,}\n"
                f"• Successfully matched: {total_matches:,}\n"
                f"• Program mismatches: {total_program_mismatches:,}\n"
                f"• Unmatched TC records: {total_unmatched_tc:,}\n"
                f"• Unmatched LMRB records: {total_unmatched_lmrb:,}\n"
                f"• Match rate: {match_rate:.2f}%\n\n"
            )
            
            # Add next steps
            completion_message += "🔄 Next Steps:\n"
            completion_message += "• Review detailed results in the Results tab\n"
            completion_message += "• Export matched/unmatched data using export buttons\n"
            
            if total_unmatched_tc > 0 or total_unmatched_lmrb > 0:
                completion_message += "• Use Manual Matching tab to handle unmatched records\n"
            
            if schedule_tab_enabled:
                completion_message += "• Use Schedule Matching tab to compare with planned schedule\n"
            else:
                completion_message += "• Schedule Matching is available if you upload schedule data\n"
            
            completion_message += "• Generate reports using PDF/Excel export options"

            # Show completion message
            messagebox.showinfo("🎉 Matching Complete", completion_message)

            # Log theme mapping statistics
            print(f"\nTheme Mapping Statistics:")
            print(f"Total mappings: {len(self.theme_mapping)}")
            full_mappings = sum(1 for m in self.theme_mapping if m.get('tc_theme', '').strip())
            sponsorship_mappings = len(self.theme_mapping) - full_mappings
            print(f"Full mappings (TC+LMRB+Schedule): {full_mappings}")
            print(f"Sponsorship mappings (LMRB+Schedule only): {sponsorship_mappings}")

        except Exception as e:
            print(f"Error processing matching results: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Reset progress on error
            self.progress_var.set(0)
            self.matching_status_var.set("Error processing results")
            
            # Show detailed error message
            error_message = (
                f"❌ An error occurred while processing matching results:\n\n"
                f"Error Details:\n{str(e)}\n\n"
                f"🔍 Troubleshooting Steps:\n"
                f"• Check that all uploaded data has the correct format\n"
                f"• Verify theme mappings are properly configured\n"
                f"• Ensure required columns (Theme, Program, Time, Date) exist\n"
                f"• Try running the matching process again\n\n"
                f"If the problem persists, check the console for detailed error logs."
            )
            
            messagebox.showerror("Processing Error", error_message)
            
            # Log detailed error information for debugging
            print("\n" + "="*60)
            print("DETAILED ERROR INFORMATION")
            print("="*60)
            print(f"Error type: {type(e).__name__}")
            print(f"Error message: {str(e)}")
            print(f"Channel: {getattr(self, 'selected_channel', 'Not set')}")
            
            # Data availability check
            print(f"\nData Availability:")
            print(f"• LMRB filtered data: {self.lmrb_filtered is not None}")
            if self.lmrb_filtered is not None:
                print(f"  - Shape: {self.lmrb_filtered.shape}")
                print(f"  - Columns: {list(self.lmrb_filtered.columns)}")
            
            print(f"• TC data: {hasattr(self, 'tc_data') and self.tc_data is not None}")
            if hasattr(self, 'tc_data') and self.tc_data is not None:
                print(f"  - Shape: {self.tc_data.shape}")
                print(f"  - Columns: {list(self.tc_data.columns)}")
            
            print(f"• Schedule data: {hasattr(self, 'schedule_data') and self.schedule_data is not None}")
            if hasattr(self, 'schedule_data') and self.schedule_data is not None:
                print(f"  - Shape: {self.schedule_data.shape}")
                print(f"  - Columns: {list(self.schedule_data.columns)}")
            
            # Matching results check
            print(f"\nMatching Results:")
            print(f"• Matched data type: {type(matched_df)}")
            if hasattr(matched_df, 'shape'):
                print(f"  - Shape: {matched_df.shape}")
            print(f"• Matched indices count: {len(matched_indices) if matched_indices else 0}")
            print(f"• TC indices count: {len(matched_tc_indices) if matched_tc_indices else 0}")
            print(f"• Program mismatches count: {len(program_mismatched_df) if program_mismatched_df is not None else 0}")
            
            # Theme mapping check
            print(f"\nTheme Mapping:")
            print(f"• Total mappings: {len(self.theme_mapping)}")
            for i, mapping in enumerate(self.theme_mapping[:3]):  # Show first 3 mappings
                print(f"  - Mapping {i+1}: {mapping}")
            if len(self.theme_mapping) > 3:
                print(f"  - ... and {len(self.theme_mapping) - 3} more mappings")
            
            print("="*60)

        finally:
            # Always update the status bar, even on error
            try:
                total_matches = len(self.matched_data) if self.matched_data is not None else 0
                if total_matches > 0:
                    self.status_var.set(f"Matching completed - {total_matches:,} matches found")
                else:
                    self.status_var.set("Matching process completed")
            except:
                self.status_var.set("Matching process finished")
                
            # Ensure progress bar is at 100% or 0% depending on success
            try:
                if self.matched_data is not None:
                    self.progress_var.set(100)
                else:
                    self.progress_var.set(0)
            except:
                pass
    
    def update_results_tab(self):
        """FIXED: Update results tab ensuring all matches are displayed."""
        # Update statistics
        total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0

        # FIXED: Ensure matched_data count is accurate
        if isinstance(self.matched_data, list):
            total_matches = len([m for m in self.matched_data if m])  # Count non-empty matches
        elif isinstance(self.matched_data, pd.DataFrame):
            total_matches = len(self.matched_data)
        else:
            total_matches = 0

        total_unmatched_tc = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
        total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0
        total_unmatched_lmrb = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
        match_rate = (total_matches / total_tc * 100) if total_tc > 0 else 0

        # Update summary variables
        self.total_lmrb_var.set(f"Total LMRB Records: {total_lmrb}")
        self.matched_count_var.set(f"Matched Records: {total_matches}")
        self.unmatched_tc_var.set(f"Unmatched TC Records: {total_unmatched_tc}")
        self.unmatched_lmrb_var.set(f"Unmatched LMRB Records: {total_unmatched_lmrb}")
        self.match_rate_var.set(f"Match Rate: {match_rate:.1f}%")

        # FIXED: Update matched data treeview ensuring all matches are shown
        if hasattr(self, 'matched_tree'):
            if self.matched_data and total_matches > 0:
                # Convert to DataFrame for display
                if isinstance(self.matched_data, list):
                    matched_df = pd.DataFrame([m for m in self.matched_data if m]).fillna("Not Available")
                else:
                    matched_df = self.matched_data.fillna("Not Available")
                
                print(f"📊 Displaying {len(matched_df)} matched records in results tree")
                self.update_treeview(self.matched_tree, matched_df)
            else:
                self.update_treeview(self.matched_tree, pd.DataFrame())

        # Update unmatched data (keeping original logic)
        if hasattr(self, 'unmatched_tree'):
            if self.unmatched_tc_data is not None and len(self.unmatched_tc_data) > 0:
                display_columns = ['Date', 'Program', 'Advt_time', 'Advt_theme', 'Duration']
                available_display_columns = [col for col in display_columns if col in self.unmatched_tc_data.columns]
                
                if available_display_columns:
                    display_df = self.unmatched_tc_data[available_display_columns]
                else:
                    display_df = self.unmatched_tc_data
                
                self.update_treeview(self.unmatched_tree, display_df)
            else:
                self.update_treeview(self.unmatched_tree, pd.DataFrame())

        if hasattr(self, 'unmatched_lmrb_tree'):
            if self.unmatched_lmrb_data is not None and len(self.unmatched_lmrb_data) > 0:
                self.update_treeview(self.unmatched_lmrb_tree, self.unmatched_lmrb_data)
            else:
                self.update_treeview(self.unmatched_lmrb_tree, pd.DataFrame())


    def setup_manual_matching_tab(self):
        """Setup the Manual Matching tab with comprehensive matching interface."""
        # Main Frame
        main_frame = ttk.Frame(self.manual_matching_tab)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Comparison Panel
        self.comparison_frame = ttk.LabelFrame(main_frame, text="Matching Comparison")
        self.comparison_frame.pack(fill=tk.X, pady=5)

        # TC Comparison Label
        self.tc_comparison_label = ttk.Label(self.comparison_frame, text="No TC Record Selected")
        self.tc_comparison_label.pack(side=tk.LEFT, padx=5, expand=True)

        # LMRB Comparison Label
        self.lmrb_comparison_label = ttk.Label(self.comparison_frame, text="No LMRB Record Selected")
        self.lmrb_comparison_label.pack(side=tk.LEFT, padx=5, expand=True)

        # Match Button
        self.match_button = ttk.Button(
            self.comparison_frame, 
            text="Confirm Match", 
            command=self.confirm_manual_match,
            state=tk.DISABLED
        )
        self.match_button.pack(side=tk.RIGHT, padx=5)

        # Create Notebook for different data sources
        data_notebook = ttk.Notebook(main_frame)
        data_notebook.pack(fill=tk.BOTH, expand=True)

        # TC Data Matching Frame
        tc_matching_frame = ttk.Frame(data_notebook)
        data_notebook.add(tc_matching_frame, text="TC Data Matching")

        # LMRB Data Matching Frame
        lmrb_matching_frame = ttk.Frame(data_notebook)
        data_notebook.add(lmrb_matching_frame, text="LMRB Data Matching")

        # Setup TC Matching Section
        self.setup_tc_matching_section(tc_matching_frame)

        # Setup LMRB Matching Section
        self.setup_lmrb_matching_section(lmrb_matching_frame)

        # Recently Matched Records Section
        self.setup_recent_matches_section(main_frame)


    def select_tc_row(self, row_dict):
        """Handle TC row selection."""
        # Store selected row
        self.selected_tc_row = row_dict

        # Update comparison panel using the keys from the dictionary
        # Handle missing program gracefully
        program_info = row_dict.get('Program', 'No Program Info')
        if program_info in ['N/A', '', None, 'nan']:
            program_info = 'No Program Info'

        tc_details = (
            f"TC Theme: {row_dict.get('Theme', 'N/A')}\n"
            f"Date: {row_dict.get('Date', 'N/A')}\n"
            f"Time: {row_dict.get('Time', 'N/A')}\n"
            f"Program: {program_info}\n"
            f"Duration: {row_dict.get('Duration', 'N/A')}"
        )
        self.tc_comparison_label.config(text=tc_details)

        # Highlight similar LMRB rows
        self.highlight_similar_lmrb_rows(row_dict)


    def select_lmrb_row(self, row_dict):
        """Handle LMRB row selection."""
        # Store selected row
        self.selected_lmrb_row = row_dict

        # Handle missing program gracefully
        program_info = row_dict.get('Program', 'No Program Info')
        if program_info in ['N/A', '', None, 'nan']:
            program_info = 'No Program Info'

        # Update comparison panel using the keys from the dictionary
        lmrb_details = (
            f"LMRB Theme: {row_dict.get('Theme', 'N/A')}\n"
            f"Date: {row_dict.get('Date', 'N/A')}\n"
            f"Time: {row_dict.get('Time', 'N/A')}\n"
            f"Program: {program_info}\n"
            f"Duration: {row_dict.get('Duration', 'N/A')}"
        )
        self.lmrb_comparison_label.config(text=lmrb_details)

        # Enable match button if TC row is also selected
        if hasattr(self, 'selected_tc_row'):
            self.match_button.config(state=tk.NORMAL)

    def refresh_all_dependent_tabs(self):
        """Refresh all tabs that depend on matched data."""
        try:
            print("Refreshing all dependent tabs...")
            
            # Update results tab
            self.update_results_tab()
            
            # Clear and reset schedule matching results to force recalculation
            if hasattr(self, 'schedule_summary_results'):
                self.schedule_summary_results = []
                
            # Clear schedule summary tree
            if hasattr(self, 'schedule_summary_tree'):
                for item in self.schedule_summary_tree.get_children():
                    self.schedule_summary_tree.delete(item)
            
            # Update statistics
            self.update_statistics()
            self.sync_summary_report_data()
            
            print("All dependent tabs refreshed successfully")
            
        except Exception as e:
            print(f"Error refreshing dependent tabs: {e}")
            import traceback
            traceback.print_exc()

    def update_statistics(self):
        """Update statistics across all tabs."""
        try:
            total_lmrb = len(self.lmrb_filtered) if self.lmrb_filtered is not None else 0
            
            # Ensure matched_data count is accurate
            total_matches = len(self.matched_data) if self.matched_data is not None else 0
            total_tc = len(self.tc_data_original) if hasattr(self, 'tc_data_original') and self.tc_data_original is not None else 0

            total_unmatched_tc = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            total_unmatched_lmrb = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            match_rate = (total_matches / total_tc * 100) if total_tc > 0 else 0

            # Update all statistics variables
            self.total_lmrb_var.set(f"Total LMRB Records: {total_lmrb}")
            self.matched_count_var.set(f"Matched Records: {total_matches}")
            self.unmatched_tc_var.set(f"Unmatched TC Records: {total_unmatched_tc}")
            self.unmatched_lmrb_var.set(f"Unmatched LMRB Records: {total_unmatched_lmrb}")
            self.match_rate_var.set(f"Match Rate: {match_rate:.1f}%")
            
            print(f"Statistics updated - Matches: {total_matches}, LMRB: {total_lmrb}, Match Rate: {match_rate:.1f}%")
            
        except Exception as e:
            print(f"Error updating statistics: {e}")

    def confirm_manual_match(self):
        """Confirm and process the manual match with proper data synchronization."""
        try:
            # Validate selections
            if not hasattr(self, 'selected_tc_row') or not hasattr(self, 'selected_lmrb_row'):
                messagebox.showerror("Error", "Please select both TC and LMRB records")
                return

            print("=== MANUAL MATCHING DEBUG ===")
            print("Selected TC row:", self.selected_tc_row)
            print("Selected LMRB row:", self.selected_lmrb_row)
            
            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Helper function to normalize dates for comparison
            def normalize_date_for_comparison(date_value):
                """Convert various date formats to a comparable string format"""
                if pd.isna(date_value):
                    return ""
                
                date_str = str(date_value).strip()
                
                # Handle datetime objects or datetime strings
                if ' 00:00:00' in date_str:
                    date_str = date_str.replace(' 00:00:00', '')
                
                # Try to parse and standardize the date
                try:
                    # Try different date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                        try:
                            parsed_date = pd.to_datetime(date_str, format=fmt)
                            return parsed_date.strftime('%Y-%m-%d')
                        except:
                            continue
                    
                    # If specific formats fail, try pandas auto-parsing
                    parsed_date = pd.to_datetime(date_str)
                    return parsed_date.strftime('%Y-%m-%d')
                except:
                    return date_str

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time', 'air_time']
            program_cols = ['Program', 'program', 'Prog','Programme']
            duration_cols = ['Dur', 'Duration', 'Spot_Duration']

            # Store original counts for debugging
            original_tc_count = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            original_lmrb_count = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            print(f"Before removal - TC: {original_tc_count}, LMRB: {original_lmrb_count}")

            # Normalize the selected dates for comparison
            selected_tc_date_normalized = normalize_date_for_comparison(self.selected_tc_row['Date'])
            selected_lmrb_date_normalized = normalize_date_for_comparison(self.selected_lmrb_row['Date'])
            
            print(f"Normalized dates - TC: {selected_tc_date_normalized}, LMRB: {selected_lmrb_date_normalized}")

            # Find and remove TC record
            tc_removed = False
            if self.unmatched_tc_data is not None and not self.unmatched_tc_data.empty:
                tc_theme_col = find_column(self.unmatched_tc_data, theme_cols)
                tc_date_col = find_column(self.unmatched_tc_data, date_cols)
                tc_time_col = find_column(self.unmatched_tc_data, time_cols)
                
                print(f"TC columns found - Theme: {tc_theme_col}, Date: {tc_date_col}, Time: {tc_time_col}")
                
                if tc_theme_col and tc_date_col and tc_time_col:
                    # Create normalized date column for comparison
                    tc_data_normalized_dates = self.unmatched_tc_data[tc_date_col].apply(normalize_date_for_comparison)
                    
                    # Create a more robust matching condition
                    tc_theme_match = (self.unmatched_tc_data[tc_theme_col].astype(str).str.strip() == str(self.selected_tc_row['Theme']).strip())
                    tc_date_match = (tc_data_normalized_dates == selected_tc_date_normalized)
                    tc_time_match = (self.unmatched_tc_data[tc_time_col].astype(str).str.strip() == str(self.selected_tc_row['Time']).strip())
                    
                    tc_mask = tc_theme_match & tc_date_match & tc_time_match
                    
                    matches_found = tc_mask.sum()
                    print(f"TC matching details:")
                    print(f"  Theme matches: {tc_theme_match.sum()}")
                    print(f"  Date matches: {tc_date_match.sum()}")
                    print(f"  Time matches: {tc_time_match.sum()}")
                    print(f"  Total matches: {matches_found}")
                    
                    if matches_found > 0:
                        # Show which record will be removed
                        matching_records = self.unmatched_tc_data[tc_mask]
                        print(f"TC records to be removed:")
                        for idx, row in matching_records.iterrows():
                            print(f"  Row {idx}: {row[tc_theme_col]} | {row[tc_date_col]} | {row[tc_time_col]}")
                        
                        # Keep all rows EXCEPT the matched ones
                        self.unmatched_tc_data = self.unmatched_tc_data[~tc_mask].reset_index(drop=True)
                        tc_removed = True
                        print(f"TC record removed. New count: {len(self.unmatched_tc_data)}")
                    else:
                        print("No TC matches found for removal")

            # Find and remove LMRB record
            lmrb_removed = False
            lmrb_row_index = None
            matched_lmrb_row = None
            
            if self.unmatched_lmrb_data is not None and not self.unmatched_lmrb_data.empty:
                lmrb_theme_col = find_column(self.unmatched_lmrb_data, theme_cols)
                lmrb_date_col = find_column(self.unmatched_lmrb_data, date_cols)
                lmrb_time_col = find_column(self.unmatched_lmrb_data, time_cols)
                
                print(f"LMRB columns found - Theme: {lmrb_theme_col}, Date: {lmrb_date_col}, Time: {lmrb_time_col}")
                
                if lmrb_theme_col and lmrb_date_col and lmrb_time_col:
                    # Create normalized date column for comparison
                    lmrb_data_normalized_dates = self.unmatched_lmrb_data[lmrb_date_col].apply(normalize_date_for_comparison)
                    
                    # Find matching LMRB rows
                    lmrb_theme_match = (self.unmatched_lmrb_data[lmrb_theme_col].astype(str).str.strip() == str(self.selected_lmrb_row['Theme']).strip())
                    lmrb_date_match = (lmrb_data_normalized_dates == selected_lmrb_date_normalized)
                    lmrb_time_match = (self.unmatched_lmrb_data[lmrb_time_col].astype(str).str.strip() == str(self.selected_lmrb_row['Time']).strip())
                    
                    lmrb_mask = lmrb_theme_match & lmrb_date_match & lmrb_time_match
                    
                    matching_lmrb_rows = self.unmatched_lmrb_data[lmrb_mask]
                    print(f"LMRB matching details:")
                    print(f"  Theme matches: {lmrb_theme_match.sum()}")
                    print(f"  Date matches: {lmrb_date_match.sum()}")
                    print(f"  Time matches: {lmrb_time_match.sum()}")
                    print(f"  Total matches: {len(matching_lmrb_rows)}")
                    
                    if not matching_lmrb_rows.empty:
                        # Get the first matching row for the matched data
                        lmrb_row_index = matching_lmrb_rows.index[0]
                        matched_lmrb_row = self.unmatched_lmrb_data.loc[lmrb_row_index].to_dict()
                        
                        print(f"LMRB record to be removed: Index {lmrb_row_index}")
                        print(f"  Theme: {matched_lmrb_row[lmrb_theme_col]}")
                        print(f"  Date: {matched_lmrb_row[lmrb_date_col]}")
                        print(f"  Time: {matched_lmrb_row[lmrb_time_col]}")
                        
                        # Remove from unmatched LMRB data
                        self.unmatched_lmrb_data = self.unmatched_lmrb_data[~lmrb_mask].reset_index(drop=True)
                        lmrb_removed = True
                        print(f"LMRB record removed. New count: {len(self.unmatched_lmrb_data)}")
                    else:
                        print("No LMRB matches found for removal")

            # Proceed only if both records were found and removed
            if not (tc_removed and lmrb_removed):
                missing = []
                if not tc_removed:
                    missing.append("TC")
                if not lmrb_removed:
                    missing.append("LMRB")
                messagebox.showerror("Error", f"Could not find and remove {' and '.join(missing)} record(s)")
                return

            # Add metadata fields to the matched LMRB row
            if matched_lmrb_row:
                matched_lmrb_row['Match_Type'] = 'Manual'
                matched_lmrb_row['Match_Timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                
                # Convert matched_data to list if it's a DataFrame
                if isinstance(self.matched_data, pd.DataFrame):
                    self.matched_data = self.matched_data.to_dict('records')

                # Ensure matched_data is a list
                if not isinstance(self.matched_data, list):
                    self.matched_data = []

                # Add to matched data
                self.matched_data.append(matched_lmrb_row)
                print(f"Added to matched data. Total matched: {len(self.matched_data)}")

            # Store records for undo functionality with normalized dates
            self.last_matched_tc = self.selected_tc_row.copy()
            self.last_matched_tc['Date_Normalized'] = selected_tc_date_normalized
            
            self.last_matched_lmrb = self.selected_lmrb_row.copy()
            self.last_matched_lmrb['Date_Normalized'] = selected_lmrb_date_normalized
            
            self.last_matched_full_lmrb = matched_lmrb_row.copy() if matched_lmrb_row else None

            # Update treeviews
            self.populate_tc_matching_tree()
            self.populate_lmrb_matching_tree()

            # Update recent matches
            self.recent_matches_tree.insert('', 'end', values=(
                self.selected_tc_row['Theme'],
                self.selected_tc_row['Date'],
                self.selected_lmrb_row['Theme'],
                self.selected_lmrb_row['Date'],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ))

            # CRITICAL: Synchronize all tabs after manual match
            self.refresh_all_dependent_tabs()
            self.sync_summary_report_data()

            # Reset selection and comparison labels
            self.tc_comparison_label.config(text="No TC Record Selected")
            self.lmrb_comparison_label.config(text="No LMRB Record Selected")
            self.match_button.config(state=tk.DISABLED)

            # Remove selection attributes
            if hasattr(self, 'selected_tc_row'):
                delattr(self, 'selected_tc_row')
            if hasattr(self, 'selected_lmrb_row'):
                delattr(self, 'selected_lmrb_row')

            # Show success message
            messagebox.showinfo("Success", "Records matched successfully! All tabs synchronized.")
            print("=== MANUAL MATCHING COMPLETED ===")

        except Exception as e:
            messagebox.showerror("Match Error", f"Error in matching: {str(e)}")
            import traceback
            traceback.print_exc()


    def setup_recent_matches_section(self, parent_frame):
        """Setup a section to show recently matched records with undo functionality."""
        # Recent Matches Frame
        recent_matches_frame = ttk.LabelFrame(parent_frame, text="Recent Matches")
        recent_matches_frame.pack(fill=tk.X, pady=5)

        # Treeview for Recent Matches
        self.recent_matches_tree = ttk.Treeview(
            recent_matches_frame, 
            columns=('TC_Theme', 'TC_Date', 'LMRB_Theme', 'LMRB_Date', 'Match_Time'),
            show='headings'
        )

        # Configure columns
        for col in self.recent_matches_tree['columns']:
            self.recent_matches_tree.heading(col, text=col.replace('_', ' '))
            self.recent_matches_tree.column(col, width=100)

        # Add scrollbar
        recent_matches_scrolly = ttk.Scrollbar(
            recent_matches_frame, 
            orient="vertical", 
            command=self.recent_matches_tree.yview
        )
        self.recent_matches_tree.configure(yscrollcommand=recent_matches_scrolly.set)

        # Pack treeview and scrollbar
        recent_matches_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.recent_matches_tree.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Undo Button
        undo_button = ttk.Button(
            recent_matches_frame, 
            text="Undo Last Match", 
            command=self.undo_last_match
        )
        undo_button.pack(side=tk.RIGHT, padx=5, pady=5)

    def undo_last_match(self):
        """Undo the last manual match with proper synchronization."""
        try:
            # Check if there are any recent matches
            recent_matches = self.recent_matches_tree.get_children()
            if not recent_matches:
                messagebox.showinfo("Undo", "No recent matches to undo.")
                return

            # Get the last match
            last_match = recent_matches[-1]
            match_details = self.recent_matches_tree.item(last_match)['values']

            # Remove from recent matches treeview
            self.recent_matches_tree.delete(last_match)

            # Remove from matched data
            if self.matched_data:
                removed_match = self.matched_data.pop()
                print(f"Removed match: {removed_match.get('Match_Type', 'Unknown')} match")

            # Restore the unmatched records
            self.restore_unmatched_record(match_details)
            
            # CRITICAL: Synchronize all tabs after undo
            self.sync_data_across_tabs()

            messagebox.showinfo("Undo", "Last match has been undone. All tabs synchronized.")

        except Exception as e:
            messagebox.showerror("Undo Error", f"Could not undo last match: {str(e)}")
            import traceback
            traceback.print_exc()

    def restore_unmatched_record(self, match_details):
        """Restore unmatched records based on match details."""
        try:
            # Unpack match details
            tc_theme, tc_date, lmrb_theme, lmrb_date, _ = match_details

            print(f"=== RESTORING RECORDS ===")
            print(f"TC to restore: theme='{tc_theme}', date='{tc_date}'")
            print(f"LMRB to restore: theme='{lmrb_theme}', date='{lmrb_date}'")
            
            # Helper function to normalize dates for comparison
            def normalize_date_for_comparison(date_value):
                """Convert various date formats to a comparable string format"""
                if pd.isna(date_value):
                    return ""
                
                date_str = str(date_value).strip()
                
                # Handle datetime objects or datetime strings
                if ' 00:00:00' in date_str:
                    date_str = date_str.replace(' 00:00:00', '')
                
                # Try to parse and standardize the date
                try:
                    # Try different date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                        try:
                            parsed_date = pd.to_datetime(date_str, format=fmt)
                            return parsed_date.strftime('%Y-%m-%d')
                        except:
                            continue
                    
                    # If specific formats fail, try pandas auto-parsing
                    parsed_date = pd.to_datetime(date_str)
                    return parsed_date.strftime('%Y-%m-%d')
                except:
                    return date_str

            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time', 'air_time']
            program_cols = ['Program', 'program', 'Prog']
            duration_cols = ['Dur', 'Duration', 'Spot_Duration']

            # Store original counts
            original_tc_count = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            original_lmrb_count = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            print(f"Before restore - TC: {original_tc_count}, LMRB: {original_lmrb_count}")
            
            # Normalize the dates we're looking for
            tc_date_normalized = normalize_date_for_comparison(tc_date)
            lmrb_date_normalized = normalize_date_for_comparison(lmrb_date)
            
            print(f"Normalized search dates - TC: '{tc_date_normalized}', LMRB: '{lmrb_date_normalized}'")
            
            # Helper function to create Date column if needed
            def ensure_date_column(df):
                if df is not None and not df.empty:
                    date_col = find_column(df, date_cols)
                    if not date_col and all(col in df.columns for col in ['Dd', 'Mn', 'Yr']):
                        df['Date'] = df.apply(
                            lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                            axis=1
                        )
                return df

            # Restore TC record
            tc_restored = False
            if hasattr(self, 'tc_data_original') and self.tc_data_original is not None:
                # Ensure Date column exists
                tc_original = ensure_date_column(self.tc_data_original.copy())
                
                # Find theme and date columns in original TC data
                tc_original_theme_col = find_column(tc_original, theme_cols)
                tc_original_date_col = find_column(tc_original, date_cols)
                
                print(f"TC original columns - Theme: {tc_original_theme_col}, Date: {tc_original_date_col}")
                
                if tc_original_theme_col and tc_original_date_col:
                    # Create normalized date column for original data
                    tc_original_normalized_dates = tc_original[tc_original_date_col].apply(normalize_date_for_comparison)
                    
                    # Find matching TC record with normalized date matching
                    tc_theme_match = (tc_original[tc_original_theme_col].astype(str).str.strip() == str(tc_theme).strip())
                    tc_date_match = (tc_original_normalized_dates == tc_date_normalized)
                    
                    tc_restore_mask = tc_theme_match & tc_date_match
                    tc_restore_rows = tc_original[tc_restore_mask].copy()
                    
                    print(f"TC restore matching:")
                    print(f"  Theme matches: {tc_theme_match.sum()}")
                    print(f"  Date matches: {tc_date_match.sum()}")
                    print(f"  Total restore candidates: {len(tc_restore_rows)}")
                    
                    if not tc_restore_rows.empty:
                        # Show what we're restoring
                        for idx, row in tc_restore_rows.iterrows():
                            print(f"  Restoring TC: {row[tc_original_theme_col]} | {row[tc_original_date_col]}")
                        
                        # Initialize unmatched_tc_data if needed
                        if self.unmatched_tc_data is None or self.unmatched_tc_data.empty:
                            self.unmatched_tc_data = tc_restore_rows.reset_index(drop=True)
                        else:
                            # Concatenate and remove duplicates
                            self.unmatched_tc_data = pd.concat([
                                self.unmatched_tc_data, tc_restore_rows
                            ]).drop_duplicates().reset_index(drop=True)
                        
                        tc_restored = True
                        print(f"TC data restored. New count: {len(self.unmatched_tc_data)}")
                    else:
                        print(f"No TC records found to restore with theme='{tc_theme}' and normalized_date='{tc_date_normalized}'")

            # Restore LMRB record
            lmrb_restored = False
            if hasattr(self, 'lmrb_filtered') and self.lmrb_filtered is not None:
                # Ensure Date column exists
                lmrb_original = ensure_date_column(self.lmrb_filtered.copy())
                
                # Find theme and date columns in LMRB data
                lmrb_original_theme_col = find_column(lmrb_original, theme_cols)
                lmrb_original_date_col = find_column(lmrb_original, date_cols)
                
                print(f"LMRB original columns - Theme: {lmrb_original_theme_col}, Date: {lmrb_original_date_col}")
                
                if lmrb_original_theme_col and lmrb_original_date_col:
                    # Create normalized date column for original data
                    lmrb_original_normalized_dates = lmrb_original[lmrb_original_date_col].apply(normalize_date_for_comparison)
                    
                    # Find matching LMRB record with normalized date matching
                    lmrb_theme_match = (lmrb_original[lmrb_original_theme_col].astype(str).str.strip() == str(lmrb_theme).strip())
                    lmrb_date_match = (lmrb_original_normalized_dates == lmrb_date_normalized)
                    
                    lmrb_restore_mask = lmrb_theme_match & lmrb_date_match
                    lmrb_restore_rows = lmrb_original[lmrb_restore_mask].copy()
                    
                    print(f"LMRB restore matching:")
                    print(f"  Theme matches: {lmrb_theme_match.sum()}")
                    print(f"  Date matches: {lmrb_date_match.sum()}")
                    print(f"  Total restore candidates: {len(lmrb_restore_rows)}")
                    
                    if not lmrb_restore_rows.empty:
                        # Show what we're restoring
                        for idx, row in lmrb_restore_rows.iterrows():
                            print(f"  Restoring LMRB: {row[lmrb_original_theme_col]} | {row[lmrb_original_date_col]}")
                        
                        # Initialize unmatched_lmrb_data if needed
                        if self.unmatched_lmrb_data is None or self.unmatched_lmrb_data.empty:
                            self.unmatched_lmrb_data = lmrb_restore_rows.reset_index(drop=True)
                        else:
                            # Concatenate and remove duplicates
                            self.unmatched_lmrb_data = pd.concat([
                                self.unmatched_lmrb_data, lmrb_restore_rows
                            ]).drop_duplicates().reset_index(drop=True)
                        
                        lmrb_restored = True
                        print(f"LMRB data restored. New count: {len(self.unmatched_lmrb_data)}")
                    else:
                        print(f"No LMRB records found to restore with theme='{lmrb_theme}' and normalized_date='{lmrb_date_normalized}'")

            # Update the trees
            self.populate_tc_matching_tree()
            self.populate_lmrb_matching_tree()
            
            # Update the results tab
            self.update_results_tab()
            
            # Show status
            if tc_restored and lmrb_restored:
                print("Both TC and LMRB records restored successfully")
            elif tc_restored:
                print("Only TC record restored")
            elif lmrb_restored:
                print("Only LMRB record restored")
            else:
                print("No records were restored")
                
            print("=== RESTORE COMPLETED ===")
            
        except Exception as e:
            print(f"Error in restore_unmatched_record: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Restore Error", f"Could not restore unmatched records: {str(e)}")

    def populate_tc_matching_tree(self):
        """Populate TC matching treeview with unmatched TC data."""
        # Clear existing items
        for item in self.tc_matching_tree.get_children():
            self.tc_matching_tree.delete(item)

        # Validate unmatched TC data
        if self.unmatched_tc_data is not None and not self.unmatched_tc_data.empty:
            try:
                print(f"Populating TC matching tree with {len(self.unmatched_tc_data)} rows")
                # Flexible column finding
                theme_col = next((col for col in self.unmatched_tc_data.columns if 'theme' in col.lower()), None)
                date_col = next((col for col in self.unmatched_tc_data.columns if 'date' in col.lower()), None)
                time_col = next((col for col in self.unmatched_tc_data.columns if ('time' in col.lower()) and ('prog' not in col.lower())), None)
                program_col = next((col for col in self.unmatched_tc_data.columns if 'program' in col.lower()), None)
                duration_col = next((col for col in self.unmatched_tc_data.columns if 'dur' in col.lower() or 'duration' in col.lower()), None)
                
                print(f"TC Tree using cols - Theme:{theme_col}, Date:{date_col}, Time:{time_col}, Program:{program_col}, Duration:{duration_col}")
                
                # Iterate through records and add to tree
                for idx, row in self.unmatched_tc_data.iterrows():
                    # Handle missing program data gracefully
                    program_value = "No Program Info"
                    if program_col and not pd.isna(row[program_col]):
                        program_value = str(row[program_col])
                        if program_value.strip() == '' or program_value.lower() == 'nan':
                            program_value = "No Program Info"
                    
                    values = (
                        str(row[theme_col]) if theme_col and not pd.isna(row[theme_col]) else "N/A",
                        str(row[date_col]) if date_col and not pd.isna(row[date_col]) else "N/A",
                        str(row[time_col]) if time_col and not pd.isna(row[time_col]) else "N/A",
                        program_value,
                        str(row[duration_col]) if duration_col and not pd.isna(row[duration_col]) else "N/A"
                    )
                    self.tc_matching_tree.insert('', 'end', values=values)
                        
            except Exception as e:
                print(f"Error populating TC tree: {str(e)}")
                import traceback
                traceback.print_exc()
        else:
            print("No unmatched TC data to display")


    def populate_lmrb_matching_tree(self):
        """Populate LMRB matching treeview with unmatched LMRB data."""
        # Clear existing items
        for item in self.lmrb_matching_tree.get_children():
            self.lmrb_matching_tree.delete(item)

        if self.unmatched_lmrb_data is not None and not self.unmatched_lmrb_data.empty:
            try:
                # Define possible column names (in order of preference)
                theme_cols = ['Advt_Theme', 'Theme', 'Product Details', 'Advt_theme']
                date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
                time_cols = ['Advt_time', 'Time', 'air_time', 'Advt_Time']
                program_cols = ['Program', 'program', 'Prog']
                duration_cols = ['Dur', 'Duration', 'Spot_Duration']
                
                # Find best matching columns
                def find_best_column(df, possible_names):
                    for name in possible_names:
                        if name in df.columns:
                            return name
                    # Fallback: case-insensitive search
                    for col in df.columns:
                        for name in possible_names:
                            if col.lower() == name.lower():
                                return col
                    return None
                
                theme_col = find_best_column(self.unmatched_lmrb_data, theme_cols)
                date_col = find_best_column(self.unmatched_lmrb_data, date_cols)
                time_col = find_best_column(self.unmatched_lmrb_data, time_cols)
                program_col = find_best_column(self.unmatched_lmrb_data, program_cols)
                duration_col = find_best_column(self.unmatched_lmrb_data, duration_cols)
                
                print(f"LMRB Tree columns - Theme:{theme_col}, Date:{date_col}, Time:{time_col}, Program:{program_col}, Duration:{duration_col}")
                
                # Populate treeview with consistent column order
                for idx, row in self.unmatched_lmrb_data.iterrows():
                    values = (
                        str(row[theme_col]) if theme_col and not pd.isna(row[theme_col]) else "N/A",
                        str(row[date_col]) if date_col and not pd.isna(row[date_col]) else "N/A", 
                        str(row[time_col]) if time_col and not pd.isna(row[time_col]) else "N/A",
                        str(row[program_col]) if program_col and not pd.isna(row[program_col]) else "N/A",
                        str(row[duration_col]) if duration_col and not pd.isna(row[duration_col]) else "N/A"
                    )
                    self.lmrb_matching_tree.insert('', 'end', values=values)
                        
            except Exception as e:
                print(f"Error populating LMRB tree: {str(e)}")
                import traceback
                traceback.print_exc()
        else:
            print("No unmatched LMRB data to display")

    def setup_tc_matching_section(self, parent_frame):
        """Setup TC data matching section."""
        # Create filter variables if not already exists
        if not hasattr(self, 'tc_match_theme_var'):
            self.tc_match_theme_var = tk.StringVar()
        if not hasattr(self, 'tc_match_date_var'):
            self.tc_match_date_var = tk.StringVar()
        if not hasattr(self, 'tc_match_program_var'):
            self.tc_match_program_var = tk.StringVar()

        # Main container
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill=tk.BOTH, expand=True)

        # Filter Frame
        filter_frame = ttk.LabelFrame(main_container, text="TC Data Filters")
        filter_frame.pack(fill=tk.X, padx=5, pady=5)

        # Theme Filter
        theme_label = ttk.Label(filter_frame, text="Theme:")
        theme_label.pack(side=tk.LEFT, padx=5)

        self.tc_match_theme_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.tc_match_theme_var, 
            state="readonly", 
            width=50
        )
        self.tc_match_theme_combobox.pack(side=tk.LEFT, padx=5)
        self.tc_match_theme_combobox.bind("<<ComboboxSelected>>", self.on_tc_theme_selected)

        # Date Filter
        date_label = ttk.Label(filter_frame, text="Date:")
        date_label.pack(side=tk.LEFT, padx=5)

        self.tc_match_date_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.tc_match_date_var, 
            state="readonly", 
            width=15
        )
        self.tc_match_date_combobox.pack(side=tk.LEFT, padx=5)
        self.tc_match_date_combobox.bind("<<ComboboxSelected>>", self.on_tc_date_selected)

        # Program Filter
        program_label = ttk.Label(filter_frame, text="Program:")
        program_label.pack(side=tk.LEFT, padx=5)

        self.tc_match_program_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.tc_match_program_var, 
            state="readonly", 
            width=20
        )
        self.tc_match_program_combobox.pack(side=tk.LEFT, padx=5)
        self.tc_match_program_combobox.bind("<<ComboboxSelected>>", self.on_tc_program_selected)

        # Filter and Reset Buttons
        filter_button = ttk.Button(
            filter_frame, 
            text="Filter", 
            command=self.filter_tc_matching_data
        )
        filter_button.pack(side=tk.LEFT, padx=5)

        reset_button = ttk.Button(
            filter_frame, 
            text="Reset", 
            command=self.reset_tc_matching_filter
        )
        reset_button.pack(side=tk.LEFT, padx=5)

        # Treeview Frame
        treeview_frame = ttk.Frame(main_container)
        treeview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Treeview Scrollbars
        tc_tree_scrolly = ttk.Scrollbar(treeview_frame, orient="vertical")
        tc_tree_scrollx = ttk.Scrollbar(treeview_frame, orient="horizontal")

        # TC Matching Treeview
        self.tc_matching_tree = ttk.Treeview(
            treeview_frame, 
            columns=('Theme', 'Date', 'Time', 'Program', 'Duration'),
            show='headings',
            selectmode='browse'
        )

        # Configure Treeview Columns
        self.tc_matching_tree.heading('Theme', text='Theme')
        self.tc_matching_tree.heading('Date', text='Date')
        self.tc_matching_tree.heading('Time', text='Time')
        self.tc_matching_tree.heading('Program', text='Program')
        self.tc_matching_tree.heading('Duration', text='Duration')

        # Configure column widths
        self.tc_matching_tree.column('#0', width=0, stretch=tk.NO)
        self.tc_matching_tree.column('Theme', width=150, anchor=tk.W)
        self.tc_matching_tree.column('Date', width=100, anchor=tk.CENTER)
        self.tc_matching_tree.column('Time', width=100, anchor=tk.CENTER)
        self.tc_matching_tree.column('Program', width=150, anchor=tk.W)
        self.tc_matching_tree.column('Duration', width=80, anchor=tk.CENTER)

        # Configure Scrollbars
        tc_tree_scrolly.config(command=self.tc_matching_tree.yview)
        tc_tree_scrollx.config(command=self.tc_matching_tree.xview)
        self.tc_matching_tree.configure(
            yscrollcommand=tc_tree_scrolly.set, 
            xscrollcommand=tc_tree_scrollx.set
        )

        # Pack Treeview and Scrollbars
        tc_tree_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        tc_tree_scrollx.pack(side=tk.BOTTOM, fill=tk.X)
        self.tc_matching_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bind selection event
        self.tc_matching_tree.bind('<<TreeviewSelect>>', self.on_tc_matching_selection)

    def setup_lmrb_matching_section(self, parent_frame):
        """Setup LMRB data matching section."""
        # Create filter variables if not already exists
        if not hasattr(self, 'lmrb_match_theme_var'):
            self.lmrb_match_theme_var = tk.StringVar()
        if not hasattr(self, 'lmrb_match_date_var'):
            self.lmrb_match_date_var = tk.StringVar()
        if not hasattr(self, 'lmrb_match_program_var'):
            self.lmrb_match_program_var = tk.StringVar()
        if not hasattr(self, 'lmrb_match_duration_var'):
            self.lmrb_match_duration_var = tk.StringVar()

        # Main container
        main_container = ttk.Frame(parent_frame)
        main_container.pack(fill=tk.BOTH, expand=True)

        # Filter Frame
        filter_frame = ttk.LabelFrame(main_container, text="LMRB Data Filters")
        filter_frame.pack(fill=tk.X, padx=5, pady=5)

        # Theme Filter
        theme_label = ttk.Label(filter_frame, text="Theme:")
        theme_label.pack(side=tk.LEFT, padx=5)

        self.lmrb_match_theme_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.lmrb_match_theme_var, 
            state="readonly", 
            width=50
        )
        self.lmrb_match_theme_combobox.pack(side=tk.LEFT, padx=5)
        self.lmrb_match_theme_combobox.bind("<<ComboboxSelected>>", self.on_lmrb_theme_selected)

        # Date Filter
        date_label = ttk.Label(filter_frame, text="Date:")
        date_label.pack(side=tk.LEFT, padx=5)

        self.lmrb_match_date_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.lmrb_match_date_var, 
            state="readonly", 
            width=15
        )
        self.lmrb_match_date_combobox.pack(side=tk.LEFT, padx=5)
        self.lmrb_match_date_combobox.bind("<<ComboboxSelected>>", self.on_lmrb_date_selected)

        # Program Filter
        program_label = ttk.Label(filter_frame, text="Program:")
        program_label.pack(side=tk.LEFT, padx=5)

        self.lmrb_match_program_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.lmrb_match_program_var, 
            state="readonly", 
            width=20
        )
        self.lmrb_match_program_combobox.pack(side=tk.LEFT, padx=5)
        self.lmrb_match_program_combobox.bind("<<ComboboxSelected>>", self.on_lmrb_program_selected)

        # Duration Filter
        duration_label = ttk.Label(filter_frame, text="Duration:")
        duration_label.pack(side=tk.LEFT, padx=5)

        self.lmrb_match_duration_combobox = ttk.Combobox(
            filter_frame, 
            textvariable=self.lmrb_match_duration_var, 
            state="readonly", 
            width=10
        )
        self.lmrb_match_duration_combobox.pack(side=tk.LEFT, padx=5)
        self.lmrb_match_duration_combobox.bind("<<ComboboxSelected>>", self.on_lmrb_duration_selected)

        # Filter and Reset Buttons
        filter_button = ttk.Button(
            filter_frame, 
            text="Filter", 
            command=self.filter_lmrb_matching_data
        )
        filter_button.pack(side=tk.LEFT, padx=5)

        reset_button = ttk.Button(
            filter_frame, 
            text="Reset", 
            command=self.reset_lmrb_matching_filter
        )
        reset_button.pack(side=tk.LEFT, padx=5)

        # Treeview Frame
        treeview_frame = ttk.Frame(main_container)
        treeview_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Treeview Scrollbars
        lmrb_tree_scrolly = ttk.Scrollbar(treeview_frame, orient="vertical")
        lmrb_tree_scrollx = ttk.Scrollbar(treeview_frame, orient="horizontal")

        # LMRB Matching Treeview
        self.lmrb_matching_tree = ttk.Treeview(
            treeview_frame, 
            columns=('Theme', 'Date', 'Time', 'Program', 'Duration'),
            show='headings',
            selectmode='browse'
        )

        # Configure Treeview Columns
        self.lmrb_matching_tree.heading('Theme', text='Theme')
        self.lmrb_matching_tree.heading('Date', text='Date')
        self.lmrb_matching_tree.heading('Time', text='Time')
        self.lmrb_matching_tree.heading('Program', text='Program')
        self.lmrb_matching_tree.heading('Duration', text='Duration')

        # Configure column widths
        self.lmrb_matching_tree.column('#0', width=0, stretch=tk.NO)
        self.lmrb_matching_tree.column('Theme', width=150, anchor=tk.W)
        self.lmrb_matching_tree.column('Date', width=100, anchor=tk.CENTER)
        self.lmrb_matching_tree.column('Time', width=100, anchor=tk.CENTER)
        self.lmrb_matching_tree.column('Program', width=150, anchor=tk.W)
        self.lmrb_matching_tree.column('Duration', width=80, anchor=tk.CENTER)

        # Configure Scrollbars
        lmrb_tree_scrolly.config(command=self.lmrb_matching_tree.yview)
        lmrb_tree_scrollx.config(command=self.lmrb_matching_tree.xview)
        self.lmrb_matching_tree.configure(
            yscrollcommand=lmrb_tree_scrolly.set, 
            xscrollcommand=lmrb_tree_scrollx.set
        )

        # Pack Treeview and Scrollbars
        lmrb_tree_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        lmrb_tree_scrollx.pack(side=tk.BOTTOM, fill=tk.X)
        self.lmrb_matching_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Bind selection event
        self.lmrb_matching_tree.bind('<<TreeviewSelect>>', self.on_lmrb_matching_selection)


    def on_tc_tree_click(self, event):
        """Handle TC treeview click event."""
        # Get the clicked item
        item = self.tc_matching_tree.identify_row(event.y)
        if item:
            # Get the row data
            row_data = self.tc_matching_tree.item(item, 'values')
            
            # Create a dictionary representation of the row
            row_dict = {
                'Advt_Theme': row_data[0],
                'Date': row_data[1],
                'Advt_time': row_data[2],
                'Program': row_data[3],
                'Dur': row_data[4]
            }
            
            # Select the row
            self.select_tc_row(row_dict)

    def on_lmrb_tree_click(self, event):
        """Handle LMRB treeview click event."""
        # Get the clicked item
        item = self.lmrb_matching_tree.identify_row(event.y)
        if item:
            # Get the row data
            row_data = self.lmrb_matching_tree.item(item, 'values')
            
            # Create a dictionary representation of the row
            row_dict = {
                'Advt_Theme': row_data[0],
                'Date': row_data[1],
                'Advt_time': row_data[2],
                'Program': row_data[3],
                'Dur': row_data[4]
            }
            
            # Select the row
            self.select_lmrb_row(row_dict)

    def on_tc_theme_selected(self, event=None):
        """Handle TC theme selection for filtering."""
        self.filter_tc_matching_data()

    def on_tc_date_selected(self, event=None):
        """Handle TC date selection for filtering."""
        self.filter_tc_matching_data()

    def on_tc_program_selected(self, event=None):
        """Handle TC program selection for filtering."""
        self.filter_tc_matching_data()

    def on_lmrb_theme_selected(self, event=None):
        """Handle LMRB theme selection for filtering."""
        self.filter_lmrb_matching_data()

    def on_lmrb_date_selected(self, event=None):
        """Handle LMRB date selection for filtering."""
        self.filter_lmrb_matching_data()

    def on_lmrb_program_selected(self, event=None):
        """Handle LMRB program selection for filtering."""
        self.filter_lmrb_matching_data()

    def on_lmrb_duration_selected(self, event=None):
        """Handle LMRB duration selection for filtering."""
        self.filter_lmrb_matching_data()

    def filter_tc_matching_data(self):
        """Filter TC matching data based on selected filters."""
        if self.unmatched_tc_data is None:
            return

        # Get filter values
        theme_filter = self.tc_match_theme_var.get()
        date_filter = self.tc_match_date_var.get()
        program_filter = self.tc_match_program_var.get()

        # Start with full unmatched TC data
        filtered_df = self.unmatched_tc_data.copy()

        # Apply theme filter if selected
        if theme_filter:
            filtered_df = filtered_df[filtered_df['Advt_Theme'] == theme_filter]

        # Apply date filter if selected
        if date_filter:
            filtered_df = filtered_df[filtered_df['Date'] == date_filter]

        # Apply program filter if selected
        if program_filter:
            filtered_df = filtered_df[filtered_df['Program'] == program_filter]

        # Update treeview
        self.update_treeview(self.tc_matching_tree, 
            filtered_df[['Advt_Theme', 'Date', 'Advt_time', 'Program', 'Dur']])

    def filter_tc_matching_data(self):
        """Filter TC matching data based on selected filters."""
        if self.unmatched_tc_data is None:
            return

        # Get filter values
        theme_filter = self.tc_match_theme_var.get()
        date_filter = self.tc_match_date_var.get()
        program_filter = self.tc_match_program_var.get()

        # Flexible column finding function
        def find_column(df, possible_names):
            for name in possible_names:
                matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                if matching_cols:
                    return matching_cols[0]
            return None

        # Define possible column names
        theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
        date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
        time_cols = ['Advt_time', 'Time', 'air_time']
        program_cols = ['Program', 'program', 'Prog']
        duration_cols = ['Dur', 'Duration', 'Spot_Duration']

        # Find appropriate columns
        theme_col = find_column(self.unmatched_tc_data, theme_cols)
        date_col = find_column(self.unmatched_tc_data, date_cols)
        time_col = find_column(self.unmatched_tc_data, time_cols)
        program_col = find_column(self.unmatched_tc_data, program_cols)
        duration_col = find_column(self.unmatched_tc_data, duration_cols)

        # Start with full unmatched TC data
        filtered_df = self.unmatched_tc_data.copy()

        # Apply filters
        try:
            # Theme Filter
            if theme_filter and theme_col:
                filtered_df = filtered_df[filtered_df[theme_col] == theme_filter]

            # Date Filter
            if date_filter and date_col:
                filtered_df = filtered_df[filtered_df[date_col] == date_filter]

            # Program Filter - handle missing program data
            if program_filter and program_col:
                if program_filter == "No Program Info":
                    # Filter for rows with no program info
                    program_mask = (
                        filtered_df[program_col].isna() | 
                        (filtered_df[program_col].astype(str).str.strip() == '') |
                        (filtered_df[program_col].astype(str).str.lower() == 'nan')
                    )
                    filtered_df = filtered_df[program_mask]
                else:
                    # Filter for specific program
                    filtered_df = filtered_df[filtered_df[program_col] == program_filter]

            # Prepare display columns with program handling
            def get_program_display(row):
                if not program_col or pd.isna(row[program_col]):
                    return "No Program Info"
                program_val = str(row[program_col]).strip()
                if program_val == '' or program_val.lower() == 'nan':
                    return "No Program Info"
                return program_val

            display_data = pd.DataFrame({
                'Theme': filtered_df[theme_col] if theme_col else 'N/A',
                'Date': filtered_df[date_col] if date_col else 'N/A',
                'Time': filtered_df[time_col] if time_col else 'N/A',
                'Program': filtered_df.apply(get_program_display, axis=1),
                'Duration': filtered_df[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.tc_matching_tree, display_data)

        except Exception as e:
            print(f"Error filtering TC matching data: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showwarning("Filter Error", f"Could not apply filters: {e}")

    def filter_lmrb_matching_data(self):
        """Filter LMRB matching data based on selected filters."""
        if self.unmatched_lmrb_data is None:
            return

        # Get filter values
        theme_filter = self.lmrb_match_theme_var.get()
        date_filter = self.lmrb_match_date_var.get()
        program_filter = self.lmrb_match_program_var.get()
        duration_filter = self.lmrb_match_duration_var.get()

        # Flexible column finding function
        def find_column(df, possible_names):
            for name in possible_names:
                matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                if matching_cols:
                    return matching_cols[0]
            return None

        # Define possible column names
        theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
        date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
        time_cols = ['Advt_time', 'Time', 'air_time']
        program_cols = ['Program', 'program', 'Prog']
        duration_cols = ['Dur', 'Duration', 'Spot_Duration']

        # Find appropriate columns
        theme_col = find_column(self.unmatched_lmrb_data, theme_cols)
        date_col = find_column(self.unmatched_lmrb_data, date_cols)
        time_col = find_column(self.unmatched_lmrb_data, time_cols)
        program_col = find_column(self.unmatched_lmrb_data, program_cols)
        duration_col = find_column(self.unmatched_lmrb_data, duration_cols)

        # Start with full unmatched LMRB data
        filtered_df = self.unmatched_lmrb_data.copy()

        # Apply filters
        try:
            # Theme Filter
            if theme_filter and theme_col:
                filtered_df = filtered_df[filtered_df[theme_col] == theme_filter]

            # Date Filter
            if date_filter and date_col:
                filtered_df = filtered_df[filtered_df[date_col] == date_filter]

            # Program Filter
            if program_filter and program_col:
                filtered_df = filtered_df[filtered_df[program_col] == program_filter]

            # Duration Filter
            if duration_filter and duration_col:
                filtered_df = filtered_df[filtered_df[duration_col].astype(str) == duration_filter]

            # Prepare display columns
            display_data = pd.DataFrame({
                'Theme': filtered_df[theme_col] if theme_col else 'N/A',
                'Date': filtered_df[date_col] if date_col else 'N/A',
                'Time': filtered_df[time_col] if time_col else 'N/A',
                'Program': filtered_df[program_col] if program_col else 'N/A',
                'Duration': filtered_df[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.lmrb_matching_tree, display_data)

        except Exception as e:
            print(f"Error filtering LMRB matching data: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showwarning("Filter Error", f"Could not apply filters: {e}")

    def reset_tc_matching_filter(self):
        """Reset TC matching filters."""
        # Reset filter variables
        self.tc_match_theme_var.set('')
        self.tc_match_date_var.set('')
        self.tc_match_program_var.set('')

                # Repopulate full data
        if self.unmatched_tc_data is not None:
            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time', 'air_time']
            program_cols = ['Program', 'program', 'Prog']
            duration_cols = ['Dur', 'Duration', 'Spot_Duration']

            # Find appropriate columns
            theme_col = find_column(self.unmatched_tc_data, theme_cols)
            date_col = find_column(self.unmatched_tc_data, date_cols)
            time_col = find_column(self.unmatched_tc_data, time_cols)
            program_col = find_column(self.unmatched_tc_data, program_cols)
            duration_col = find_column(self.unmatched_tc_data, duration_cols)

            # Prepare display data
            display_data = pd.DataFrame({
                'Theme': self.unmatched_tc_data[theme_col] if theme_col else 'N/A',
                'Date': self.unmatched_tc_data[date_col] if date_col else 'N/A',
                'Time': self.unmatched_tc_data[time_col] if time_col else 'N/A',
                'Program': self.unmatched_tc_data[program_col] if program_col else 'N/A',
                'Duration': self.unmatched_tc_data[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.tc_matching_tree, display_data)

            # Repopulate filter dropdowns
            if theme_col:
                themes = [''] + sorted(self.unmatched_tc_data[theme_col].dropna().unique())
                self.tc_match_theme_combobox['values'] = themes

            if date_col:
                dates = [''] + sorted(self.unmatched_tc_data[date_col].dropna().unique())
                self.tc_match_date_combobox['values'] = dates

            if program_col:
                programs = [''] + sorted(self.unmatched_tc_data[program_col].dropna().unique())
                self.tc_match_program_combobox['values'] = programs
    def reset_lmrb_matching_filter(self):
        """Reset LMRB matching filters and repopulate data."""
        # Reset filter variables
        self.lmrb_match_theme_var.set('')
        self.lmrb_match_date_var.set('')
        self.lmrb_match_program_var.set('')
        self.lmrb_match_duration_var.set('')

        # Repopulate full data
        if self.unmatched_lmrb_data is not None:
            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time', 'air_time']
            program_cols = ['Program', 'program', 'Prog']
            duration_cols = ['Dur', 'Duration', 'Spot_Duration']

            # Find appropriate columns
            theme_col = find_column(self.unmatched_lmrb_data, theme_cols)
            date_col = find_column(self.unmatched_lmrb_data, date_cols)
            time_col = find_column(self.unmatched_lmrb_data, time_cols)
            program_col = find_column(self.unmatched_lmrb_data, program_cols)
            duration_col = find_column(self.unmatched_lmrb_data, duration_cols)

            # Prepare display data
            display_data = pd.DataFrame({
                'Theme': self.unmatched_lmrb_data[theme_col] if theme_col else 'N/A',
                'Date': self.unmatched_lmrb_data[date_col] if date_col else 'N/A',
                'Time': self.unmatched_lmrb_data[time_col] if time_col else 'N/A',
                'Program': self.unmatched_lmrb_data[program_col] if program_col else 'N/A',
                'Duration': self.unmatched_lmrb_data[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.lmrb_matching_tree, display_data)

            # Repopulate filter dropdowns
            if theme_col:
                themes = [''] + sorted(self.unmatched_lmrb_data[theme_col].dropna().unique())
                self.lmrb_match_theme_combobox['values'] = themes

            if date_col:
                dates = [''] + sorted(self.unmatched_lmrb_data[date_col].dropna().unique())
                self.lmrb_match_date_combobox['values'] = dates

            if program_col:
                programs = [''] + sorted(self.unmatched_lmrb_data[program_col].dropna().unique())
                self.lmrb_match_program_combobox['values'] = programs

            if duration_col:
                durations = [''] + sorted(str(d) for d in self.unmatched_lmrb_data[duration_col].dropna().unique())
                self.lmrb_match_duration_combobox['values'] = durations

    def populate_manual_matching_data(self):
        """Populate manual matching data for TC and LMRB with robust handling of missing program information."""
        try:
            # Print available data for debugging
            print("\n--- Populating Manual Matching Data ---")
            print("Unmatched TC Data:")
            if self.unmatched_tc_data is not None:
                print("TC Columns:", self.unmatched_tc_data.columns.tolist())
                print("TC Shape:", self.unmatched_tc_data.shape)
                print("TC Sample:")
                print(self.unmatched_tc_data.head())
            else:
                print("No unmatched TC data available")

            print("\nUnmatched LMRB Data:")
            if self.unmatched_lmrb_data is not None:
                print("LMRB Columns:", self.unmatched_lmrb_data.columns.tolist())
                print("LMRB Shape:", self.unmatched_lmrb_data.shape)
                print("LMRB Sample:")
                print(self.unmatched_lmrb_data.head())
            else:
                print("No unmatched LMRB data available")

            # Flexible column finding function
            def find_column(df, possible_names):
                """Find the best matching column from possible names."""
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Helper function to safely get program value
            def get_program_display_value(row, program_col):
                """Get program display value, handling missing data gracefully."""
                if not program_col or pd.isna(row.get(program_col)):
                    return "No Program Info"
                
                program_val = str(row[program_col]).strip()
                if program_val == '' or program_val.lower() in ['nan', 'none', 'null']:
                    return "No Program Info"
                return program_val

            # Helper function to prepare unique filter values
            def get_unique_filter_values(df, column, include_missing_option=False):
                """Get unique values for filter dropdowns."""
                values = ['']  # Always start with empty option
                
                if column and column in df.columns:
                    if include_missing_option:
                        # Check if there are rows with missing program info
                        has_missing = (
                            df[column].isna().any() or
                            (df[column].astype(str).str.strip() == '').any() or
                            (df[column].astype(str).str.lower().isin(['nan', 'none', 'null'])).any()
                        )
                        if has_missing:
                            values.append("No Program Info")
                    
                    # Get valid non-empty values
                    valid_values = df[column].dropna()
                    valid_values = valid_values[valid_values.astype(str).str.strip() != '']
                    valid_values = valid_values[~valid_values.astype(str).str.lower().isin(['nan', 'none', 'null'])]
                    
                    if len(valid_values) > 0:
                        values.extend(sorted(valid_values.unique()))
                
                return values

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme', 'Product Details']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time', 'air_time', 'Advt_Time']
            program_cols = ['Program', 'program', 'Prog', 'Programme']
            duration_cols = ['Dur', 'Duration', 'Spot_Duration']

            # =====================================================================
            # POPULATE TC MATCHING DATA
            # =====================================================================
            if self.unmatched_tc_data is not None and not self.unmatched_tc_data.empty:
                try:
                    print("\n=== Processing TC Data ===")
                    
                    # Find appropriate columns for TC data
                    tc_theme_col = find_column(self.unmatched_tc_data, theme_cols)
                    tc_date_col = find_column(self.unmatched_tc_data, date_cols)
                    tc_time_col = find_column(self.unmatched_tc_data, time_cols)
                    tc_program_col = find_column(self.unmatched_tc_data, program_cols)
                    tc_duration_col = find_column(self.unmatched_tc_data, duration_cols)

                    print(f"TC Column Mapping:")
                    print(f"  Theme Column: {tc_theme_col}")
                    print(f"  Date Column: {tc_date_col}")
                    print(f"  Time Column: {tc_time_col}")
                    print(f"  Program Column: {tc_program_col}")
                    print(f"  Duration Column: {tc_duration_col}")

                    # Clear existing TC tree items
                    for item in self.tc_matching_tree.get_children():
                        self.tc_matching_tree.delete(item)

                    # Populate TC Matching Treeview
                    tc_rows_added = 0
                    for index, row in self.unmatched_tc_data.iterrows():
                        try:
                            # Get program value with proper handling
                            program_display = get_program_display_value(row, tc_program_col)
                            
                            # Prepare row values for treeview
                            row_values = [
                                str(row.get(tc_theme_col, 'N/A')) if tc_theme_col else 'N/A',
                                str(row.get(tc_date_col, 'N/A')) if tc_date_col else 'N/A',
                                str(row.get(tc_time_col, 'N/A')) if tc_time_col else 'N/A',
                                program_display,
                                str(row.get(tc_duration_col, 'N/A')) if tc_duration_col else 'N/A'
                            ]
                            
                            # Clean up any NaN values
                            row_values = [str(val).replace('nan', 'N/A') for val in row_values]
                            
                            self.tc_matching_tree.insert('', 'end', values=row_values)
                            tc_rows_added += 1
                            
                        except Exception as row_error:
                            print(f"Error processing TC row {index}: {row_error}")
                            continue

                    print(f"TC Tree populated with {tc_rows_added} rows")

                    # Populate TC Filter Dropdowns
                    try:
                        # Theme filter
                        if tc_theme_col:
                            tc_themes = get_unique_filter_values(self.unmatched_tc_data, tc_theme_col)
                            self.tc_match_theme_combobox['values'] = tc_themes
                            print(f"TC Theme filter populated with {len(tc_themes)} options")

                        # Date filter
                        if tc_date_col:
                            tc_dates = get_unique_filter_values(self.unmatched_tc_data, tc_date_col)
                            self.tc_match_date_combobox['values'] = tc_dates
                            print(f"TC Date filter populated with {len(tc_dates)} options")

                        # Program filter - special handling for missing programs
                        if tc_program_col:
                            tc_programs = get_unique_filter_values(
                                self.unmatched_tc_data, 
                                tc_program_col, 
                                include_missing_option=True
                            )
                            self.tc_match_program_combobox['values'] = tc_programs
                            print(f"TC Program filter populated with {len(tc_programs)} options: {tc_programs[:5]}...")
                        else:
                            # If no program column exists, only show "No Program Info"
                            self.tc_match_program_combobox['values'] = ['', 'No Program Info']
                            print("TC Program filter set to 'No Program Info' only")

                    except Exception as filter_error:
                        print(f"Error populating TC filters: {filter_error}")

                except Exception as tc_error:
                    print(f"Error processing TC data: {tc_error}")
                    import traceback
                    traceback.print_exc()

            else:
                print("No TC data to populate")
                # Clear TC tree if no data
                for item in self.tc_matching_tree.get_children():
                    self.tc_matching_tree.delete(item)

            # =====================================================================
            # POPULATE LMRB MATCHING DATA
            # =====================================================================
            if self.unmatched_lmrb_data is not None and not self.unmatched_lmrb_data.empty:
                try:
                    print("\n=== Processing LMRB Data ===")
                    
                    # Find appropriate columns for LMRB data
                    lmrb_theme_col = find_column(self.unmatched_lmrb_data, theme_cols)
                    lmrb_date_col = find_column(self.unmatched_lmrb_data, date_cols)
                    lmrb_time_col = find_column(self.unmatched_lmrb_data, time_cols)
                    lmrb_program_col = find_column(self.unmatched_lmrb_data, program_cols)
                    lmrb_duration_col = find_column(self.unmatched_lmrb_data, duration_cols)

                    print(f"LMRB Column Mapping:")
                    print(f"  Theme Column: {lmrb_theme_col}")
                    print(f"  Date Column: {lmrb_date_col}")
                    print(f"  Time Column: {lmrb_time_col}")
                    print(f"  Program Column: {lmrb_program_col}")
                    print(f"  Duration Column: {lmrb_duration_col}")

                    # Clear existing LMRB tree items
                    for item in self.lmrb_matching_tree.get_children():
                        self.lmrb_matching_tree.delete(item)

                    # Populate LMRB Matching Treeview
                    lmrb_rows_added = 0
                    for index, row in self.unmatched_lmrb_data.iterrows():
                        try:
                            # Get program value with proper handling
                            program_display = get_program_display_value(row, lmrb_program_col)
                            
                            # Prepare row values for treeview
                            row_values = [
                                str(row.get(lmrb_theme_col, 'N/A')) if lmrb_theme_col else 'N/A',
                                str(row.get(lmrb_date_col, 'N/A')) if lmrb_date_col else 'N/A',
                                str(row.get(lmrb_time_col, 'N/A')) if lmrb_time_col else 'N/A',
                                program_display,
                                str(row.get(lmrb_duration_col, 'N/A')) if lmrb_duration_col else 'N/A'
                            ]
                            
                            # Clean up any NaN values
                            row_values = [str(val).replace('nan', 'N/A') for val in row_values]
                            
                            self.lmrb_matching_tree.insert('', 'end', values=row_values)
                            lmrb_rows_added += 1
                            
                        except Exception as row_error:
                            print(f"Error processing LMRB row {index}: {row_error}")
                            continue

                    print(f"LMRB Tree populated with {lmrb_rows_added} rows")

                    # Populate LMRB Filter Dropdowns
                    try:
                        # Theme filter
                        if lmrb_theme_col:
                            lmrb_themes = get_unique_filter_values(self.unmatched_lmrb_data, lmrb_theme_col)
                            self.lmrb_match_theme_combobox['values'] = lmrb_themes
                            print(f"LMRB Theme filter populated with {len(lmrb_themes)} options")

                        # Date filter
                        if lmrb_date_col:
                            lmrb_dates = get_unique_filter_values(self.unmatched_lmrb_data, lmrb_date_col)
                            self.lmrb_match_date_combobox['values'] = lmrb_dates
                            print(f"LMRB Date filter populated with {len(lmrb_dates)} options")

                        # Program filter - special handling for missing programs
                        if lmrb_program_col:
                            lmrb_programs = get_unique_filter_values(
                                self.unmatched_lmrb_data, 
                                lmrb_program_col, 
                                include_missing_option=True
                            )
                            self.lmrb_match_program_combobox['values'] = lmrb_programs
                            print(f"LMRB Program filter populated with {len(lmrb_programs)} options: {lmrb_programs[:5]}...")
                        else:
                            # If no program column exists, only show "No Program Info"
                            self.lmrb_match_program_combobox['values'] = ['', 'No Program Info']
                            print("LMRB Program filter set to 'No Program Info' only")

                        # Duration filter
                        if lmrb_duration_col:
                            lmrb_durations = get_unique_filter_values(self.unmatched_lmrb_data, lmrb_duration_col)
                            self.lmrb_match_duration_combobox['values'] = lmrb_durations
                            print(f"LMRB Duration filter populated with {len(lmrb_durations)} options")

                    except Exception as filter_error:
                        print(f"Error populating LMRB filters: {filter_error}")

                except Exception as lmrb_error:
                    print(f"Error processing LMRB data: {lmrb_error}")
                    import traceback
                    traceback.print_exc()

            else:
                print("No LMRB data to populate")
                # Clear LMRB tree if no data
                for item in self.lmrb_matching_tree.get_children():
                    self.lmrb_matching_tree.delete(item)

            # =====================================================================
            # FINAL STATUS UPDATE
            # =====================================================================
            tc_count = len(self.unmatched_tc_data) if self.unmatched_tc_data is not None else 0
            lmrb_count = len(self.unmatched_lmrb_data) if self.unmatched_lmrb_data is not None else 0
            
            print(f"\n=== Manual Matching Data Population Complete ===")
            print(f"TC Records Available: {tc_count}")
            print(f"LMRB Records Available: {lmrb_count}")
            print("Manual matching interface is ready for use.")

            # Enable the manual matching interface if we have data
            if tc_count > 0 and lmrb_count > 0:
                print("Manual matching is enabled - both TC and LMRB data available")
            elif tc_count == 0:
                print("Warning: No TC data available for manual matching")
            elif lmrb_count == 0:
                print("Warning: No LMRB data available for manual matching")
            else:
                print("Warning: No data available for manual matching")

        except Exception as e:
            print(f"Critical error in populate_manual_matching_data: {e}")
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Could not populate manual matching data: {e}")

            # Ensure treeviews are cleared on error
            try:
                for item in self.tc_matching_tree.get_children():
                    self.tc_matching_tree.delete(item)
                for item in self.lmrb_matching_tree.get_children():
                    self.lmrb_matching_tree.delete(item)
            except:
                pass


    def reset_tc_matching_filter(self):
        """Reset TC matching filters and repopulate data."""
        # Reset filter variables
        self.tc_match_theme_var.set('')
        self.tc_match_date_var.set('')
        self.tc_match_program_var.set('')

        # Repopulate full data
        if self.unmatched_tc_data is not None:
            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Find appropriate columns
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time']
            program_cols = ['Program']
            duration_cols = ['Dur', 'Duration']

            # Identify columns
            theme_col = find_column(self.unmatched_tc_data, theme_cols)
            date_col = find_column(self.unmatched_tc_data, date_cols)
            time_col = find_column(self.unmatched_tc_data, time_cols)
            program_col = find_column(self.unmatched_tc_data, program_cols)
            duration_col = find_column(self.unmatched_tc_data, duration_cols)

            # Create Date column if not exists
            if not date_col and all(col in self.unmatched_tc_data.columns for col in ['Dd', 'Mn', 'Yr']):
                self.unmatched_tc_data['Date'] = self.unmatched_tc_data.apply(
                    lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                    axis=1
                )
                date_col = 'Date'

            # Prepare display data
            display_data = pd.DataFrame({
                'Theme': self.unmatched_tc_data[theme_col] if theme_col else 'N/A',
                'Date': self.unmatched_tc_data[date_col] if date_col else 'N/A',
                'Time': self.unmatched_tc_data[time_col] if time_col else 'N/A',
                'Program': self.unmatched_tc_data[program_col] if program_col else 'N/A',
                'Duration': self.unmatched_tc_data[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.tc_matching_tree, display_data)

            # Repopulate filter dropdowns
            if theme_col:
                themes = [''] + sorted(self.unmatched_tc_data[theme_col].unique())
                if hasattr(self, 'tc_match_theme_combobox'):
                    self.tc_match_theme_combobox['values'] = themes

            if date_col:
                dates = [''] + sorted(self.unmatched_tc_data[date_col].unique())
                if hasattr(self, 'tc_match_date_combobox'):
                    self.tc_match_date_combobox['values'] = dates

            if program_col:
                programs = [''] + sorted(self.unmatched_tc_data[program_col].unique())
                if hasattr(self, 'tc_match_program_combobox'):
                    self.tc_match_program_combobox['values'] = programs

    def reset_lmrb_matching_filter(self):
        """Reset LMRB matching filters and repopulate data."""
        # Reset filter variables
        self.lmrb_match_theme_var.set('')
        self.lmrb_match_date_var.set('')
        self.lmrb_match_program_var.set('')
        self.lmrb_match_duration_var.set('')

        # Repopulate full data
        if self.unmatched_lmrb_data is not None:
            # Flexible column finding function
            def find_column(df, possible_names):
                for name in possible_names:
                    matching_cols = [col for col in df.columns if col.lower() == name.lower()]
                    if matching_cols:
                        return matching_cols[0]
                return None

            # Define possible column names
            theme_cols = ['Advt_Theme', 'Advt_theme', 'Theme']
            date_cols = ['Date', 'DATE', 'Dd/Mn/Yr']
            time_cols = ['Advt_time', 'Time']
            program_cols = ['Program']
            duration_cols = ['Dur', 'Duration']

            # Identify columns
            theme_col = find_column(self.unmatched_lmrb_data, theme_cols)
            date_col = find_column(self.unmatched_lmrb_data, date_cols)
            time_col = find_column(self.unmatched_lmrb_data, time_cols)
            program_col = find_column(self.unmatched_lmrb_data, program_cols)
            duration_col = find_column(self.unmatched_lmrb_data, duration_cols)

            # Create Date column if not exists
            if not date_col and all(col in self.unmatched_lmrb_data.columns for col in ['Dd', 'Mn', 'Yr']):
                self.unmatched_lmrb_data['Date'] = self.unmatched_lmrb_data.apply(
                    lambda x: f"{int(x['Dd'])}/{int(x['Mn'])}/{int(x['Yr'])}", 
                    axis=1
                )
                date_col = 'Date'

            # Prepare display data
            display_data = pd.DataFrame({
                'Theme': self.unmatched_lmrb_data[theme_col] if theme_col else 'N/A',
                'Date': self.unmatched_lmrb_data[date_col] if date_col else 'N/A',
                'Time': self.unmatched_lmrb_data[time_col] if time_col else 'N/A',
                'Program': self.unmatched_lmrb_data[program_col] if program_col else 'N/A',
                'Duration': self.unmatched_lmrb_data[duration_col] if duration_col else 'N/A'
            })

            # Update treeview
            self.update_treeview(self.lmrb_matching_tree, display_data)

            # Repopulate filter dropdowns
            if theme_col:
                themes = [''] + sorted(self.unmatched_lmrb_data[theme_col].unique())
                if hasattr(self, 'lmrb_match_theme_combobox'):
                    self.lmrb_match_theme_combobox['values'] = themes

            if date_col:
                dates = [''] + sorted(self.unmatched_lmrb_data[date_col].unique())
                if hasattr(self, 'lmrb_match_date_combobox'):
                    self.lmrb_match_date_combobox['values'] = dates

            if program_col:
                programs = [''] + sorted(self.unmatched_lmrb_data[program_col].unique())
                if hasattr(self, 'lmrb_match_program_combobox'):
                    self.lmrb_match_program_combobox['values'] = programs

            if duration_col:
                durations = [''] + sorted(str(d) for d in self.unmatched_lmrb_data[duration_col].unique())
                if hasattr(self, 'lmrb_match_duration_combobox'):
                    self.lmrb_match_duration_combobox['values'] = durations

    def on_tc_matching_selection(self, event):
        """Handle TC row selection and prepare for matching."""
        selected_items = self.tc_matching_tree.selection()
        if selected_items:
            # Get the selected row data
            selected_item = selected_items[0]
            row_values = self.tc_matching_tree.item(selected_item)['values']
            
            if row_values and len(row_values) >= 5:
                # Create a dictionary with consistent keys
                row_dict = {
                    'Theme': row_values[0],
                    'Date': row_values[1],
                    'Time': row_values[2],
                    'Program': row_values[3],
                    'Duration': row_values[4]
                }
                
                # Select the row
                self.select_tc_row(row_dict)

    def on_lmrb_matching_selection(self, event):
        """Handle LMRB row selection and prepare for matching."""
        selected_items = self.lmrb_matching_tree.selection()
        if selected_items:
            # Get the selected row data
            selected_item = selected_items[0]
            row_values = self.lmrb_matching_tree.item(selected_item)['values']
            
            if row_values and len(row_values) >= 5:
                # Create a dictionary with consistent keys
                row_dict = {
                    'Theme': row_values[0],
                    'Date': row_values[1],
                    'Time': row_values[2],
                    'Program': row_values[3],
                    'Duration': row_values[4]
                }
                
                # Select the row
                self.select_lmrb_row(row_dict)

    def highlight_similar_lmrb_rows(self, tc_row_dict):
        """Highlight LMRB rows similar to the selected TC row - Date First Priority."""
        try:
            # Extract TC row data with safe access
            tc_theme = str(tc_row_dict.get('Theme', '')).strip()
            tc_date = str(tc_row_dict.get('Date', '')).strip()
            tc_time = str(tc_row_dict.get('Time', '')).strip()
            tc_program = str(tc_row_dict.get('Program', '')).strip()
            tc_duration = str(tc_row_dict.get('Duration', '')).strip()

            print(f"TC Selection - Theme: {tc_theme}, Date: {tc_date}, Time: {tc_time}, Program: {tc_program}, Duration: {tc_duration}")

            # Reset previous highlighting
            for item in self.lmrb_matching_tree.get_children():
                self.lmrb_matching_tree.item(item, tags=())

            # STEP 1: First filter by DATE similarity (highest priority)
            date_matched_rows = []
            non_date_matched_rows = []
            
            for item in self.lmrb_matching_tree.get_children():
                try:
                    row_values = self.lmrb_matching_tree.item(item)['values']
                    if row_values and len(row_values) >= 5:
                        lmrb_date = str(row_values[1]).strip()
                        
                        # Check if dates are similar
                        if self.dates_are_similar(tc_date, lmrb_date):
                            date_matched_rows.append((item, row_values))
                            print(f"Date match found: TC({tc_date}) == LMRB({lmrb_date})")
                        else:
                            non_date_matched_rows.append((item, row_values))
                except Exception as e:
                    print(f"Error processing LMRB row: {e}")
                    continue

            print(f"Step 1: Found {len(date_matched_rows)} rows with matching dates")

            # STEP 2: Apply scoring system to date-matched rows first
            def calculate_similarity(lmrb_row_values, is_date_matched=False):
                """Calculate similarity score between TC and LMRB rows."""
                score = 0
                
                # Ensure we have enough values (Theme, Date, Time, Program, Duration)
                if len(lmrb_row_values) < 5:
                    return 0
                
                # Extract LMRB values with safe conversion
                lmrb_theme = str(lmrb_row_values[0]).strip()
                lmrb_date = str(lmrb_row_values[1]).strip()
                lmrb_time = str(lmrb_row_values[2]).strip()
                lmrb_program = str(lmrb_row_values[3]).strip()
                lmrb_duration = str(lmrb_row_values[4]).strip()
                
                # Date similarity - HIGHEST PRIORITY
                if is_date_matched or self.dates_are_similar(tc_date, lmrb_date):
                    score += 10  # Very high score for date match
                
                # Theme similarity (case-insensitive)
                if lmrb_theme.lower() == tc_theme.lower():
                    score += 5  # High weight for exact theme match
                elif tc_theme.lower() in lmrb_theme.lower() or lmrb_theme.lower() in tc_theme.lower():
                    score += 2  # Partial theme match
                
                # Duration similarity
                if lmrb_duration == tc_duration:
                    score += 3  # High weight for duration match
                
                # Time similarity (within tolerance)
                try:
                    if tc_time and lmrb_time and tc_time != '00:00:00' and lmrb_time != '00:00:00':
                        def parse_time_flexible(time_str):
                            time_str = str(time_str).strip()
                            if ':' in time_str:
                                parts = time_str.split(':')
                                if len(parts) >= 2:
                                    hours = int(parts[0])
                                    minutes = int(parts[1])
                                    seconds = int(parts[2]) if len(parts) > 2 else 0
                                    return hours * 3600 + minutes * 60 + seconds
                            return 0
                        
                        tc_seconds = parse_time_flexible(tc_time)
                        lmrb_seconds = parse_time_flexible(lmrb_time)
                        
                        time_diff = abs(tc_seconds - lmrb_seconds)
                        if time_diff <= 30:  # Within 30 seconds
                            score += 5  # Very high score for close time match
                        elif time_diff <= 60:  # Within 1 minute
                            score += 3
                        elif time_diff <= 300:  # Within 5 minutes
                            score += 1
                except Exception as e:
                    print(f"Time parsing error: {e}")
                    pass

                # Program similarity (lower priority)
                if lmrb_program.lower() == tc_program.lower():
                    score += 1
                elif tc_program.lower() in lmrb_program.lower() or lmrb_program.lower() in tc_program.lower():
                    score += 0.5  # Partial program match

                return score

            # Score date-matched rows (these get priority)
            date_matched_with_scores = []
            for item, row_values in date_matched_rows:
                similarity_score = calculate_similarity(row_values, is_date_matched=True)
                date_matched_with_scores.append((item, similarity_score, row_values, True))  # True indicates date match

            # Score non-date-matched rows (lower priority)
            non_date_matched_with_scores = []
            for item, row_values in non_date_matched_rows:
                similarity_score = calculate_similarity(row_values, is_date_matched=False)
                non_date_matched_with_scores.append((item, similarity_score, row_values, False))  # False indicates no date match

            # Combine all rows with date-matched rows first
            all_rows_with_scores = date_matched_with_scores + non_date_matched_with_scores

            # Sort: Date-matched rows first (by score), then non-date-matched rows (by score)
            date_matched_with_scores.sort(key=lambda x: x[1], reverse=True)
            non_date_matched_with_scores.sort(key=lambda x: x[1], reverse=True)
            
            # Combine with date matches first
            sorted_rows = date_matched_with_scores + non_date_matched_with_scores

            print(f"Top date-matched scores: {[x[1] for x in date_matched_with_scores[:3]]}")
            print(f"Top non-date-matched scores: {[x[1] for x in non_date_matched_with_scores[:3]]}")

            # Apply highlighting and reordering
            for index, (item, score, row_values, is_date_matched) in enumerate(sorted_rows):
                try:
                    # Special highlighting for date-matched rows
                    if is_date_matched:
                        if score >= 18:  # Date + Theme + Duration + Time match
                            self.lmrb_matching_tree.item(item, tags=('perfect_match',))
                        elif score >= 15:  # Date + Theme + Duration match
                            self.lmrb_matching_tree.item(item, tags=('excellent_match',))
                        elif score >= 12:  # Date + Theme match
                            self.lmrb_matching_tree.item(item, tags=('good_match',))
                        else:  # Date match only
                            self.lmrb_matching_tree.item(item, tags=('date_match',))
                    else:
                        # Non-date matched rows get lower priority colors
                        if score >= 8:
                            self.lmrb_matching_tree.item(item, tags=('medium_similar',))
                        elif score >= 3:
                            self.lmrb_matching_tree.item(item, tags=('low_similar',))
                    
                    # Move rows to reflect priority order
                    self.lmrb_matching_tree.move(item, '', index)
                    
                except Exception as e:
                    print(f"Error highlighting/moving row: {e}")
                    continue

            # Configure highlighting tags with different colors
            self.lmrb_matching_tree.tag_configure('perfect_match', background='#00FF00', foreground='black')      # Bright green
            self.lmrb_matching_tree.tag_configure('excellent_match', background="#60A0F3", foreground='black')   # Lime green
            #self.lmrb_matching_tree.tag_configure('good_match', background="#5AC2E2", foreground='black')        # Light green
            #self.lmrb_matching_tree.tag_configure('date_match', background='#FFFF99', foreground='black')        # Light yellow
            #self.lmrb_matching_tree.tag_configure('medium_similar', background='#FFA500', foreground='black')    # Orange
            #self.lmrb_matching_tree.tag_configure('low_similar', background='#FFE4B5', foreground='black')       # Light orange

            # Auto-select the best match (first row after sorting)
            if sorted_rows:
                best_match_item = sorted_rows[0][0]
                best_match_score = sorted_rows[0][1]
                is_best_date_matched = sorted_rows[0][3]
                
                self.lmrb_matching_tree.selection_set(best_match_item)
                self.lmrb_matching_tree.focus(best_match_item)
                self.lmrb_matching_tree.see(best_match_item)  # Ensure it's visible
                
                print(f"BEST MATCH selected - Score: {best_match_score}, Date Matched: {is_best_date_matched}")

            print(f"Total rows processed: {len(sorted_rows)}")
            print(f"Date-matched rows: {len(date_matched_with_scores)}")
            print(f"Non-date-matched rows: {len(non_date_matched_with_scores)}")

        except Exception as e:
            print(f"Error in highlight_similar_lmrb_rows: {e}")
            import traceback
            traceback.print_exc()

    def dates_are_similar(self, date1, date2):
        """Check if two dates are similar (exact match or various formats)."""
        if not date1 or not date2 or date1.lower() == 'n/a' or date2.lower() == 'n/a':
            return False
        
        # Direct string comparison first
        if date1.lower().strip() == date2.lower().strip():
            return True
        
        try:
            # Try to parse and compare dates
            from datetime import datetime
            
            # Common date formats to try
            date_formats = [
                '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d',
                '%d-%m-%Y', '%m-%d-%Y', '%Y%m%d', '%d.%m.%Y'
            ]
            
            parsed_date1 = None
            parsed_date2 = None
            
            # Try to parse date1
            for fmt in date_formats:
                try:
                    parsed_date1 = datetime.strptime(date1.strip(), fmt).date()
                    break
                except:
                    continue
            
            # Try to parse date2
            for fmt in date_formats:
                try:
                    parsed_date2 = datetime.strptime(date2.strip(), fmt).date()
                    break
                except:
                    continue
            
            # If both dates were parsed successfully, compare them
            if parsed_date1 and parsed_date2:
                return parsed_date1 == parsed_date2
            
        except Exception as e:
            print(f"Date parsing error: {e}")
        
        return False

    def reset_app(self):
        """Reset the application to its initial state."""
        # Clear data variables
        self.lmrb_data = None
        self.tc_data = None
        self.schedule_data = None  # Add this
        self.lmrb_filtered = None
        self.tc_filtered = None
        self.theme_mapping = []
        self.matched_data = None
        self.unmatched_tc_data = None
        self.unmatched_lmrb_data = None  # Add this
        self.matched_lmrb_indices = []
        self.selected_channel = None
        self.tc_data_original = None
        self.program_mismatched_data = None
        self.matched_indices = None
        self.filtered_lmrb_df = None
        self.manually_matched_schedule_items = []

        # Clear schedule matching data
        self.schedule_summary_results = []
        self.selected_trailers = []

        # ✅ NEW: Reset product filter variables
        self.selected_products = []
        self.available_products = []
        self.lmrb_product_filtered = None

        # Reset Tkinter variables
        self.lmrb_file_var.set("")
        self.tc_file_var.set("")
        self.schedule_file_var.set("")  # Add this
        self.channel_var.set("")
        self.lmrb_theme_var.set("")
        self.lmrb_dur_var.set("")
        self.tc_theme_var.set("")
        self.tc_dur_var.set("")
        self.schedule_theme_var.set("")  # Add this
        self.schedule_dur_var.set("")   # Add this
        self.mapping_type_var.set("COMMERCIAL BENEFITS")  # Add this

        # ✅ NEW: Reset product filter variables
        if hasattr(self, 'product_var'):
            self.product_var.set("")

        # Reset manual matching filter variables
        self.tc_match_theme_var.set('')
        self.tc_match_date_var.set('')
        self.tc_match_program_var.set('')
        self.lmrb_match_theme_var.set('')
        self.lmrb_match_date_var.set('')
        self.lmrb_match_program_var.set('')
        self.lmrb_match_duration_var.set('')

        # Reset other variables
        self.ignore_date_var.set(False)
        self.time_tolerance_var.set(30)
        self.progress_var.set(0)
        self.matching_status_var.set("Ready to match")
        self.total_lmrb_var.set("Total LMRB Records: 0")
        self.matched_count_var.set("Matched Records: 0")
        self.unmatched_tc_var.set("Unmatched TC Records: 0")
        self.unmatched_lmrb_var.set("Unmatched LMRB Records: 0")  # Add this
        self.match_rate_var.set("Match Rate: 0%")
        self.status_var.set("Ready")

        # Clear manual matching selection variables
        self.selected_tc_row = None
        self.selected_lmrb_row = None
        self.last_matched_tc = None
        self.last_matched_lmrb = None
        self.last_matched_full_lmrb = None

        # Clear treeviews
        trees_to_clear = [
            self.lmrb_tree, self.tc_tree, self.schedule_tree,  # Add schedule_tree
            self.mappings_tree, self.matched_tree, self.unmatched_tree, 
            self.unmatched_lmrb_tree  # Add this if it exists
        ]

        # Clear schedule matching treeviews
        if hasattr(self, 'schedule_summary_tree'):
            trees_to_clear.append(self.schedule_summary_tree)
        if hasattr(self, 'trailer_selection_tree'):
            trees_to_clear.append(self.trailer_selection_tree)

        for tree in trees_to_clear:
            try:
                for item in tree.get_children():
                    tree.delete(item)
            except:
                pass  # Skip if tree doesn't exist

        # ✅ NEW: Clear product UI elements
        if hasattr(self, 'selected_products_listbox'):
            self.selected_products_listbox.delete(0, tk.END)
        if hasattr(self, 'product_combobox'):
            self.product_combobox["values"] = []
        if hasattr(self, 'product_filter_status'):
            self.product_filter_status.config(text="No products selected - showing all products", foreground="gray")

        # Reset channel combobox
        self.channel_combobox["values"] = []

        # Reset notebook tabs
        self.notebook.select(0)
        self.notebook.tab(1, state="disabled")  # Theme Mapping
        self.notebook.tab(2, state="disabled")  # Matching
        self.notebook.tab(3, state="disabled")  # Results
        self.notebook.tab(4, state="disabled")  # Manual Matching
        self.notebook.tab(5, state="disabled")  # Schedule Matching

        # Always enable mapping treeview and delete button
        self.mappings_tree.config(selectmode="extended")
        for child in self.mapping_tab.winfo_children():
            if isinstance(child, ttk.Button) and child['text'] == "Delete Selected Mapping":
                child['state'] = 'normal'

        self.status_bar.config(text="Application reset. Ready.")

        # Reset new summary report variables
        self.report_month_var.set("")
        self.estimate_no_var.set("")
        self.supplier_invoice_var.set("")
        self.po_no_var.set("")
        self.invoice_no_var.set("")

        # Clear summary report trees if they exist
        if hasattr(self, 'commercial_tree'):
            for item in self.commercial_tree.get_children():
                self.commercial_tree.delete(item)

        if hasattr(self, 'sponsorship_tree'):
            for item in self.sponsorship_tree.get_children():
                self.sponsorship_tree.delete(item)

        # Update summary report totals
        if hasattr(self, 'commercial_totals_label'):
            self.commercial_totals_label.config(text="Grand Total: 0 | 0 | 0 | 0 | 0 | 0")

        if hasattr(self, 'sponsorship_totals_label'):
            self.sponsorship_totals_label.config(text="Grand Total: 0 | 0 | 0 | 0 | 0")

        # Update channel label in summary report
        if hasattr(self, 'report_channel_label'):
            self.report_channel_label.config(text="")

    def setup_summary_report_tab(self):
        """Setup the Summary Report tab with header inputs and data display."""

        # Title
        title_label = ttk.Label(self.summary_report_tab, text="Summary Report Generator", style="Heading.TLabel")
        title_label.grid(row=0, column=0, columnspan=4, pady=10, sticky="w")

        # Header Information Frame
        header_frame = ttk.LabelFrame(self.summary_report_tab, text="Report Header Information", padding=10)
        header_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=10)

        # Row 1: Month and Channel
        ttk.Label(header_frame, text="Month:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        month_entry = ttk.Entry(header_frame, textvariable=self.report_month_var, width=20)
        month_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(header_frame, text="Channel:").grid(row=0, column=2, sticky="w", padx=5, pady=5)
        channel_label = ttk.Label(header_frame, text="", foreground="blue", font=("Arial", 10, "bold"))
        channel_label.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.report_channel_label = channel_label  # Store reference to update later

        # Row 2: Estimate No and Supplier Invoice
        ttk.Label(header_frame, text="Estimate No:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        estimate_entry = ttk.Entry(header_frame, textvariable=self.estimate_no_var, width=20)
        estimate_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(header_frame, text="Supplier Invoice No:").grid(row=1, column=2, sticky="w", padx=5, pady=5)
        supplier_entry = ttk.Entry(header_frame, textvariable=self.supplier_invoice_var, width=20)
        supplier_entry.grid(row=1, column=3, padx=5, pady=5, sticky="w")

        # Row 3: PO No and Invoice No
        ttk.Label(header_frame, text="PO No:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        po_entry = ttk.Entry(header_frame, textvariable=self.po_no_var, width=20)
        po_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

        ttk.Label(header_frame, text="Invoice No:").grid(row=2, column=2, sticky="w", padx=5, pady=5)
        invoice_entry = ttk.Entry(header_frame, textvariable=self.invoice_no_var, width=20)
        invoice_entry.grid(row=2, column=3, padx=5, pady=5, sticky="w")

        # Control Buttons Frame
        control_frame = ttk.Frame(self.summary_report_tab)
        control_frame.grid(row=2, column=0, columnspan=4, sticky="ew", pady=10)

        # Generate Report Button
        generate_button = ttk.Button(control_frame, text="🔄 Generate Report", command=self.generate_and_display_summary)
        generate_button.pack(side=tk.LEFT, padx=5)

        # Export Button
        export_button = ttk.Button(control_frame, text="📊 Export to Excel", command=self.export_summary_report)
        export_button.pack(side=tk.LEFT, padx=5)

        # Refresh Button
        refresh_button = ttk.Button(control_frame, text="🔄 Refresh Data", command=self.sync_summary_report_data)
        refresh_button.pack(side=tk.LEFT, padx=5)

        # Commercial Benefits Frame
        commercial_frame = ttk.LabelFrame(self.summary_report_tab, text="Commercial Benefits", padding=10)
        commercial_frame.grid(row=3, column=0, columnspan=4, sticky="nsew", pady=5)

        # Commercial Benefits Treeview
        self.commercial_tree = ttk.Treeview(commercial_frame, columns=(
            'product', 'duration', 'planned', 'aired', 'missed', 'extra', 'third_party', 'average_30_sec'
        ), show='headings', height=8)

        # Configure Commercial tree columns
        self.commercial_tree.heading('product', text='PRODUCT')
        self.commercial_tree.heading('duration', text='DUR')
        self.commercial_tree.heading('planned', text='PLANNED')
        self.commercial_tree.heading('aired', text='AIRED')
        self.commercial_tree.heading('missed', text='MISSED')
        self.commercial_tree.heading('extra', text='EXTRA')
        self.commercial_tree.heading('third_party', text='3RD PARTY')
        self.commercial_tree.heading('average_30_sec', text='Average 30 sec.')

        # Set column widths
        self.commercial_tree.column('product', width=200)
        self.commercial_tree.column('duration', width=60)
        self.commercial_tree.column('planned', width=80)
        self.commercial_tree.column('aired', width=80)
        self.commercial_tree.column('missed', width=80)
        self.commercial_tree.column('extra', width=80)
        self.commercial_tree.column('third_party', width=80)
        self.commercial_tree.column('average_30_sec', width=100)

        # Commercial scrollbars
        commercial_scrolly = ttk.Scrollbar(commercial_frame, orient="vertical", command=self.commercial_tree.yview)
        self.commercial_tree.configure(yscrollcommand=commercial_scrolly.set)

        commercial_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.commercial_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Commercial Totals Frame
        commercial_totals_frame = ttk.Frame(commercial_frame)
        commercial_totals_frame.pack(fill=tk.X, pady=5)

        self.commercial_totals_label = ttk.Label(commercial_totals_frame, text="Grand Total: 0 | 0 | 0 | 0 | 0 | 0", 
                                            font=("Arial", 10, "bold"))
        self.commercial_totals_label.pack(side=tk.LEFT)

        # Sponsorship Benefits Frame
        sponsorship_frame = ttk.LabelFrame(self.summary_report_tab, text="Sponsorship Benefits", padding=10)
        sponsorship_frame.grid(row=4, column=0, columnspan=4, sticky="nsew", pady=5)

        # Sponsorship Benefits Treeview
        self.sponsorship_tree = ttk.Treeview(sponsorship_frame, columns=(
            'product', 'duration', 'planned', 'aired', 'missed', 'extra', 'third_party'
        ), show='headings', height=6)

        # Configure Sponsorship tree columns
        self.sponsorship_tree.heading('product', text='PRODUCT')
        self.sponsorship_tree.heading('duration', text='DUR')
        self.sponsorship_tree.heading('planned', text='PLANNED')
        self.sponsorship_tree.heading('aired', text='AIRED')
        self.sponsorship_tree.heading('missed', text='MISSED')
        self.sponsorship_tree.heading('extra', text='EXTRA')
        self.sponsorship_tree.heading('third_party', text='3RD PARTY')

        # Set sponsorship column widths
        self.sponsorship_tree.column('product', width=200)
        self.sponsorship_tree.column('duration', width=60)
        self.sponsorship_tree.column('planned', width=80)
        self.sponsorship_tree.column('aired', width=80)
        self.sponsorship_tree.column('missed', width=80)
        self.sponsorship_tree.column('extra', width=80)
        self.sponsorship_tree.column('third_party', width=80)

        # Sponsorship scrollbars
        sponsorship_scrolly = ttk.Scrollbar(sponsorship_frame, orient="vertical", command=self.sponsorship_tree.yview)
        self.sponsorship_tree.configure(yscrollcommand=sponsorship_scrolly.set)

        sponsorship_scrolly.pack(side=tk.RIGHT, fill=tk.Y)
        self.sponsorship_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Sponsorship Totals Frame
        sponsorship_totals_frame = ttk.Frame(sponsorship_frame)
        sponsorship_totals_frame.pack(fill=tk.X, pady=5)

        self.sponsorship_totals_label = ttk.Label(sponsorship_totals_frame, text="Grand Total: 0 | 0 | 0 | 0 | 0", 
                                                font=("Arial", 10, "bold"))
        self.sponsorship_totals_label.pack(side=tk.LEFT)

        # Make the grid expandable
        self.summary_report_tab.grid_rowconfigure(3, weight=1)
        self.summary_report_tab.grid_rowconfigure(4, weight=1)
        self.summary_report_tab.grid_columnconfigure(0, weight=1)

    def find_schedule_column(self, possible_names):
        """Find column in schedule data by possible names."""
        if self.schedule_data is None:
            return None

        for name in possible_names:
            # Exact match first
            if name in self.schedule_data.columns:
                return name
            # Case insensitive search
            for col in self.schedule_data.columns:
                if name.lower() in col.lower():
                    return col
        return None

    def find_lmrb_column(self, possible_names):
        """Find column in LMRB data by possible names."""
        # Check matched data first
        matched_df = self.get_current_matched_data_df()
        if matched_df.empty:
            return None

        for name in possible_names:
            # Exact match first
            if name in matched_df.columns:
                return name
            # Case insensitive search
            for col in matched_df.columns:
                if name.lower() in col.lower():
                    return col
        return None

    def get_planned_count(self, theme, duration):
        """Get planned count from schedule data, accounting for manual matches."""
        if self.schedule_data is None or self.schedule_data.empty:
            return 0

        # Find schedule theme and duration columns
        theme_col = self.find_schedule_column(['theme', 'aadvt', 'advt_theme'])
        dur_col = self.find_schedule_column(['dur', 'duration'])

        if not theme_col or not dur_col:
            print(f"Could not find theme ({theme_col}) or duration ({dur_col}) columns in schedule data")
            return 0

        try:
            # Filter and count original schedule
            filtered = self.schedule_data[
                (self.schedule_data[theme_col].astype(str).str.strip() == str(theme).strip()) & 
                (self.schedule_data[dur_col].astype(str).str.strip() == str(duration).strip())
            ]
            
            original_count = len(filtered)
            
            # NEW: Subtract manually matched schedule items
            manually_used_count = 0
            if hasattr(self, 'manually_matched_schedule_items'):
                for match_record in self.manually_matched_schedule_items:
                    if (match_record['schedule_theme'] == theme):
                        manually_used_count += match_record['schedule_count_used']
                        print(f"Subtracting {match_record['schedule_count_used']} from {theme} (manually matched)")
            
            remaining_count = max(0, original_count - manually_used_count)
            
            print(f"Planned count for {theme} ({duration}s): {original_count} original - {manually_used_count} manually used = {remaining_count}")
            return remaining_count
            
        except Exception as e:
            print(f"Error calculating planned count for {theme}: {e}")
            return 0

    def get_aired_count(self, theme, duration, matched_data):
        """Get aired count from matched LMRB data."""
        if matched_data is None or matched_data.empty:
            return 0

        # Find LMRB theme and duration columns
        theme_col = self.find_lmrb_column(['Advt_Theme', 'theme', 'Product Details'])
        dur_col = self.find_lmrb_column(['Dur', 'duration'])

        if not theme_col or not dur_col:
            print(f"Could not find theme ({theme_col}) or duration ({dur_col}) columns in matched data")
            return 0

        try:
            # Filter and count
            filtered = matched_data[
                (matched_data[theme_col].astype(str).str.strip() == str(theme).strip()) & 
                (matched_data[dur_col].astype(str).str.strip() == str(duration).strip())
            ]
            
            count = len(filtered)
            print(f"Aired count for {theme} ({duration}s): {count}")
            return count
            
        except Exception as e:
            print(f"Error calculating aired count for {theme}: {e}")
            return 0

    def calculate_average_30_sec(self, planned_count, duration):
        """Calculate average 30 second equivalent."""
        try:
            duration_num = float(duration)
            if duration_num == 0:
                return 0
            result = (planned_count * duration_num) / 30
            return round(result, 2)
        except (ValueError, ZeroDivisionError):
            return 0

    def generate_summary_report(self):
        """
        ENHANCED: Generate summary report with proper ad type categorization
        """
        print("\n" + "="*80)
        print("🚀 ENHANCED: Generating Summary Report with Schedule Ad Type Tracking")
        print("="*80)

        commercial_benefits = []
        sponsorship_benefits = []

        # Get current matched data
        try:
            matched_lmrb_data = self.get_current_matched_data_df()
            print(f"📊 Got matched LMRB data: {len(matched_lmrb_data)} records")
            
            if matched_lmrb_data.empty:
                print("⚠️ No matched LMRB data available")
            else:
                lmrb_theme_col = self._find_column(matched_lmrb_data, ['Advt_Theme', 'theme', 'Product Details'])
                print(f"   LMRB theme column: {lmrb_theme_col}")
                if lmrb_theme_col:
                    unique_themes = matched_lmrb_data[lmrb_theme_col].nunique()
                    print(f"   Unique LMRB themes in matched data: {unique_themes}")
                    
        except Exception as e:
            print(f"❌ Error getting matched data: {e}")
            matched_lmrb_data = pd.DataFrame()

        # Update schedule mapped themes before processing
        print("\n🔄 Updating schedule mapped themes...")
        self.update_schedule_mapped_themes()

        # Create grouped schedule DataFrame
        print("\n📊 Creating grouped schedule DataFrame...")
        grouped_schedule = self.create_grouped_schedule_df()
        if grouped_schedule is None:
            print("❌ No grouped schedule data available - using fallback method")
            return self._generate_summary_from_mappings_only(matched_lmrb_data)

        print(f"✅ Processing {len(grouped_schedule)} grouped schedule entries")

        # Find required columns in grouped schedule
        schedule_theme_col = self._find_schedule_column(['theme', 'aadvt', 'advt_theme'])
        schedule_dur_col = self._find_schedule_column(['dur', 'duration'])
        schedule_adtype_col = self._find_schedule_column(['advertisement_type', 'ad_type'])

        if not schedule_adtype_col:
            schedule_adtype_col = 'temp_ad_type'  # Use temp column if created

        # Process each grouped schedule entry
        print(f"\n📋 Processing grouped schedule entries:")
        print("-" * 60)

        for idx, group in grouped_schedule.iterrows():
            try:
                # Extract data from grouped schedule
                schedule_theme = group[schedule_theme_col]
                schedule_duration = str(group[schedule_dur_col])
                planned_count = group['planned_count']
                mapped_theme = group['mapped_theme']
                has_mapping = group['has_mapping']

                print(f"\n🔍 [{idx+1}/{len(grouped_schedule)}] Processing: '{schedule_theme}' ({schedule_duration}s)")
                print(f"   📋 Planned: {planned_count}")
                print(f"   🗺️ Mapped to: '{mapped_theme}' {'✅' if has_mapping else '❌'}")

                # FIXED: Get schedule ad type from mapping first
                schedule_ad_type = self._get_schedule_ad_type(schedule_theme)
                print(f"   🏷️ Schedule Ad Type: {schedule_ad_type}")

                # FIXED: Set current processing context for filtering
                self._current_schedule_ad_type = schedule_ad_type
                self._current_schedule_theme = schedule_theme

                # Calculate aired count using mapped theme (now filtered by ad type)
                total_aired_count = 0
                automatic_count = 0
                manual_count = 0
                
                if mapped_theme and has_mapping:
                    # Count from matched LMRB data using mapped theme and ad type filter
                    automatic_count = self._get_aired_count_by_mapped_theme(
                        mapped_theme, schedule_duration, matched_lmrb_data
                    )
                    print(f"   📺 Automatic aired (filtered by ad type): {automatic_count}")
                    
                # Add manual matches (always check, even without mapping)
                manual_count = self._get_manual_match_count(schedule_theme, schedule_duration)
                if manual_count > 0:
                    print(f"   🖐️ Manual matches: {manual_count}")
                
                # Calculate consumption-based matches
                consumption_count = self._get_consumption_match_count(schedule_theme, schedule_duration)
                if consumption_count > 0:
                    print(f"   🔄 Consumption matches: {consumption_count}")
                
                total_aired_count = automatic_count + manual_count + consumption_count
                print(f"   🎯 Total aired: {total_aired_count}")

                # Clear processing context
                self._current_schedule_ad_type = None
                self._current_schedule_theme = None

                # Calculate metrics
                missed_count = max(0, planned_count - total_aired_count)
                extra_count = max(0, total_aired_count - planned_count)
                third_party_count = total_aired_count
                average_30_sec = self.calculate_average_30_sec(planned_count, schedule_duration)
                
                print(f"   📊 Missed: {missed_count}, Extra: {extra_count}, 30sec: {average_30_sec:.1f}")

                # Create report row
                report_row = {
                    'product': schedule_theme,
                    'duration': schedule_duration,
                    'planned': planned_count,
                    'aired': total_aired_count,
                    'missed': missed_count,
                    'extra': extra_count,
                    'third_party': third_party_count,
                    'average_30_sec': average_30_sec,
                    'has_mapping': has_mapping,
                    'mapped_theme': mapped_theme,
                    'automatic_count': automatic_count,
                    'manual_count': manual_count,
                    'consumption_count': consumption_count,
                    'schedule_ad_type': schedule_ad_type,
                    'mapping_status': 'MAPPED' if has_mapping else 'NO MAPPING'
                }

                # FIXED: Categorize by SCHEDULE ad type (not mapping type)
                if 'SPONSORSHIP' in str(schedule_ad_type).upper():
                    sponsorship_benefits.append(report_row)
                    print(f"   📋 ✅ Added to Sponsorship Benefits")
                else:
                    commercial_benefits.append(report_row)
                    print(f"   📋 ✅ Added to Commercial Benefits")

            except Exception as e:
                print(f"❌ Error processing grouped schedule entry {idx}: {e}")
                import traceback
                traceback.print_exc()
                continue

        print(f"\n" + "="*60)
        print("🎉 ENHANCED Summary Report Results:")
        print("="*60)
        print(f"📋 Commercial Benefits: {len(commercial_benefits)} items")
        print(f"📋 Sponsorship Benefits: {len(sponsorship_benefits)} items")

        # Calculate totals
        if commercial_benefits:
            c_planned = sum(item['planned'] for item in commercial_benefits)
            c_aired = sum(item['aired'] for item in commercial_benefits)
            c_auto = sum(item['automatic_count'] for item in commercial_benefits)
            c_manual = sum(item['manual_count'] for item in commercial_benefits)
            print(f"📊 Commercial Totals - Planned: {c_planned}, Aired: {c_aired} (Auto: {c_auto}, Manual: {c_manual})")

        if sponsorship_benefits:
            s_planned = sum(item['planned'] for item in sponsorship_benefits)
            s_aired = sum(item['aired'] for item in sponsorship_benefits)
            s_auto = sum(item['automatic_count'] for item in sponsorship_benefits)
            s_manual = sum(item['manual_count'] for item in sponsorship_benefits)
            print(f"📊 Sponsorship Totals - Planned: {s_planned}, Aired: {s_aired} (Auto: {s_auto}, Manual: {s_manual})")

        print("="*60)

        return commercial_benefits, sponsorship_benefits

    def _get_aired_count_by_mapped_theme(self, mapped_theme, duration, matched_lmrb_data):
        """FIXED: Get aired count and exclude already consumed records."""
        if matched_lmrb_data is None or matched_lmrb_data.empty or not mapped_theme:
            return 0

        # Find LMRB theme and duration columns
        lmrb_theme_col = self._find_column(matched_lmrb_data, ['Advt_Theme', 'theme', 'Product Details'])
        lmrb_dur_col = self._find_column(matched_lmrb_data, ['Dur', 'duration'])

        if not lmrb_theme_col or not lmrb_dur_col:
            print(f"     ❌ Could not find LMRB columns - Theme: {lmrb_theme_col}, Duration: {lmrb_dur_col}")
            return 0

        try:
            print(f"     🔍 Searching for LMRB theme: '{mapped_theme}' with duration: {duration}s")
            
            # Filter by mapped theme and duration
            theme_match = matched_lmrb_data[lmrb_theme_col].astype(str).str.strip() == str(mapped_theme).strip()
            duration_match = matched_lmrb_data[lmrb_dur_col].astype(str).str.strip() == str(duration).strip()
            
            filtered = matched_lmrb_data[theme_match & duration_match]
            original_count = len(filtered)
            
            print(f"     📊 Total LMRB records for theme: {original_count}")
            
            # FIXED: Exclude records that were consumed through schedule matching
            if hasattr(self, '_current_schedule_ad_type') and self._current_schedule_ad_type:
                
                # Separate automatic matches from manual/consumption matches
                automatic_matches = filtered[
                    (~filtered.get('Match_Type', pd.Series(dtype='object')).isin([
                        'Dual-Consumption Manual Match', 
                        'Count-Based Manual Match',
                        'Manual'
                    ])) | 
                    (filtered.get('Match_Type', pd.Series(dtype='object')).isna())
                ]
                
                # Manual/consumption matches with specific ad type
                manual_matches = filtered[
                    (filtered.get('Match_Type', pd.Series(dtype='object')).isin([
                        'Dual-Consumption Manual Match', 
                        'Count-Based Manual Match',
                        'Manual'
                    ])) &
                    (filtered.get('Schedule_Ad_Type', pd.Series(dtype='object')) == self._current_schedule_ad_type)
                ]
                
                print(f"     📊 Automatic matches: {len(automatic_matches)}")
                print(f"     📊 Manual matches with ad type '{self._current_schedule_ad_type}': {len(manual_matches)}")
                
                # For automatic matches, only include if current schedule mapping type matches
                if len(automatic_matches) > 0 and hasattr(self, '_current_schedule_theme'):
                    schedule_mapping_type = self._get_schedule_ad_type(self._current_schedule_theme)
                    
                    if schedule_mapping_type == self._current_schedule_ad_type:
                        # Check if this LMRB theme is being consumed by other schedule themes
                        consumed_by_others = self._check_lmrb_theme_consumed_by_others(
                            mapped_theme, duration, self._current_schedule_theme
                        )
                        
                        if consumed_by_others > 0:
                            # Reduce automatic matches by the amount consumed by others
                            adjusted_automatic = max(0, len(automatic_matches) - consumed_by_others)
                            print(f"     🔄 Adjusted automatic matches: {len(automatic_matches)} - {consumed_by_others} (consumed by others) = {adjusted_automatic}")
                            
                            # Take only the adjusted amount
                            if adjusted_automatic > 0:
                                automatic_matches = automatic_matches.head(adjusted_automatic)
                            else:
                                automatic_matches = pd.DataFrame()  # Empty
                        
                        # Combine automatic and manual matches
                        filtered = pd.concat([automatic_matches, manual_matches])
                    else:
                        # Only include manual matches
                        filtered = manual_matches
                else:
                    # Only manual matches
                    filtered = manual_matches
                
                final_count = len(filtered)
                print(f"     🎯 Final count after filtering: {original_count} → {final_count}")
            else:
                final_count = original_count
            
            count = len(filtered)
            
            if count > 0:
                print(f"     ✅ Found {count} aired records for mapped theme: '{mapped_theme}' ({duration}s)")
            else:
                print(f"     ⚠️ No aired records found for mapped theme: '{mapped_theme}' ({duration}s)")
            
            return count
            
        except Exception as e:
            print(f"     ❌ Error getting aired count for mapped theme {mapped_theme}: {e}")
            import traceback
            traceback.print_exc()
            return 0

    def _check_lmrb_theme_consumed_by_others(self, lmrb_theme, duration, current_schedule_theme):
        """Check how many records of this LMRB theme were consumed by other schedule themes."""
        try:
            consumed_count = 0
            
            # Check consumption tracker
            if hasattr(self, 'schedule_consumption_tracker'):
                for consumption_key, consumed_amount in self.schedule_consumption_tracker.items():
                    try:
                        # Parse consumption key: schedule_theme_duration_adtype
                        parts = consumption_key.split('_')
                        if len(parts) >= 2:
                            key_schedule_theme = '_'.join(parts[:-2])  # Handle themes with underscores
                            key_duration = parts[-2]
                            
                            # Skip if this is the current schedule theme
                            if key_schedule_theme == current_schedule_theme:
                                continue
                            
                            # Check if this schedule theme maps to the same LMRB theme
                            for mapping in self.theme_mapping:
                                if (mapping.get('schedule_theme') == key_schedule_theme and
                                    mapping.get('media_watch_theme', mapping.get('lmrb_theme', '')) == lmrb_theme and
                                    str(mapping.get('lmrb_duration', '')) == str(duration)):
                                    
                                    consumed_count += consumed_amount
                                    print(f"       🔍 Found consumption by '{key_schedule_theme}': {consumed_amount}")
                                    break
                    
                    except Exception as e:
                        print(f"Error parsing consumption key {consumption_key}: {e}")
                        continue
            
            print(f"     📊 Total consumed by other themes: {consumed_count}")
            return consumed_count
            
        except Exception as e:
            print(f"Error checking consumption by others: {e}")
            return 0

    def _get_consumption_match_count(self, schedule_theme, duration):
        """Get count from consumption-based matches (dual consumption, schedule consumption)"""
        try:
            consumption_count = 0
            
            # Check dual consumption matches
            if hasattr(self, 'dual_consumption_matches'):
                for match in self.dual_consumption_matches:
                    if (match.get('schedule_theme') == schedule_theme and 
                        str(match.get('schedule_duration', '')) == str(duration)):
                        consumption_count += match.get('consumption_amount', 0)
            
            # Check schedule consumption tracker
            if hasattr(self, 'schedule_consumption_tracker'):
                # Try different possible consumption keys
                possible_keys = [
                    f"{schedule_theme}_{duration}_Unknown",
                    f"{schedule_theme}_{duration}_COMMERCIAL BENEFITS", 
                    f"{schedule_theme}_{duration}_SPONSORSHIP BENEFITS"
                ]
                
                for key in possible_keys:
                    consumption_count += self.schedule_consumption_tracker.get(key, 0)
            
            return consumption_count
            
        except Exception as e:
            print(f"     ❌ Error calculating consumption matches for {schedule_theme}: {e}")
            return 0


    def _get_aired_count_exact_match(self, lmrb_theme, duration, matched_lmrb_data):
        """FIXED: Get aired count using EXACT LMRB theme match"""
        if matched_lmrb_data is None or matched_lmrb_data.empty:
            return 0

        # Find LMRB theme and duration columns
        lmrb_theme_col = self._find_column(matched_lmrb_data, ['Advt_Theme', 'theme', 'Product Details'])
        lmrb_dur_col = self._find_column(matched_lmrb_data, ['Dur', 'duration'])

        if not lmrb_theme_col or not lmrb_dur_col:
            print(f"  ❌ Could not find LMRB columns - Theme: {lmrb_theme_col}, Duration: {lmrb_dur_col}")
            return 0

        try:
            # FIXED: Use exact string matching (strip whitespace)
            filtered = matched_lmrb_data[
                (matched_lmrb_data[lmrb_theme_col].astype(str).str.strip() == str(lmrb_theme).strip()) & 
                (matched_lmrb_data[lmrb_dur_col].astype(str).str.strip() == str(duration).strip())
            ]
            
            count = len(filtered)
            if count > 0:
                print(f"  ✅ Found {count} aired records for LMRB theme: '{lmrb_theme}' ({duration}s)")
            else:
                print(f"  ⚠️ No aired records found for LMRB theme: '{lmrb_theme}' ({duration}s)")
            
            return count
            
        except Exception as e:
            print(f"  ❌ Error calculating aired count for LMRB theme {lmrb_theme}: {e}")
            return 0

    def _get_manual_match_count(self, schedule_theme, duration):
        """FIXED: Get count of manual matches for this schedule theme"""
        try:
            manual_count = 0
            
            # Check dual consumption matches
            if hasattr(self, 'dual_consumption_matches'):
                for match in self.dual_consumption_matches:
                    if (match.get('schedule_theme') == schedule_theme and 
                        str(match.get('schedule_duration', '')) == str(duration)):
                        manual_count += match.get('consumption_amount', 0)
                        print(f"  📊 Found dual consumption match: {match.get('consumption_amount', 0)}")
            
            # Check other manual matches
            if hasattr(self, 'manually_matched_schedule_items'):
                for match in self.manually_matched_schedule_items:
                    if (match.get('schedule_theme') == schedule_theme and 
                        str(match.get('schedule_duration', '')) == str(duration)):
                        manual_count += match.get('schedule_count_consumed', 0)
                        print(f"  📊 Found manual schedule match: {match.get('schedule_count_consumed', 0)}")
            
            if manual_count > 0:
                print(f"  ✅ Total manual matches for {schedule_theme}: {manual_count}")
            
            return manual_count
            
        except Exception as e:
            print(f"  ❌ Error calculating manual matches for {schedule_theme}: {e}")
            return 0

    def _generate_summary_from_mappings_only(self, matched_lmrb_data):
        """Fallback method when no schedule data is available"""
        print("⚠️ Using fallback: generating summary from theme mappings only")

        commercial_benefits = []
        sponsorship_benefits = []

        for mapping in self.theme_mapping:
            try:
                mapping_type = mapping.get('mapping_type', 'COMMERCIAL BENEFITS')
                schedule_theme = mapping.get('schedule_theme', '')
                schedule_duration = mapping.get('schedule_duration', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                
                if not schedule_theme or not lmrb_theme:
                    continue
                
                # No planned count available - use 0
                planned_count = 0
                aired_count = self.get_aired_count(lmrb_theme, schedule_duration, matched_lmrb_data)
                
                report_row = {
                    'product': schedule_theme,
                    'duration': schedule_duration,
                    'planned': planned_count,
                    'aired': aired_count,
                    'missed': 0,
                    'extra': aired_count,  # All are "extra" since no planned count
                    'third_party': aired_count,
                    'average_30_sec': 0,
                    'has_mapping': True,
                    'mapping_status': 'MAPPING ONLY'
                }
                
                if 'SPONSORSHIP' in mapping_type.upper():
                    sponsorship_benefits.append(report_row)
                else:
                    commercial_benefits.append(report_row)
                    
            except Exception as e:
                print(f"Error processing mapping: {e}")
                continue

        return commercial_benefits, sponsorship_benefits

    def _determine_ad_type_from_schedule(self, ad_type_value):
        """Determine if schedule entry is commercial or sponsorship based on ad_type column"""
        if pd.isna(ad_type_value):
            return 'COMMERCIAL BENEFITS'  # Default

        ad_type_str = str(ad_type_value).lower()

        if 'sponsor' in ad_type_str:
            return 'SPONSORSHIP BENEFITS'
        elif 'commercial' in ad_type_str:
            return 'COMMERCIAL BENEFITS'
        else:
            return 'COMMERCIAL BENEFITS'  # Default

    def _find_schedule_column(self, possible_names):
        """Find column in schedule data by possible names"""
        if not hasattr(self, 'schedule_data') or self.schedule_data is None:
            return None
            
        for name in possible_names:
            # Exact match first
            if name in self.schedule_data.columns:
                return name
            # Case insensitive search
            for col in self.schedule_data.columns:
                if name.lower() in col.lower():
                    return col
        return None

    def generate_and_display_summary(self):
        """Generate and display the summary report."""
        try:
            # Update channel label
            if hasattr(self, 'report_channel_label'):
                self.report_channel_label.config(text=self.selected_channel or "No Channel Selected")
            
            # Generate report data
            commercial_data, sponsorship_data = self.generate_summary_report()
            
            # Update displays
            self.update_summary_trees(commercial_data, sponsorship_data)
            
            # Update totals
            self.update_summary_totals(commercial_data, sponsorship_data)
            
            messagebox.showinfo("Success", 
                f"Summary report generated successfully!\n\n"
                f"Commercial Benefits: {len(commercial_data)} items\n"
                f"Sponsorship Benefits: {len(sponsorship_data)} items")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generating summary report: {str(e)}")
            import traceback
            traceback.print_exc()

    def update_summary_trees(self, commercial_data, sponsorship_data):
        """
        ENHANCED: Update summary trees with mapping status indicators
        """
        # Clear existing data
        for item in self.commercial_tree.get_children():
            self.commercial_tree.delete(item)
        for item in self.sponsorship_tree.get_children():
            self.sponsorship_tree.delete(item)

        # Populate commercial tree with mapping status
        for item in commercial_data:
            # Add visual indicator for mapping status
            product_name = item['product']
            if not item.get('has_mapping', True):
                product_name = f"🔸 {product_name}"  # Indicator for no mapping
            
            self.commercial_tree.insert('', 'end', values=(
                product_name,
                item['duration'], 
                item['planned'],
                item['aired'], 
                item['missed'], 
                item['extra'],
                item['third_party'], 
                f"{item['average_30_sec']:.1f}"
            ))

        # Populate sponsorship tree with mapping status
        for item in sponsorship_data:
            # Add visual indicator for mapping status
            product_name = item['product']
            if not item.get('has_mapping', True):
                product_name = f"🔸 {product_name}"  # Indicator for no mapping
            
            self.sponsorship_tree.insert('', 'end', values=(
                product_name,
                item['duration'], 
                item['planned'],
                item['aired'], 
                item['missed'], 
                item['extra'],
                item['third_party']
            ))

    def update_summary_totals(self, commercial_data, sponsorship_data):
        """
        ENHANCED: Update totals with mapping status information
        """
        # Calculate commercial totals
        c_planned = sum(item['planned'] for item in commercial_data)
        c_aired = sum(item['aired'] for item in commercial_data)
        c_missed = sum(item['missed'] for item in commercial_data)
        c_extra = sum(item['extra'] for item in commercial_data)
        c_third_party = sum(item['third_party'] for item in commercial_data)
        c_avg_30_sec = sum(item['average_30_sec'] for item in commercial_data)

        # Count mapping status
        c_mapped = len([x for x in commercial_data if x.get('has_mapping', True)])
        c_unmapped = len(commercial_data) - c_mapped

        self.commercial_totals_label.config(
            text=f"Grand Total: {c_planned} | {c_aired} | {c_missed} | {c_extra} | {c_third_party} | {c_avg_30_sec:.1f} | Mapped: {c_mapped} | Unmapped: {c_unmapped}"
        )

        # Calculate sponsorship totals
        s_planned = sum(item['planned'] for item in sponsorship_data)
        s_aired = sum(item['aired'] for item in sponsorship_data)
        s_missed = sum(item['missed'] for item in sponsorship_data)
        s_extra = sum(item['extra'] for item in sponsorship_data)
        s_third_party = sum(item['third_party'] for item in sponsorship_data)

        # Count mapping status
        s_mapped = len([x for x in sponsorship_data if x.get('has_mapping', True)])
        s_unmapped = len(sponsorship_data) - s_mapped

        self.sponsorship_totals_label.config(
            text=f"Grand Total: {s_planned} | {s_aired} | {s_missed} | {s_extra} | {s_third_party} | Mapped: {s_mapped} | Unmapped: {s_unmapped}"
        )

    def sync_summary_report_data(self):
        """Sync summary report when data changes in other tabs."""
        try:
            print("Syncing summary report data...")
            
            # Check if we have the necessary data
            if not hasattr(self, 'commercial_tree') or not hasattr(self, 'sponsorship_tree'):
                print("Summary report trees not initialized yet")
                return
            
            # Generate and display updated data
            self.generate_and_display_summary()
            
            print("Summary report data synchronized")
            
        except Exception as e:
            print(f"Error syncing summary report data: {e}")

    def export_summary_report(self):
        """Export summary report to Excel with proper formatting matching the template."""
        try:
            # Check if we have data
            if not hasattr(self, 'commercial_tree') or not hasattr(self, 'sponsorship_tree'):
                messagebox.showerror("Error", "Please generate the summary report first")
                return
            
            # Generate current data - FIXED: Ensure data is available
            commercial_data, sponsorship_data = self.generate_summary_report()
            
            # Validate data exists
            if not commercial_data and not sponsorship_data:
                messagebox.showerror("Error", "No data available to export")
                return
            
            # Get file path
            default_name = f"Summary_Report_{self.selected_channel or 'Channel'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not save_path:
                return
            
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"
            
            # === HEADER SECTION ===
            # Company name (merged cells)
            ws.merge_cells('A1:H1')
            ws['A1'] = "MOBITEL (PVT) LTD"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Report title (merged cells)
            ws.merge_cells('A2:H2')
            ws['A2'] = "SUMMARY SHEET REPORT"
            ws['A2'].font = Font(bold=True, size=12)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Company details (right side)
            ws['H4'] = "Phoenix O & M (Pvt) Ltd"
            ws['H4'].font = Font(bold=True)
            ws['H5'] = "No 16, Barnes Place"
            ws['H6'] = "Colombo 7"
            
            # Report details (left side)
            current_row = 4
            
            # Month
            ws[f'A{current_row}'] = "MONTH"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.report_month_var.get()
            current_row += 1
            
            # Channel
            ws[f'A{current_row}'] = "CHANNEL"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.selected_channel or ""
            current_row += 1
            
            # Estimate No
            ws[f'A{current_row}'] = "ESTIMATE NO"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.estimate_no_var.get()
            current_row += 1
            
            # Supplier Invoice No
            ws[f'A{current_row}'] = "SUPPLIER INVOICE NO"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.supplier_invoice_var.get()
            current_row += 1
            
            # PO No
            ws[f'A{current_row}'] = "PO NO"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.po_no_var.get()
            current_row += 1
            
            # Invoice No
            ws[f'A{current_row}'] = "INVOICE NO"
            ws[f'B{current_row}'] = ":"
            ws[f'C{current_row}'] = self.invoice_no_var.get()
            current_row += 2
            
            # === TABLE SECTION ===
            # Main header row
            ws[f'A{current_row}'] = "PRODUCT"
            ws[f'B{current_row}'] = "DUR"
            ws.merge_cells(f'C{current_row}:G{current_row}')
            ws[f'C{current_row}'] = "SPOTS"
            ws[f'H{current_row}'] = "Average 30 sec."
            
            # Make header row bold
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Sub-header row
            ws[f'A{current_row}'] = ""
            ws[f'B{current_row}'] = ""
            ws[f'C{current_row}'] = "PLANNED"
            ws[f'D{current_row}'] = "AIRED"
            ws[f'E{current_row}'] = "MISSED"
            ws[f'F{current_row}'] = "EXTRA"
            ws[f'G{current_row}'] = "3RD PARTY"
            ws[f'H{current_row}'] = ""
            
            # Make sub-header row bold
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            start_data_row = current_row
            
            # === COMMERCIAL BENEFITS DATA ===
            print(f"Exporting {len(commercial_data)} commercial items")
            
            # Calculate commercial totals
            c_planned = sum(item['planned'] for item in commercial_data) if commercial_data else 0
            c_aired = sum(item['aired'] for item in commercial_data) if commercial_data else 0
            c_missed = sum(item['missed'] for item in commercial_data) if commercial_data else 0
            c_extra = sum(item['extra'] for item in commercial_data) if commercial_data else 0
            c_third_party = sum(item['third_party'] for item in commercial_data) if commercial_data else 0
            c_avg_30_sec = sum(item['average_30_sec'] for item in commercial_data) if commercial_data else 0
            
            # Write commercial data
            for item in commercial_data:
                ws[f'A{current_row}'] = item['product']
                ws[f'B{current_row}'] = item['duration']
                ws[f'C{current_row}'] = item['planned']
                ws[f'D{current_row}'] = item['aired']
                ws[f'E{current_row}'] = item['missed']
                ws[f'F{current_row}'] = item['extra']
                ws[f'G{current_row}'] = item['third_party']
                ws[f'H{current_row}'] = round(item['average_30_sec'], 1)
                current_row += 1
            
            # Commercial Grand Total
            ws[f'A{current_row}'] = "Grand Total"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'C{current_row}'] = c_planned
            ws[f'D{current_row}'] = c_aired
            ws[f'E{current_row}'] = c_missed
            ws[f'F{current_row}'] = c_extra
            ws[f'G{current_row}'] = c_third_party
            ws[f'H{current_row}'] = round(c_avg_30_sec, 1)
            
            # Make totals row bold
            for col in range(1, 9):
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            
            current_row += 2
            
            # === SPONSORSHIP BENEFITS SECTION ===
            print(f"Exporting {len(sponsorship_data)} sponsorship items")
            
            ws[f'A{current_row}'] = "Sponsorship Benefits"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # Calculate sponsorship totals
            s_planned = sum(item['planned'] for item in sponsorship_data) if sponsorship_data else 0
            s_aired = sum(item['aired'] for item in sponsorship_data) if sponsorship_data else 0
            s_missed = sum(item['missed'] for item in sponsorship_data) if sponsorship_data else 0
            s_extra = sum(item['extra'] for item in sponsorship_data) if sponsorship_data else 0
            s_third_party = sum(item['third_party'] for item in sponsorship_data) if sponsorship_data else 0
            
            # Write sponsorship data
            for item in sponsorship_data:
                ws[f'A{current_row}'] = item['product']
                ws[f'B{current_row}'] = item['duration']
                ws[f'C{current_row}'] = item['planned']
                ws[f'D{current_row}'] = item['aired']
                ws[f'E{current_row}'] = item['missed']
                ws[f'F{current_row}'] = item['extra']
                ws[f'G{current_row}'] = item['third_party']
                current_row += 1
            
            # Sponsorship Grand Total
            ws[f'A{current_row}'] = "Grand Total"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'C{current_row}'] = s_planned
            ws[f'D{current_row}'] = s_aired
            ws[f'E{current_row}'] = s_missed
            ws[f'F{current_row}'] = s_extra
            ws[f'G{current_row}'] = s_third_party
            
            # Make sponsorship totals row bold
            for col in range(1, 8):  # Only 7 columns for sponsorship
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            
            current_row += 3
            
            # === FOOTER SECTION ===
            # Note section
            ws[f'A{current_row}'] = "NOTE:"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 2
            
            # Signature lines
            ws[f'A{current_row}'] = "…………………………………….."
            ws[f'D{current_row}'] = "…………………………………….."
            ws[f'G{current_row}'] = "…………………………………….."
            current_row += 1
            
            ws[f'A{current_row}'] = "PREPARED BY"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'D{current_row}'] = "CHECKED BY"
            ws[f'D{current_row}'].font = Font(bold=True)
            ws[f'G{current_row}'] = "AUTHORISED BY"
            ws[f'G{current_row}'].font = Font(bold=True)
            
            # === FORMATTING ===
            # Apply borders to the data table
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply borders to data section
            for row in range(start_data_row - 2, current_row - 5):
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Set column widths
            column_widths = {
                'A': 25,  # PRODUCT
                'B': 8,   # DUR
                'C': 10,  # PLANNED
                'D': 10,  # AIRED
                'E': 10,  # MISSED
                'F': 10,  # EXTRA
                'G': 12,  # 3RD PARTY
                'H': 15   # Average 30 sec
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Save the workbook
            wb.save(save_path)
            
            # Show success message
            total_items = len(commercial_data) + len(sponsorship_data)
            messagebox.showinfo("Export Successful", 
                f"Summary report exported successfully!\n\n"
                f"File: {save_path}\n"
                f"Commercial Benefits: {len(commercial_data)} items\n"
                f"Sponsorship Benefits: {len(sponsorship_data)} items\n"
                f"Total Items: {total_items}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting summary report: {str(e)}")
            import traceback
            traceback.print_exc()




if __name__ == "__main__":
    from src.splash_screen import SplashScreen
    import os
    import sys
    from PIL import Image, ImageTk

    # Function to get resource path (works for both dev and exe)
    def resource_path(relative_path):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
        
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

# Path to your splash image
splash_img_path = resource_path("assets/splash.png")

if not os.path.exists(splash_img_path):
    # fallback: use a default image from PIL if not found
    try:
        from PIL import ImageDraw
        img = Image.new('RGB', (400, 220), color="#b7d700")
        d = ImageDraw.Draw(img)
        d.text((100, 100), "Loading...", fill=(255,255,255))
        # Save to current directory as fallback
        fallback_path = "temp_splash.png"
        img.save(fallback_path)
        splash_img_path = fallback_path
    except Exception as e:
        print(f"Could not create fallback image: {e}")
        splash_img_path = None

root = tk.Tk()
root.withdraw()  # Hide main window during splash

# Only show splash if image is available
if splash_img_path and os.path.exists(splash_img_path):
    splash = SplashScreen(root, splash_img_path, load_time=2)
    root.wait_window(splash)

root.deiconify()
app = MediaReconciliationTool(root)
root.mainloop()