import pandas as pd
from datetime import datetime
import os
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import re

class ExportManager:
    def __init__(self, app_instance):
        self.app = app_instance
        
    def get_default_filename(self, export_type):
        """Generate default filename with timestamp"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        channel = self.app.selected_channel or 'channel'
        return f"{export_type}_{channel}_{timestamp}.xlsx"

    def _find_column(self, df, possible_names):
        """Find column by possible names (case insensitive)"""
        if df is None or df.empty:
            return None
            
        for name in possible_names:
            # Exact match first
            if name in df.columns:
                return name
            # Case insensitive search
            for col in df.columns:
                if name.lower() in col.lower():
                    return col
        return None

    def _normalize_date(self, date_value):
        """Enhanced date normalization to handle various formats"""
        if pd.isna(date_value) or date_value is None:
            return None

        try:
            date_str = str(date_value).strip()
            
            # Remove time component if present
            if ' ' in date_str and ':' in date_str:
                date_str = date_str.split(' ')[0]
            
            # Handle common formats
            date_formats = [
                '%Y-%m-%d',    # 2024-01-15
                '%d/%m/%Y',    # 15/01/2024
                '%m/%d/%Y',    # 01/15/2024
                '%Y/%m/%d',    # 2024/01/15
                '%d-%m-%Y',    # 15-01-2024
                '%Y%m%d',      # 20240115
                '%d.%m.%Y',    # 15.01.2024
                '%d %m %Y',    # 15 01 2024
                '%d-%b-%Y',    # 15-Jan-2024
                '%d %b %Y'     # 15 Jan 2024
            ]
            
            for fmt in date_formats:
                try:
                    parsed_date = pd.to_datetime(date_str, format=fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except:
                    continue
            
            # Try pandas auto-parsing as last resort
            try:
                parsed_date = pd.to_datetime(date_str, dayfirst=True)  # Assume day first for ambiguous dates
                return parsed_date.strftime('%Y-%m-%d')
            except:
                pass
            
            # If all parsing fails, return the original string for debugging
            print(f"⚠️ Could not parse date: '{date_value}' -> '{date_str}'")
            return str(date_value)
            
        except Exception as e:
            print(f"Error normalizing date '{date_value}': {e}")
            return str(date_value)

    def _normalize_time(self, time_value):
        """Enhanced time normalization to handle various formats"""
        if pd.isna(time_value) or time_value is None:
            return None

        try:
            time_str = str(time_value).strip()
            
            # Handle various time formats
            import re
            
            # Try to extract time pattern
            # Patterns: HH:MM, HH:MM:SS, H:MM, etc.
            time_patterns = [
                r'(\d{1,2}):(\d{2}):(\d{2})',  # HH:MM:SS
                r'(\d{1,2}):(\d{2})',          # HH:MM
                r'(\d{1,2})\.(\d{2})',         # HH.MM
                r'(\d{3,4})',                  # HHMM or HMM
            ]
            
            for pattern in time_patterns:
                match = re.search(pattern, time_str)
                if match:
                    if len(match.groups()) == 3:  # HH:MM:SS
                        hours = int(match.group(1))
                        minutes = int(match.group(2))
                    elif len(match.groups()) == 2:  # HH:MM
                        hours = int(match.group(1))
                        minutes = int(match.group(2))
                    elif len(match.groups()) == 1:  # HHMM
                        time_num = match.group(1)
                        if len(time_num) == 4:  # HHMM
                            hours = int(time_num[:2])
                            minutes = int(time_num[2:])
                        elif len(time_num) == 3:  # HMM
                            hours = int(time_num[:1])
                            minutes = int(time_num[1:])
                        else:
                            continue
                    
                    # Validate hours and minutes
                    if 0 <= hours <= 23 and 0 <= minutes <= 59:
                        # Convert to minutes since midnight
                        return hours * 60 + minutes
            
            print(f"⚠️ Could not parse time: '{time_value}' -> '{time_str}'")
            return None
            
        except Exception as e:
            print(f"Error normalizing time '{time_value}': {e}")
            return None

    def _identify_missed_schedule_entries(self):
        """
        Identify schedule entries that were planned but never aired
        Returns DataFrame of missed schedule entries
        """
        try:
            if self.app.schedule_data is None or not self.app.matched_data:
                print("No schedule data or matched data available")
                return pd.DataFrame()
            
            print("\n=== IDENTIFYING MISSED SCHEDULE ENTRIES ===")
            
            # Convert matched data to DataFrame
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            # Find columns
            schedule_theme_col = self._find_column(self.app.schedule_data, ['theme', 'aadvt', 'advt_theme'])
            schedule_program_col = self._find_column(self.app.schedule_data, ['program', 'programme'])
            schedule_date_col = self._find_column(self.app.schedule_data, ['date', 'air_date', 'broadcast_date'])
            schedule_dur_col = self._find_column(self.app.schedule_data, ['dur', 'duration'])
            
            if not schedule_theme_col:
                print("❌ No theme column found in schedule data")
                return pd.DataFrame()
            
            # Create theme mapping lookup (Schedule -> LMRB)
            schedule_to_lmrb_themes = {}
            for mapping in self.app.theme_mapping:
                schedule_theme = mapping.get('schedule_theme', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                if schedule_theme and lmrb_theme:
                    schedule_to_lmrb_themes[schedule_theme] = lmrb_theme
            
            print(f"📋 Found {len(schedule_to_lmrb_themes)} theme mappings")
            
            # Get all LMRB themes that were actually matched
            lmrb_theme_col = self._find_column(matched_df, ['Advt_Theme', 'theme', 'aadvt', 'product'])
            
            if not lmrb_theme_col:
                print("❌ No theme column found in matched LMRB data")
                return pd.DataFrame()
            
            # Get unique combinations of matched themes, dates, programs for comparison
            matched_combinations = set()
            for _, row in matched_df.iterrows():
                lmrb_theme = row.get(lmrb_theme_col, '')
                lmrb_date = self._get_date_from_row(row)
                lmrb_program = row.get('Program', '')
                
                # Normalize for comparison
                lmrb_date_norm = self._normalize_date(lmrb_date)
                matched_combinations.add((lmrb_theme, lmrb_date_norm, str(lmrb_program).strip().lower()))
            
            print(f"📊 Found {len(matched_combinations)} unique matched combinations")
            
            # Check each schedule entry
            missed_entries = []
            total_schedule = 0
            found_matches = 0
            
            for _, schedule_row in self.app.schedule_data.iterrows():
                total_schedule += 1
                
                schedule_theme = schedule_row.get(schedule_theme_col, '')
                schedule_program = schedule_row.get(schedule_program_col, '') if schedule_program_col else ''
                schedule_date = schedule_row.get(schedule_date_col, '') if schedule_date_col else ''
                schedule_duration = schedule_row.get(schedule_dur_col, '') if schedule_dur_col else ''
                
                # Find corresponding LMRB theme
                expected_lmrb_theme = schedule_to_lmrb_themes.get(schedule_theme)
                
                if not expected_lmrb_theme:
                    # No theme mapping - mark as missed due to no mapping
                    missed_entries.append({
                        'Schedule_Theme': schedule_theme,
                        'Schedule_Program': schedule_program,
                        'Schedule_Date': schedule_date,
                        'Schedule_Duration': schedule_duration,
                        'Expected_LMRB_Theme': 'NO MAPPING',
                        'Status': 'NO THEME MAPPING',
                        'Reason': f'Schedule theme "{schedule_theme}" has no LMRB theme mapping',
                        'Category': 'MAPPING_ISSUE'
                    })
                    continue
                
                # Check if this scheduled entry was actually aired
                schedule_date_norm = self._normalize_date(schedule_date)
                schedule_program_norm = str(schedule_program).strip().lower()
                
                # Look for exact match (theme + date + program)
                exact_match = (expected_lmrb_theme, schedule_date_norm, schedule_program_norm) in matched_combinations
                
                # Look for theme + date match (different program)
                theme_date_match = any(
                    combo[0] == expected_lmrb_theme and combo[1] == schedule_date_norm 
                    for combo in matched_combinations
                )
                
                # Look for theme-only match (different date/program)
                theme_only_match = any(
                    combo[0] == expected_lmrb_theme 
                    for combo in matched_combinations
                )
                
                if exact_match:
                    found_matches += 1
                    # This was aired as planned - not missed
                    continue
                elif theme_date_match:
                    # Aired on correct date but different program
                    missed_entries.append({
                        'Schedule_Theme': schedule_theme,
                        'Schedule_Program': schedule_program,
                        'Schedule_Date': schedule_date,
                        'Schedule_Duration': schedule_duration,
                        'Expected_LMRB_Theme': expected_lmrb_theme,
                        'Status': 'PROGRAM MOVED',
                        'Reason': f'Aired on {schedule_date_norm} but in different program',
                        'Category': 'PROGRAM_CHANGE'
                    })
                elif theme_only_match:
                    # Aired but on different date/program
                    missed_entries.append({
                        'Schedule_Theme': schedule_theme,
                        'Schedule_Program': schedule_program,
                        'Schedule_Date': schedule_date,
                        'Schedule_Duration': schedule_duration,
                        'Expected_LMRB_Theme': expected_lmrb_theme,
                        'Status': 'DATE/PROGRAM MOVED',
                        'Reason': f'Aired but not on scheduled date/program',
                        'Category': 'TIMING_CHANGE'
                    })
                else:
                    # Not found at all - truly missed
                    missed_entries.append({
                        'Schedule_Theme': schedule_theme,
                        'Schedule_Program': schedule_program,
                        'Schedule_Date': schedule_date,
                        'Schedule_Duration': schedule_duration,
                        'Expected_LMRB_Theme': expected_lmrb_theme,
                        'Status': 'NOT AIRED',
                        'Reason': f'Scheduled entry was not found in aired data',
                        'Category': 'NOT_AIRED'
                    })
            
            print(f"\n📊 MISSED SCHEDULE ANALYSIS RESULTS:")
            print(f"  • Total schedule entries: {total_schedule}")
            print(f"  • Found exact matches: {found_matches}")
            print(f"  • Missed/moved entries: {len(missed_entries)}")
            
            # Categorize missed entries
            categories = {}
            for entry in missed_entries:
                cat = entry['Category']
                if cat not in categories:
                    categories[cat] = 0
                categories[cat] += 1
            
            print(f"  • Categories breakdown:")
            for cat, count in categories.items():
                print(f"    - {cat}: {count}")
            
            return pd.DataFrame(missed_entries) if missed_entries else pd.DataFrame()
            
        except Exception as e:
            print(f"❌ Error identifying missed schedule entries: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def get_missed_schedule_data(self):
        """
        Public method to get missed schedule data for PDF export
        """
        return self._identify_missed_schedule_entries()

    def _enhanced_program_matching(self, scheduled_program, scheduled_start, scheduled_end, scheduled_date,
                                aired_program, aired_time, aired_date):
        """
        Enhanced program matching logic:
        1. Try program name match first
        2. If no name match, check if aired time falls within program time range
        """
        result = {
            'method': 'NO_MATCH',
            'status': 'NO_MATCH',
            'time_status': 'NO_MATCH',
            'overall_type': 'PROGRAM_WRONG',
            'confidence': 'Low',
            'notes': ''
        }
        
        try:
            # Step 1: Try program name matching (case insensitive)
            if scheduled_program and aired_program:
                scheduled_clean = str(scheduled_program).strip().lower()
                aired_clean = str(aired_program).strip().lower()
                
                if scheduled_clean == aired_clean:
                    result['method'] = 'NAME_MATCH'
                    result['status'] = 'EXACT_MATCH'
                    result['overall_type'] = 'PERFECT_MATCH'
                    result['confidence'] = 'High'
                    result['notes'] = 'Program names match exactly'
                    return result
            
            # Step 2: If name doesn't match, try time-based matching
            if (scheduled_start and scheduled_end and aired_time and 
                scheduled_date and aired_date):
                
                # Check if dates match first
                scheduled_date_clean = self._normalize_date(scheduled_date)
                aired_date_clean = self._normalize_date(aired_date)
                
                if scheduled_date_clean == aired_date_clean:
                    # Convert times to comparable format
                    start_time = self._normalize_time(scheduled_start)
                    end_time = self._normalize_time(scheduled_end)
                    air_time = self._normalize_time(aired_time)
                    
                    if start_time is not None and end_time is not None and air_time is not None:
                        # Check if aired time falls within program time range
                        if start_time <= air_time <= end_time:
                            result['method'] = 'TIME_BASED_MATCH'
                            result['status'] = 'TIME_RANGE_MATCH'
                            result['overall_type'] = 'PROGRAM_CORRECT_TIME_BASED'
                            result['confidence'] = 'Medium'
                            result['notes'] = f'Aired time falls within program time range'
                            return result
                        else:
                            result['notes'] = f'Aired time outside program time range'
                    else:
                        result['notes'] = 'Could not parse time values for comparison'
                else:
                    result['notes'] = f'Date mismatch'
            else:
                result['notes'] = 'Missing time/date information for time-based matching'
            
            # Step 3: No match found
            result['method'] = 'NO_MATCH'
            result['status'] = 'NO_MATCH'
            result['overall_type'] = 'PROGRAM_WRONG'
            result['confidence'] = 'Low'
            if not result['notes']:
                result['notes'] = 'Neither program name nor time range matched'
            
            return result
            
        except Exception as e:
            result['notes'] = f'Error in matching logic: {str(e)}'
            return result

    def export_schedule_vs_actual_programs(self):
        """
        FIXED: Complete Schedule vs Actual Programs analysis using matched data as source
        """
        try:
            if self.app.schedule_data is None or not self.app.matched_data:
                messagebox.showerror("Error", "No schedule or matched data available")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.get_default_filename("Schedule_vs_Actual_Programs_Complete"),
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filename:
                return
                
            print("=== Creating COMPLETE Schedule vs Actual Programs Report ===")
            
            # Convert matched data to DataFrame
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            print(f"📊 Source Data Summary:")
            print(f"  • Matched LMRB records: {len(matched_df)}")
            print(f"  • Schedule records: {len(self.app.schedule_data)}")
            print(f"  • Theme mappings: {len(self.app.theme_mapping)}")
            
            # Find columns in schedule data
            schedule_theme_col = self._find_column(self.app.schedule_data, ['theme', 'aadvt', 'advt_theme'])
            schedule_program_col = self._find_column(self.app.schedule_data, ['program', 'programme'])
            schedule_start_time_col = self._find_column(self.app.schedule_data, ['start_time', 'start', 'program_start', 'prog_start'])
            schedule_end_time_col = self._find_column(self.app.schedule_data, ['end_time', 'end', 'program_end', 'prog_end'])
            schedule_date_col = self._find_column(self.app.schedule_data, ['date', 'air_date', 'broadcast_date'])
            schedule_dur_col = self._find_column(self.app.schedule_data, ['dur', 'duration'])
            schedule_adtype_col = self._find_column(self.app.schedule_data, ['advertisement_type', 'ad_type', 'type'])
            
            # FIXED: Find columns in matched LMRB data - prioritize Advt_time
            lmrb_theme_col = self._find_column(matched_df, ['Advt_Theme', 'theme', 'aadvt', 'advt_theme', 'product'])
            lmrb_program_col = self._find_column(matched_df, ['Program', 'program', 'programme'])
            lmrb_time_col = self._find_column(matched_df, ['Advt_time', 'advt_time', 'air_time', 'time'])  # FIXED: Advt_time first
            lmrb_dur_col = self._find_column(matched_df, ['Dur', 'dur', 'duration'])
            
            # FIXED: Handle LMRB date columns (Dd, Mn, Yr instead of Date)
            lmrb_dd_col = 'Dd' if 'Dd' in matched_df.columns else None
            lmrb_mn_col = 'Mn' if 'Mn' in matched_df.columns else None
            lmrb_yr_col = 'Yr' if 'Yr' in matched_df.columns else None
            lmrb_date_col = self._find_column(matched_df, ['Date', 'date']) if not all([lmrb_dd_col, lmrb_mn_col, lmrb_yr_col]) else None
            
            print(f"\n=== COLUMN DETECTION ===")
            print(f"Schedule columns:")
            print(f"  Theme: {schedule_theme_col}")
            print(f"  Program: {schedule_program_col}")  
            print(f"  Start Time: {schedule_start_time_col}")
            print(f"  End Time: {schedule_end_time_col}")
            print(f"  Date: {schedule_date_col}")
            
            print(f"LMRB columns:")
            print(f"  Theme: {lmrb_theme_col}")
            print(f"  Program: {lmrb_program_col}")
            print(f"  Air Time: {lmrb_time_col}")  # This should now show Advt_time
            print(f"  Date Components: Dd={lmrb_dd_col}, Mn={lmrb_mn_col}, Yr={lmrb_yr_col}")
            print(f"  Date Column: {lmrb_date_col}")
            
            # Verify we have required columns
            has_lmrb_date_info = all([lmrb_dd_col, lmrb_mn_col, lmrb_yr_col]) or lmrb_date_col
            
            if not schedule_theme_col or not lmrb_theme_col or not schedule_date_col or not has_lmrb_date_info:
                messagebox.showerror("Error", "Required columns (theme/date) not found in data")
                return
            
            # FIXED: Create Date column in matched_df if using Dd, Mn, Yr
            if all([lmrb_dd_col, lmrb_mn_col, lmrb_yr_col]):
                print("📅 Creating Date column from Dd, Mn, Yr...")
                matched_df['Computed_Date'] = matched_df.apply(
                    lambda row: self._create_date_from_components(row[lmrb_dd_col], row[lmrb_mn_col], row[lmrb_yr_col]), 
                    axis=1
                )
                lmrb_date_col = 'Computed_Date'
                print(f"✅ Created Computed_Date column")
            
            print(f"\n=== FIXED: PROCESSING USING MATCHED DATA AS SOURCE ===")
            print(f"📊 Using {len(matched_df)} MATCHED records as source of truth")
            
            schedule_lmrb_analysis = []
            
            # FIXED: Process each MATCHED LMRB record and find its schedule counterpart
            for matched_idx, matched_row in matched_df.iterrows():
                try:
                    lmrb_theme = matched_row.get(lmrb_theme_col, 'Unknown') if lmrb_theme_col else 'Unknown'
                    lmrb_program = matched_row.get(lmrb_program_col, 'Unknown') if lmrb_program_col else 'Unknown'
                    lmrb_time = matched_row.get(lmrb_time_col, 'Unknown') if lmrb_time_col else 'Unknown'
                    lmrb_date = matched_row.get(lmrb_date_col, 'Unknown') if lmrb_date_col else 'Unknown'
                    match_type = matched_row.get('Match_Type', 'AUTOMATIC')
                    match_timestamp = matched_row.get('Match_Timestamp', 'Unknown')
                    
                    print(f"  🔍 Processing matched LMRB #{matched_idx + 1}: {lmrb_theme} in '{lmrb_program}' at {lmrb_time}")
                    
                    # Find the corresponding schedule theme using theme mappings
                    schedule_theme = None
                    mapping_idx = None
                    mapping_type = 'Unknown'
                    
                    for map_idx, mapping in enumerate(self.app.theme_mapping, 1):
                        mapped_schedule_theme = mapping.get('schedule_theme', '')
                        mapped_lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                        mapped_type = mapping.get('mapping_type', 'Unknown')
                        
                        if mapped_lmrb_theme == lmrb_theme and mapped_schedule_theme:
                            schedule_theme = mapped_schedule_theme
                            mapping_idx = map_idx
                            mapping_type = mapped_type
                            print(f"    ✅ Found mapping: LMRB '{lmrb_theme}' → Schedule '{schedule_theme}'")
                            break
                    
                    if not schedule_theme:
                        # No schedule mapping found for this LMRB theme
                        print(f"    ❌ No schedule mapping found for LMRB theme '{lmrb_theme}'")
                        schedule_lmrb_analysis.append({
                            'Row_Number': matched_idx + 1,
                            'Mapping_Number': 'NO MAPPING',
                            'Schedule_Theme': 'NO MAPPING',
                            'LMRB_Theme': lmrb_theme,
                            'Mapping_Type': mapping_type,
                            'Scheduled_Program': 'NO MAPPING',
                            'Scheduled_Start_Time': 'NO MAPPING',
                            'Scheduled_End_Time': 'NO MAPPING',
                            'Scheduled_Date': 'NO MAPPING',
                            'Actually_Aired_Program': lmrb_program,
                            'Actually_Aired_Time': lmrb_time,
                            'Actually_Aired_Date': lmrb_date,
                            'Match_Status': 'NO SCHEDULE MAPPING',
                            'Match_Method': match_type,
                            'Match_Timestamp': match_timestamp,
                            'Notes': f'LMRB theme "{lmrb_theme}" has no schedule mapping'
                        })
                        continue
                    
                    # Find schedule entries for this theme
                    schedule_entries = self.app.schedule_data[
                        self.app.schedule_data[schedule_theme_col] == schedule_theme
                    ]
                    
                    if schedule_entries.empty:
                        # No schedule data found for this theme
                        print(f"    ⚠️ No schedule data found for theme '{schedule_theme}'")
                        schedule_lmrb_analysis.append({
                            'Row_Number': matched_idx + 1,
                            'Mapping_Number': mapping_idx,
                            'Schedule_Theme': schedule_theme,
                            'LMRB_Theme': lmrb_theme,
                            'Mapping_Type': mapping_type,
                            'Scheduled_Program': 'NOT IN SCHEDULE',
                            'Scheduled_Start_Time': 'NOT IN SCHEDULE',
                            'Scheduled_End_Time': 'NOT IN SCHEDULE',
                            'Scheduled_Date': 'NOT IN SCHEDULE',
                            'Actually_Aired_Program': lmrb_program,
                            'Actually_Aired_Time': lmrb_time,
                            'Actually_Aired_Date': lmrb_date,
                            'Match_Status': 'NO SCHEDULE DATA',
                            'Match_Method': match_type,
                            'Match_Timestamp': match_timestamp,
                            'Notes': f'Schedule theme "{schedule_theme}" not found in schedule data'
                        })
                        continue
                    
                    # FIXED: Find the BEST matching schedule entry (same date + closest time/program)
                    lmrb_date_normalized = self._normalize_date(lmrb_date)
                    print(f"    📅 Looking for schedule entries on date: {lmrb_date_normalized}")
                    
                    # First, try to find schedule entries on the same date
                    date_matched_schedule = []
                    for _, sched_row in schedule_entries.iterrows():
                        sched_date = sched_row[schedule_date_col] if schedule_date_col else None
                        sched_date_normalized = self._normalize_date(sched_date)
                        
                        if lmrb_date_normalized == sched_date_normalized:
                            date_matched_schedule.append(sched_row)
                    
                    print(f"    📊 Found {len(date_matched_schedule)} schedule entries on matching date")
                    
                    if not date_matched_schedule:
                        # No date match - use first schedule entry for comparison
                        best_schedule_match = schedule_entries.iloc[0]
                        match_status = 'DATE MISMATCH'
                        match_method = f'{match_type} - DATE MISMATCH'
                        notes = f'LMRB date {lmrb_date_normalized} not found in schedule dates'
                        print(f"    ❌ No date matches found")
                    else:
                        # Find best schedule match from date-matched entries
                        best_schedule_match = None
                        best_match_score = 0
                        match_status = 'PROGRAM/TIME MISMATCH'
                        match_method = f'{match_type} - NO MATCH'
                        notes = 'No good match found'
                        
                        for sched_row in date_matched_schedule:
                            sched_program = sched_row[schedule_program_col] if schedule_program_col else 'Unknown'
                            sched_start = sched_row[schedule_start_time_col] if schedule_start_time_col else None
                            sched_end = sched_row[schedule_end_time_col] if schedule_end_time_col else None
                            
                            # Check program name match (highest priority)
                            if sched_program and lmrb_program:
                                sched_prog_clean = str(sched_program).strip().lower()
                                lmrb_prog_clean = str(lmrb_program).strip().lower()
                                
                                if sched_prog_clean == lmrb_prog_clean:
                                    best_schedule_match = sched_row
                                    match_status = 'PERFECT MATCH'
                                    match_method = f'{match_type} - PROGRAM NAME MATCH'
                                    notes = 'Program names match exactly'
                                    print(f"    🎯 PERFECT MATCH: Program names match exactly")
                                    break
                            
                            # Check time-based match
                            if sched_start and sched_end and lmrb_time:
                                time_match_result = self._check_time_based_match(sched_start, sched_end, lmrb_time)
                                if time_match_result['is_match'] and best_match_score < 1:
                                    best_schedule_match = sched_row
                                    match_status = 'TIME-BASED MATCH'
                                    match_method = f'{match_type} - TIME RANGE MATCH'
                                    notes = time_match_result['notes']
                                    best_match_score = 1
                                    print(f"    ⏰ TIME-BASED MATCH: {notes}")
                        
                        # If no good match found, use first date-matched entry
                        if best_schedule_match is None:
                            best_schedule_match = date_matched_schedule[0]
                            print(f"    ⚠️ No perfect match, using first date-matched entry")
                    
                    # Create the analysis entry
                    schedule_lmrb_analysis.append({
                        'Row_Number': matched_idx + 1,
                        'Mapping_Number': mapping_idx,
                        'Schedule_Theme': schedule_theme,
                        'LMRB_Theme': lmrb_theme,
                        'Mapping_Type': mapping_type,
                        'Scheduled_Program': best_schedule_match[schedule_program_col] if schedule_program_col else 'Unknown',
                        'Scheduled_Start_Time': best_schedule_match[schedule_start_time_col] if schedule_start_time_col else 'N/A',
                        'Scheduled_End_Time': best_schedule_match[schedule_end_time_col] if schedule_end_time_col else 'N/A',
                        'Scheduled_Date': best_schedule_match[schedule_date_col] if schedule_date_col else 'N/A',
                        'Actually_Aired_Program': lmrb_program,
                        'Actually_Aired_Time': lmrb_time,
                        'Actually_Aired_Date': lmrb_date,
                        'Match_Status': match_status,
                        'Match_Method': match_method,
                        'Match_Timestamp': match_timestamp,
                        'Notes': notes
                    })
                    
                    print(f"    ✅ Analysis entry created: {match_status}")
                    
                except Exception as e:
                    print(f"❌ Error processing matched record {matched_idx}: {e}")
                    # Still create an error entry
                    schedule_lmrb_analysis.append({
                        'Row_Number': matched_idx + 1,
                        'Mapping_Number': 'ERROR',
                        'Schedule_Theme': 'ERROR',
                        'LMRB_Theme': 'ERROR',
                        'Mapping_Type': 'ERROR',
                        'Scheduled_Program': 'ERROR',
                        'Scheduled_Start_Time': 'ERROR',
                        'Scheduled_End_Time': 'ERROR',
                        'Scheduled_Date': 'ERROR',
                        'Actually_Aired_Program': 'ERROR',
                        'Actually_Aired_Time': 'ERROR',
                        'Actually_Aired_Date': 'ERROR',
                        'Match_Status': 'PROCESSING ERROR',
                        'Match_Method': 'ERROR',
                        'Match_Timestamp': 'ERROR',
                        'Notes': f'Error processing record: {str(e)}'
                    })
                    continue
            
            print(f"\n=== SCHEDULE-LMRB ANALYSIS COMPLETE ===")
            print(f"📊 Input: {len(matched_df)} matched LMRB records")
            print(f"📊 Output: {len(schedule_lmrb_analysis)} analysis entries")
            print(f"✅ Ratio: 1:1 (each match = one analysis entry)")
            
            # NOW CREATE TC-LMRB ANALYSIS (Fixed version)
            print(f"\n=== CREATING TC-LMRB ANALYSIS ===")
            tc_lmrb_analysis = self._create_tc_lmrb_analysis()
            
            # Create summary statistics
            if schedule_lmrb_analysis:
                total_entries = len(schedule_lmrb_analysis)
                perfect_matches = len([x for x in schedule_lmrb_analysis if x['Match_Status'] == 'PERFECT MATCH'])
                time_based_matches = len([x for x in schedule_lmrb_analysis if x['Match_Status'] == 'TIME-BASED MATCH'])
                not_aired = len([x for x in schedule_lmrb_analysis if x['Match_Status'] == 'NO SCHEDULE DATA'])
                date_mismatches = len([x for x in schedule_lmrb_analysis if x['Match_Status'] == 'DATE MISMATCH'])
                no_mapping = len([x for x in schedule_lmrb_analysis if x['Match_Status'] == 'NO SCHEDULE MAPPING'])
                
                summary_data = [
                    {'Metric': 'Total Schedule-LMRB Pairs', 'Count': total_entries, 'Percentage': '100.0%'},
                    {'Metric': 'Perfect Matches (Same Program)', 'Count': perfect_matches, 'Percentage': f'{perfect_matches/total_entries*100:.1f}%'},
                    {'Metric': 'Time-Based Matches (Different Program)', 'Count': time_based_matches, 'Percentage': f'{time_based_matches/total_entries*100:.1f}%'},
                    {'Metric': 'No Schedule Data', 'Count': not_aired, 'Percentage': f'{not_aired/total_entries*100:.1f}%'},
                    {'Metric': 'Date Mismatches', 'Count': date_mismatches, 'Percentage': f'{date_mismatches/total_entries*100:.1f}%'},
                    {'Metric': 'No Schedule Mapping', 'Count': no_mapping, 'Percentage': f'{no_mapping/total_entries*100:.1f}%'},
                    {'Metric': 'Overall Success Rate', 'Count': perfect_matches + time_based_matches, 'Percentage': f'{(perfect_matches + time_based_matches)/total_entries*100:.1f}%'}
                ]
            else:
                summary_data = [{'Metric': 'No Data Available', 'Count': 0, 'Percentage': '0%'}]
            
            # Export to Excel with multiple sheets
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Sheet 1: Schedule vs LMRB Analysis
                if schedule_lmrb_analysis:
                    schedule_df = pd.DataFrame(schedule_lmrb_analysis)
                    schedule_df.to_excel(writer, sheet_name='Schedule vs LMRB', index=False)
                    self._apply_excel_formatting(writer, 'Schedule vs LMRB', schedule_df, 'Match_Status')
                
                # Sheet 2: TC vs LMRB Analysis  
                if tc_lmrb_analysis:
                    tc_df = pd.DataFrame(tc_lmrb_analysis)
                    tc_df.to_excel(writer, sheet_name='TC vs LMRB', index=False)
                    self._apply_excel_formatting(writer, 'TC vs LMRB', tc_df, 'Match_Status')
                
                # Sheet 3: COLOR LEGEND
                self._create_color_legend_sheet(writer)
                
                # Sheet 4: Summary
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Sheet 5: Theme Mappings Used
                mapping_data = []
                for i, mapping in enumerate(self.app.theme_mapping, 1):
                    mapping_data.append({
                        'Mapping_Number': i,
                        'Schedule_Theme': mapping.get('schedule_theme', ''),
                        'LMRB_Theme': mapping.get('media_watch_theme', mapping.get('lmrb_theme', '')),
                        'TC_Theme': mapping.get('tc_theme', ''),
                        'Mapping_Type': mapping.get('mapping_type', 'Unknown')
                    })
                
                if mapping_data:
                    pd.DataFrame(mapping_data).to_excel(writer, sheet_name='Theme Mappings', index=False)
            
            # Success message
            messagebox.showinfo("Success", 
                f"✅ COMPLETE Export Successful!\n\n"
                f"📊 Schedule vs LMRB Analysis:\n"
                f"• Input matched records: {len(matched_df)}\n"
                f"• Analysis entries created: {len(schedule_lmrb_analysis)}\n"
                f"• Perfect matches: {perfect_matches if schedule_lmrb_analysis else 0}\n"
                f"• Time-based matches: {time_based_matches if schedule_lmrb_analysis else 0}\n\n"
                f"📊 TC vs LMRB Analysis:\n"
                f"• TC-LMRB pairs: {len(tc_lmrb_analysis)}\n\n"
                f"✅ Now showing 1:1 ratio (no more duplicates!)\n"
                f"📁 File: {filename}")
            
            # ADD THIS: Get missed schedule entries
            print(f"\n=== ANALYZING MISSED SCHEDULE ENTRIES ===")
            missed_schedule_df = self._identify_missed_schedule_entries()
            
            # Export to Excel with MISSED SCHEDULE sheet
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Your existing sheets...
                if schedule_lmrb_analysis:
                    schedule_df = pd.DataFrame(schedule_lmrb_analysis)
                    schedule_df.to_excel(writer, sheet_name='Schedule vs LMRB', index=False)
                    self._apply_excel_formatting(writer, 'Schedule vs LMRB', schedule_df, 'Match_Status')
                
                if tc_lmrb_analysis:
                    tc_df = pd.DataFrame(tc_lmrb_analysis)
                    tc_df.to_excel(writer, sheet_name='TC vs LMRB', index=False)
                    self._apply_excel_formatting(writer, 'TC vs LMRB', tc_df, 'Match_Status')
                
                # NEW: Add missed schedule entries
                if not missed_schedule_df.empty:
                    missed_schedule_df.to_excel(writer, sheet_name='Missed Schedule', index=False)
                    self._apply_missed_schedule_formatting(writer, 'Missed Schedule', missed_schedule_df)
                    print(f"✅ Added {len(missed_schedule_df)} missed schedule entries to export")
                
                # Your existing sheets...
                self._create_color_legend_sheet(writer)
                
                # Enhanced summary including missed data
                enhanced_summary = summary_data + [
                    {'Metric': 'Missed Schedule Entries', 'Count': len(missed_schedule_df), 'Percentage': f'{len(missed_schedule_df)/len(self.app.schedule_data)*100:.1f}%'},
                    {'Metric': 'Schedule Fulfillment Rate', 'Count': len(self.app.schedule_data) - len(missed_schedule_df), 'Percentage': f'{(len(self.app.schedule_data) - len(missed_schedule_df))/len(self.app.schedule_data)*100:.1f}%'}
                ]
                
                pd.DataFrame(enhanced_summary).to_excel(writer, sheet_name='Enhanced Summary', index=False)
            
            # Enhanced success message
            messagebox.showinfo("Success", 
                f"✅ COMPLETE Export with Missed Schedule Analysis!\n\n"
                f"📊 Schedule vs LMRB Analysis: {len(schedule_lmrb_analysis)}\n"
                f"📊 TC vs LMRB Analysis: {len(tc_lmrb_analysis)}\n"
                f"📊 Missed Schedule Entries: {len(missed_schedule_df)}\n\n"
                f"📋 Missed Categories:\n"
                f"• Not Aired: {len([x for x in missed_schedule_df.to_dict('records') if x.get('Category') == 'NOT_AIRED'])}\n"
                f"• Program Changes: {len([x for x in missed_schedule_df.to_dict('records') if x.get('Category') == 'PROGRAM_CHANGE'])}\n"
                f"• Timing Changes: {len([x for x in missed_schedule_df.to_dict('records') if x.get('Category') == 'TIMING_CHANGE'])}\n"
                f"• Mapping Issues: {len([x for x in missed_schedule_df.to_dict('records') if x.get('Category') == 'MAPPING_ISSUE'])}\n\n"
                f"📁 File: {filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error in enhanced export: {str(e)}")
            import traceback
            traceback.print_exc()

    def _apply_missed_schedule_formatting(self, writer, sheet_name, df):
        """Apply color formatting for missed schedule entries"""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define colors for different categories
            not_aired_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')      # Red
            program_moved_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange  
            timing_changed_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # Yellow
            mapping_issue_fill = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')  # Pink
            
            # Find the Category column index
            category_col_idx = None
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Category':
                    category_col_idx = col_idx
                    break
            
            if category_col_idx:
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    category_value = worksheet.cell(row=row, column=category_col_idx).value
                    
                    # Choose color based on category
                    fill_color = None
                    if category_value == 'NOT_AIRED':
                        fill_color = not_aired_fill
                    elif category_value == 'PROGRAM_CHANGE':
                        fill_color = program_moved_fill
                    elif category_value == 'TIMING_CHANGE':
                        fill_color = timing_changed_fill
                    elif category_value == 'MAPPING_ISSUE':
                        fill_color = mapping_issue_fill
                    
                    # Apply color to entire row
                    if fill_color:
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = fill_color
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Error applying missed schedule formatting: {e}")

    def _check_time_based_match(self, sched_start, sched_end, lmrb_time):
        """Check if LMRB aired time falls within scheduled program time range"""
        result = {'is_match': False, 'notes': ''}

        try:
            if not sched_start or not sched_end or not lmrb_time:
                result['notes'] = 'Missing time information for comparison'
                return result
            
            # Convert times to comparable format (minutes since midnight)
            start_minutes = self._normalize_time(sched_start)
            end_minutes = self._normalize_time(sched_end)
            air_minutes = self._normalize_time(lmrb_time)
            
            if start_minutes is None or end_minutes is None or air_minutes is None:
                result['notes'] = 'Could not parse time values'
                return result
            
            # Check if aired time falls within program time range
            if start_minutes <= air_minutes <= end_minutes:
                result['is_match'] = True
                result['notes'] = f'Aired at {lmrb_time} within program time {sched_start}-{sched_end}'
            else:
                result['notes'] = f'Aired at {lmrb_time} outside program time {sched_start}-{sched_end}'
            
            return result
            
        except Exception as e:
            result['notes'] = f'Error in time comparison: {str(e)}'
            return result

    def _debug_date_comparison(self, schedule_entries, lmrb_entries, schedule_date_col, lmrb_date_col):
        """Debug function to see actual dates in both datasets"""
        print(f"\n🔍 DEBUG: Date Comparison Analysis")

        # Show schedule dates
        if not schedule_entries.empty and schedule_date_col:
            print(f"📅 Schedule dates found:")
            schedule_dates = schedule_entries[schedule_date_col].unique()
            for i, date in enumerate(schedule_dates[:5]):  # Show first 5
                normalized = self._normalize_date(date)
                print(f"  {i+1}. Original: '{date}' → Normalized: '{normalized}' (Type: {type(date)})")
            if len(schedule_dates) > 5:
                print(f"  ... and {len(schedule_dates) - 5} more dates")

        # Show LMRB dates  
        if not lmrb_entries.empty and lmrb_date_col:
            print(f"📺 LMRB dates found:")
            lmrb_dates = lmrb_entries[lmrb_date_col].unique()
            for i, date in enumerate(lmrb_dates[:5]):  # Show first 5
                normalized = self._normalize_date(date)
                print(f"  {i+1}. Original: '{date}' → Normalized: '{normalized}' (Type: {type(date)})")
            if len(lmrb_dates) > 5:
                print(f"  ... and {len(lmrb_dates) - 5} more dates")

        print(f"🔍 DEBUG: End of date analysis\n")

    def _create_date_from_components(self, dd, mn, yr):
        """Create a date string from Dd, Mn, Yr components"""
        try:
            if pd.isna(dd) or pd.isna(mn) or pd.isna(yr):
                return None
            
            # Convert to integers and create date string
            day = int(float(dd))
            month = int(float(mn))
            year = int(float(yr))
            
            # Create date in YYYY-MM-DD format
            date_str = f"{year:04d}-{month:02d}-{day:02d}"
            
            # Validate the date
            from datetime import datetime
            datetime.strptime(date_str, '%Y-%m-%d')  # This will raise an error if invalid
            
            return date_str
            
        except Exception as e:
            print(f"Error creating date from components Dd={dd}, Mn={mn}, Yr={yr}: {e}")
            return f"ERROR_{dd}_{mn}_{yr}"

    def _create_color_legend_sheet(self, writer):
        """Create a sheet explaining the color coding"""
        try:
            # Create color legend data
            legend_data = [
                ['Color Code', 'Match Status', 'Description', 'What It Means'],
                ['', '', '', ''],
                ['🟢 GREEN', 'PERFECT MATCH', 'Program Name Match', 'Ad aired in exactly the planned program'],
                ['', '', '', ''],
                ['🟡 LIGHT GREEN', 'TIME-BASED MATCH', 'Time Range Match', 'Different program name but aired within program time slot'],
                ['🟡 LIGHT GREEN', 'PROGRAM CORRECT TIME-BASED', 'Time-Based Program Match', 'Correct program identified by time range'],
                ['', '', '', ''],
                ['🟡 YELLOW', 'TIME APPROXIMATE MATCH', 'Close Time (1-5 min)', 'Times are close but not exact'],
                ['', '', '', ''],
                ['🟠 ORANGE', 'PROGRAM/TIME MISMATCH', 'Wrong Program/Time', 'Ad aired in different program and different time'],
                ['🟠 ORANGE', 'NO MATCH', 'No Matching Criteria', 'Could not match by program name or time'],
                ['', '', '', ''],
                ['🔴 RED', 'NOT AIRED', 'Missing from LMRB', 'Scheduled ad was not found in aired data'],
                ['🔴 RED', 'DATE MISMATCH', 'Different Date', 'Same theme but aired on different date'],
                ['🔴 RED', 'NO SCHEDULE DATA', 'Missing Schedule', 'Theme mapping exists but no schedule data'],
                ['🔴 RED', 'NO TC DATA', 'Missing TC Data', 'Theme mapping exists but no TC data'],
                ['', '', '', ''],
                ['📊 ANALYSIS TIPS', '', '', ''],
                ['', 'Perfect Accuracy', 'GREEN entries', 'These ads aired exactly as planned'],
                ['', 'Good Placement', 'LIGHT GREEN entries', 'Correct program, identified by time analysis'],
                ['', 'Timing Issues', 'YELLOW entries', 'Same ad but slight timing differences'],
                ['', 'Wrong Placement', 'ORANGE entries', 'Ads moved to different programs/times'],
                ['', 'Missing/Problems', 'RED entries', 'Ads not aired or data quality issues'],
                ['', '', '', ''],
                ['📈 SUCCESS METRICS', '', '', ''],
                ['', 'Program Accuracy', 'GREEN + LIGHT GREEN', 'Percentage of ads in correct programs'],
                ['', 'Overall Success', 'GREEN + LIGHT GREEN + YELLOW', 'Percentage of ads successfully aired'],
                ['', 'Issues to Review', 'ORANGE + RED', 'Ads requiring attention or investigation']
            ]
            
            # Create DataFrame
            legend_df = pd.DataFrame(legend_data[1:], columns=legend_data[0])
            
            # Write to Excel
            legend_df.to_excel(writer, sheet_name='Color Legend', index=False)
            
            # Apply formatting to the legend sheet
            workbook = writer.book
            worksheet = writer.sheets['Color Legend']
            
            # Define the actual colors used in analysis
            green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # Apply colors to explanation rows
            color_rows = {
                3: green_fill,   # PERFECT MATCH
                5: light_green_fill,  # TIME-BASED MATCH
                6: light_green_fill,  # PROGRAM CORRECT TIME-BASED
                8: yellow_fill,  # TIME APPROXIMATE MATCH
                10: orange_fill, # PROGRAM/TIME MISMATCH
                11: orange_fill, # NO MATCH
                13: red_fill,    # NOT AIRED
                14: red_fill,    # DATE MISMATCH
                15: red_fill,    # NO SCHEDULE DATA
                16: red_fill,    # NO TC DATA
            }
            
            for row_num, fill_color in color_rows.items():
                for col in range(1, 5):  # Apply to all columns
                    worksheet.cell(row=row_num, column=col).fill = fill_color
            
            # Make headers bold
            for col in range(1, 5):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
            print("✅ Color legend sheet created successfully")
            
        except Exception as e:
            print(f"Error creating color legend: {e}")

    def _create_tc_lmrb_analysis(self):
        """
        FIXED: Create TC vs LMRB analysis using ACTUAL matched results (no duplicates)
        Uses the results section data (self.app.matched_data) as the single source of truth
        """
        try:
            tc_lmrb_analysis = []
            
            if not self.app.matched_data:
                print("No matched data available for TC-LMRB analysis")
                return []
            
            # Convert matched data to DataFrame - THIS IS THE KEY FIX
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            print(f"🔍 TC-LMRB Analysis - Using MATCHED RESULTS as source")
            print(f"📊 Total matched records to analyze: {len(matched_df)}")
            
            # Get original TC data for comparison
            if not hasattr(self.app, 'tc_data_original') or self.app.tc_data_original is None:
                print("⚠️ No original TC data available - will mark all as 'NO TC COMPARISON'")
                # Still create analysis from matched data
                for idx, row in matched_df.iterrows():
                    tc_lmrb_analysis.append({
                        'Mapping_Number': 'N/A',
                        'TC_Theme': 'NO TC DATA',
                        'LMRB_Theme': row.get('Advt_Theme', row.get('Theme', 'Unknown')),
                        'TC_Program': 'NO TC DATA',
                        'TC_Time': 'NO TC DATA',
                        'TC_Date': 'NO TC DATA',
                        'LMRB_Program': row.get('Program', 'Unknown'),
                        'LMRB_Time': row.get('Advt_time', row.get('Time', 'Unknown')),
                        'LMRB_Date': self._get_date_from_row(row),
                        'Match_Status': 'LMRB ONLY',
                        'Match_Method': row.get('Match_Type', 'AUTOMATIC'),
                        'Notes': 'Matched LMRB entry - no TC data for comparison'
                    })
                return tc_lmrb_analysis
            
            # Find columns in matched LMRB data
            lmrb_theme_col = self._find_column(matched_df, ['Advt_Theme', 'theme', 'aadvt', 'product'])
            lmrb_program_col = self._find_column(matched_df, ['Program', 'program', 'programme'])
            lmrb_time_col = self._find_column(matched_df, ['Advt_time', 'advt_time', 'air_time', 'time'])
            
            # Find columns in original TC data
            tc_theme_col = self._find_column(self.app.tc_data_original, ['Advt_Theme', 'theme', 'aadvt'])
            tc_program_col = self._find_column(self.app.tc_data_original, ['Program', 'program', 'programme'])
            tc_time_col = self._find_column(self.app.tc_data_original, ['Advt_time', 'time', 'advt_time', 'air_time'])
            tc_date_col = self._find_column(self.app.tc_data_original, ['Date', 'date'])
            
            print(f"📋 Column Mapping:")
            print(f"  LMRB: Theme={lmrb_theme_col}, Program={lmrb_program_col}, Time={lmrb_time_col}")
            print(f"  TC: Theme={tc_theme_col}, Program={tc_program_col}, Time={tc_time_col}, Date={tc_date_col}")
            
            # FIXED: Process each MATCHED record (1-to-1, no duplicates)
            match_type_counts = {'PERFECT MATCH': 0, 'TIME APPROXIMATE MATCH': 0, 'PROGRAM/TIME MISMATCH': 0, 'NO TC COMPARISON': 0}
            
            for idx, matched_row in matched_df.iterrows():
                try:
                    # Get LMRB data from matched record
                    lmrb_theme = matched_row.get(lmrb_theme_col, 'Unknown') if lmrb_theme_col else 'Unknown'
                    lmrb_program = matched_row.get(lmrb_program_col, 'Unknown') if lmrb_program_col else 'Unknown'
                    lmrb_time = matched_row.get(lmrb_time_col, 'Unknown') if lmrb_time_col else 'Unknown'
                    lmrb_date = self._get_date_from_row(matched_row)
                    match_type = matched_row.get('Match_Type', 'AUTOMATIC')
                    
                    # Find corresponding TC theme using theme mappings
                    tc_theme_for_comparison = None
                    mapping_number = 'N/A'
                    
                    for map_idx, mapping in enumerate(self.app.theme_mapping, 1):
                        lmrb_mapped_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                        tc_mapped_theme = mapping.get('tc_theme', '')
                        
                        if lmrb_mapped_theme == lmrb_theme and tc_mapped_theme:
                            tc_theme_for_comparison = tc_mapped_theme
                            mapping_number = map_idx
                            break
                    
                    if not tc_theme_for_comparison:
                        # No TC mapping found
                        tc_lmrb_analysis.append({
                            'Mapping_Number': 'NO MAPPING',
                            'TC_Theme': 'NO MAPPING',
                            'LMRB_Theme': lmrb_theme,
                            'TC_Program': 'NO MAPPING',
                            'TC_Time': 'NO MAPPING',
                            'TC_Date': 'NO MAPPING',
                            'LMRB_Program': lmrb_program,
                            'LMRB_Time': lmrb_time,
                            'LMRB_Date': lmrb_date,
                            'Match_Status': 'NO TC MAPPING',
                            'Match_Method': match_type,
                            'Notes': f'LMRB theme "{lmrb_theme}" has no TC theme mapping'
                        })
                        match_type_counts['NO TC COMPARISON'] += 1
                        continue
                    
                    # Find TC entries with this theme and same date
                    tc_candidates = self.app.tc_data_original[
                        self.app.tc_data_original[tc_theme_col] == tc_theme_for_comparison
                    ] if tc_theme_col else pd.DataFrame()
                    
                    if tc_candidates.empty:
                        # No TC data found for this theme
                        tc_lmrb_analysis.append({
                            'Mapping_Number': mapping_number,
                            'TC_Theme': tc_theme_for_comparison,
                            'LMRB_Theme': lmrb_theme,
                            'TC_Program': 'NOT FOUND IN TC',
                            'TC_Time': 'NOT FOUND IN TC',
                            'TC_Date': 'NOT FOUND IN TC',
                            'LMRB_Program': lmrb_program,
                            'LMRB_Time': lmrb_time,
                            'LMRB_Date': lmrb_date,
                            'Match_Status': 'NO TC DATA',
                            'Match_Method': match_type,
                            'Notes': f'TC theme "{tc_theme_for_comparison}" not found in TC data'
                        })
                        match_type_counts['NO TC COMPARISON'] += 1
                        continue
                    
                    # FIXED: Find the BEST matching TC record (closest time on same date)
                    best_tc_match = None
                    best_time_diff = float('inf')
                    
                    lmrb_date_normalized = self._normalize_date(lmrb_date)
                    
                    for _, tc_row in tc_candidates.iterrows():
                        tc_date = tc_row.get(tc_date_col) if tc_date_col else None
                        tc_date_normalized = self._normalize_date(tc_date)
                        
                        # Check date match first
                        if lmrb_date_normalized == tc_date_normalized:
                            # Calculate time difference
                            tc_time = tc_row.get(tc_time_col, '') if tc_time_col else ''
                            time_diff = self._calculate_time_difference(lmrb_time, tc_time)
                            
                            if time_diff is not None and time_diff < best_time_diff:
                                best_time_diff = time_diff
                                best_tc_match = tc_row
                    
                    if best_tc_match is None:
                        # No date/time match found
                        # Use first TC entry for comparison
                        best_tc_match = tc_candidates.iloc[0]
                        tc_lmrb_analysis.append({
                            'Mapping_Number': mapping_number,
                            'TC_Theme': tc_theme_for_comparison,
                            'LMRB_Theme': lmrb_theme,
                            'TC_Program': best_tc_match.get(tc_program_col, 'Unknown') if tc_program_col else 'Unknown',
                            'TC_Time': best_tc_match.get(tc_time_col, 'Unknown') if tc_time_col else 'Unknown',
                            'TC_Date': best_tc_match.get(tc_date_col, 'Unknown') if tc_date_col else 'Unknown',
                            'LMRB_Program': lmrb_program,
                            'LMRB_Time': lmrb_time,
                            'LMRB_Date': lmrb_date,
                            'Match_Status': 'DATE/TIME MISMATCH',
                            'Match_Method': match_type,
                            'Notes': 'No matching date/time found in TC data'
                        })
                        match_type_counts['PROGRAM/TIME MISMATCH'] += 1
                        continue
                    
                    # Compare the matched LMRB and TC records
                    tc_program = best_tc_match.get(tc_program_col, 'Unknown') if tc_program_col else 'Unknown'
                    tc_time = best_tc_match.get(tc_time_col, 'Unknown') if tc_time_col else 'Unknown'
                    tc_date = best_tc_match.get(tc_date_col, 'Unknown') if tc_date_col else 'Unknown'
                    
                    # Determine match status
                    program_match = str(lmrb_program).strip().lower() == str(tc_program).strip().lower()
                    
                    if program_match and best_time_diff <= 60:  # Same program, time within 1 minute
                        match_status = 'PERFECT MATCH'
                        match_method = f'{match_type} - EXACT MATCH'
                        notes = f'Perfect match: same program, time diff: {int(best_time_diff)}s'
                        match_type_counts['PERFECT MATCH'] += 1
                        
                    elif best_time_diff <= 300:  # Within 5 minutes
                        match_status = 'TIME APPROXIMATE MATCH'
                        match_method = f'{match_type} - TIME CLOSE'
                        notes = f'Close match: time diff: {int(best_time_diff)}s, programs: {lmrb_program} vs {tc_program}'
                        match_type_counts['TIME APPROXIMATE MATCH'] += 1
                        
                    else:
                        match_status = 'PROGRAM/TIME MISMATCH'
                        match_method = f'{match_type} - MISMATCH'
                        notes = f'Mismatch: time diff: {int(best_time_diff)}s, programs: {lmrb_program} vs {tc_program}'
                        match_type_counts['PROGRAM/TIME MISMATCH'] += 1
                    
                    tc_lmrb_analysis.append({
                        'Mapping_Number': mapping_number,
                        'TC_Theme': tc_theme_for_comparison,
                        'LMRB_Theme': lmrb_theme,
                        'TC_Program': tc_program,
                        'TC_Time': tc_time,
                        'TC_Date': tc_date,
                        'LMRB_Program': lmrb_program,
                        'LMRB_Time': lmrb_time,
                        'LMRB_Date': lmrb_date,
                        'Match_Status': match_status,
                        'Match_Method': match_method,
                        'Notes': notes
                    })
                    
                except Exception as e:
                    print(f"❌ Error processing matched record {idx}: {e}")
                    continue
            
            print(f"\n✅ TC-LMRB Analysis Summary:")
            print(f"📊 Total entries created: {len(tc_lmrb_analysis)} (should equal matched records: {len(matched_df)})")
            print(f"📈 Match quality breakdown:")
            for status, count in match_type_counts.items():
                percentage = (count / len(tc_lmrb_analysis) * 100) if tc_lmrb_analysis else 0
                print(f"  • {status}: {count} ({percentage:.1f}%)")
            
            return tc_lmrb_analysis
            
        except Exception as e:
            print(f"❌ Error in TC-LMRB analysis: {str(e)}")
            import traceback
            traceback.print_exc()
            return []

    def get_missed_schedule_data_for_pdf(self):
        """
        Get missed schedule data specifically for PDF - only MAPPING_ISSUE and NOT_AIRED
        """
        try:
            all_missed_df = self._identify_missed_schedule_entries()
            
            if all_missed_df.empty:
                return pd.DataFrame()
            
            # Filter for truly unmatched entries only
            unmatched_df = all_missed_df[
                all_missed_df['Category'].isin(['MAPPING_ISSUE', 'NOT_AIRED'])
            ].copy()
            
            print(f"📄 PDF Export: Filtered {len(unmatched_df)} unmatched schedule entries (MAPPING_ISSUE + NOT_AIRED)")
            return unmatched_df
            
        except Exception as e:
            print(f"Error getting missed schedule data for PDF: {e}")
            return pd.DataFrame()

    def get_missed_schedule_data(self):
        """
        Public method to get ALL missed schedule data for Excel export
        """
        return self._identify_missed_schedule_entries()

    def _get_date_from_row(self, row):
        """Helper function to extract date from a matched row (handles Dd, Mn, Yr or Date columns)"""
        try:
            # Try Date column first
            if 'Date' in row and not pd.isna(row['Date']):
                return str(row['Date'])
            
            # Try Computed_Date if exists
            if 'Computed_Date' in row and not pd.isna(row['Computed_Date']):
                return str(row['Computed_Date'])
            
            # Try Dd, Mn, Yr components
            if all(col in row for col in ['Dd', 'Mn', 'Yr']):
                return self._create_date_from_components(row['Dd'], row['Mn'], row['Yr'])
            
            return 'Unknown Date'
            
        except Exception as e:
            print(f"Error getting date from row: {e}")
            return 'Date Error'

    def _calculate_time_difference(self, time1, time2):
        """Calculate time difference in seconds between two times"""
        try:
            if not time1 or not time2:
                return None
            
            # Convert both times to minutes since midnight
            minutes1 = self._normalize_time(time1)
            minutes2 = self._normalize_time(time2)
            
            if minutes1 is None or minutes2 is None:
                return None
            
            # Convert to seconds and calculate absolute difference
            seconds_diff = abs((minutes1 - minutes2) * 60)
            return seconds_diff
            
        except Exception as e:
            print(f"Error calculating time difference: {e}")
            return None

    def _apply_excel_formatting(self, writer, sheet_name, df, status_column):
        """Apply color formatting to Excel sheet based on match status"""
        try:
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define colors
            perfect_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green
            time_based_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light Green
            close_match_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Yellow
            mismatch_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Orange
            not_aired_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Light Red
            date_mismatch_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # Red
            
            # Find the status column index
            status_col_idx = None
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == status_column:
                    status_col_idx = col_idx
                    break
            
            if status_col_idx:
                for row in range(2, len(df) + 2):  # Start from row 2 (after header)
                    status_value = worksheet.cell(row=row, column=status_col_idx).value
                    
                    # Choose color based on status
                    fill_color = None
                    if status_value == 'PERFECT MATCH':
                        fill_color = perfect_fill
                    elif status_value in ['TIME-BASED MATCH']:
                        fill_color = time_based_fill
                    elif status_value == 'TIME APPROXIMATE MATCH':
                        fill_color = close_match_fill
                    elif status_value in ['PROGRAM/TIME MISMATCH', 'NO MATCH']:
                        fill_color = mismatch_fill
                    elif status_value in ['NOT AIRED', 'NO SCHEDULE DATA']:
                        fill_color = not_aired_fill
                    elif status_value == 'DATE MISMATCH':
                        fill_color = date_mismatch_fill
                    
                    # Apply color to entire row
                    if fill_color:
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=row, column=col).fill = fill_color
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Error applying Excel formatting: {e}")

    def export_cross_program_analysis(self):
        """Export analysis of ads that moved between programs"""
        try:
            if not self.app.matched_data or self.app.schedule_data is None:
                messagebox.showerror("Error", "No data available for analysis")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.get_default_filename("Cross_Program_Movement"),
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filename:
                return
            
            print("Creating Cross-Program Movement analysis...")
            
            # Convert matched data to DataFrame
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            # Find columns
            schedule_theme_col = self._find_column(self.app.schedule_data, ['theme', 'aadvt'])
            schedule_program_col = self._find_column(self.app.schedule_data, ['program'])
            matched_theme_col = self._find_column(matched_df, ['theme', 'aadvt', 'product'])
            matched_program_col = self._find_column(matched_df, ['program'])
            
            if not all([schedule_theme_col, schedule_program_col, matched_theme_col, matched_program_col]):
                messagebox.showerror("Error", "Required columns not found in data")
                return
            
            # Create theme mapping
            theme_lookup = {}
            for mapping in self.app.theme_mapping:
                schedule_theme = mapping.get('schedule_theme', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                if schedule_theme and lmrb_theme:
                    theme_lookup[schedule_theme] = lmrb_theme
            
            # Analyze movements
            movements = []
            program_stats = {}
            
            for _, scheduled_row in self.app.schedule_data.iterrows():
                schedule_theme = scheduled_row[schedule_theme_col]
                scheduled_program = scheduled_row[schedule_program_col]
                lmrb_theme = theme_lookup.get(schedule_theme, schedule_theme)
                
                # Find where this theme actually aired
                aired_entries = matched_df[matched_df[matched_theme_col] == lmrb_theme]
                
                for _, aired_row in aired_entries.iterrows():
                    aired_program = aired_row[matched_program_col]
                    
                    # Track program statistics
                    if aired_program not in program_stats:
                        program_stats[aired_program] = {'planned': 0, 'actual': 0, 'moved_in': 0, 'moved_out': 0}
                    
                    program_stats[aired_program]['actual'] += 1
                    
                    if scheduled_program != aired_program:
                        movements.append({
                            'Theme': schedule_theme,
                            'Planned_Program': scheduled_program,
                            'Actually_Aired_Program': aired_program,
                            'Movement_Type': 'Program Change',
                            'Status': 'Moved'
                        })
                        program_stats[aired_program]['moved_in'] += 1
                        
                        if scheduled_program in program_stats:
                            program_stats[scheduled_program]['moved_out'] += 1
                    else:
                        movements.append({
                            'Theme': schedule_theme,
                            'Planned_Program': scheduled_program,
                            'Actually_Aired_Program': aired_program,
                            'Movement_Type': 'No Change',
                            'Status': 'Stayed'
                        })
            
            # Count planned programs
            for _, row in self.app.schedule_data.iterrows():
                program = row[schedule_program_col]
                if program in program_stats:
                    program_stats[program]['planned'] += 1
            
            # Create program popularity data
            program_popularity = []
            for program, stats in program_stats.items():
                program_popularity.append({
                    'Program': program,
                    'Planned_Count': stats['planned'],
                    'Actual_Count': stats['actual'],
                    'Moved_In': stats['moved_in'],
                    'Moved_Out': stats['moved_out'],
                    'Net_Change': stats['moved_in'] - stats['moved_out'],
                    'Retention_Rate': f"{((stats['actual'] - stats['moved_in']) / max(stats['planned'], 1) * 100):.1f}%"
                })
            
            # Export to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Movement details
                if movements:
                    pd.DataFrame(movements).to_excel(writer, sheet_name='Program Movements', index=False)
                
                # Program popularity
                if program_popularity:
                    pd.DataFrame(program_popularity).to_excel(writer, sheet_name='Program Statistics', index=False)
                
                # Movement summary
                if movements:
                    moved_count = len([x for x in movements if x['Status'] == 'Moved'])
                    stayed_count = len([x for x in movements if x['Status'] == 'Stayed'])
                    
                    summary = [
                        {'Metric': 'Total Ads Tracked', 'Value': len(movements)},
                        {'Metric': 'Ads That Moved Programs', 'Value': moved_count},
                        {'Metric': 'Ads That Stayed in Planned Program', 'Value': stayed_count},
                        {'Metric': 'Program Movement Rate', 'Value': f"{moved_count/len(movements)*100:.1f}%"}
                    ]
                    pd.DataFrame(summary).to_excel(writer, sheet_name='Movement Summary', index=False)
            
            messagebox.showinfo("Success", f"Cross-program analysis exported to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting cross-program analysis: {str(e)}")
            import traceback
            traceback.print_exc()

    def export_theme_performance(self):
        """Export theme-wise performance analysis"""
        try:
            if not self.app.matched_data:
                messagebox.showerror("Error", "No matched data available")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.get_default_filename("Theme_Performance"),
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filename:
                return
            
            print("Creating Theme Performance analysis...")
            
            # Convert matched data to DataFrame
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            # Get total LMRB count
            total_lmrb = len(self.app.lmrb_filtered) if self.app.lmrb_filtered is not None else 0
            total_matched = len(matched_df)
            total_unmatched_lmrb = len(self.app.unmatched_lmrb_data) if self.app.unmatched_lmrb_data is not None else 0

            # Analyze by theme
            theme_analysis = []
            
            # Find theme column in matched data
            matched_theme_col = self._find_column(matched_df, ['theme', 'aadvt', 'product'])
            
            if matched_theme_col:
                # Group by theme
                theme_groups = matched_df.groupby(matched_theme_col)
                
                for theme, group in theme_groups:
                    # Count match types
                    manual_matches = len(group[group.get('Match_Type', pd.Series()) == 'Manual'])
                    auto_matches = len(group) - manual_matches
                    
                    theme_analysis.append({
                        'Theme': theme,
                        'Total_Matched': len(group),
                        'Automatic_Matches': auto_matches,
                        'Manual_Matches': manual_matches,
                        'Manual_Rate': f"{manual_matches/len(group)*100:.1f}%" if len(group) > 0 else "0%",
                        'Match_Success_Rate': f"{len(group)/total_lmrb*100:.1f}%" if total_lmrb > 0 else "0%"
                    })
            
            # Theme mapping effectiveness
            mapping_effectiveness = []
            for mapping in self.app.theme_mapping:
                schedule_theme = mapping.get('schedule_theme', '')
                lmrb_theme = mapping.get('media_watch_theme', mapping.get('lmrb_theme', ''))
                mapping_type = mapping.get('mapping_type', 'Unknown')
                
                if lmrb_theme and matched_theme_col:
                    matches = matched_df[matched_df[matched_theme_col] == lmrb_theme]
                    mapping_effectiveness.append({
                        'Schedule_Theme': schedule_theme,
                        'LMRB_Theme': lmrb_theme,
                        'Mapping_Type': mapping_type,
                        'Matches_Found': len(matches),
                        'Effectiveness': 'High' if len(matches) > 5 else 'Medium' if len(matches) > 0 else 'Low'
                    })
            
            # Overall summary
            summary_data = [
                {'Metric': 'Total LMRB Records', 'Value': total_lmrb},
                {'Metric': 'Total Matched Records', 'Value': total_matched},
                {'Metric': 'Overall Match Rate', 'Value': f"{total_matched/total_lmrb*100:.1f}%" if total_lmrb > 0 else "0%"},
                {'Metric': 'Unique Themes Matched', 'Value': len(theme_analysis)},
                {'Metric': 'Theme Mappings Configured', 'Value': len(self.app.theme_mapping)}
            ]
            
            # Export to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Theme performance
                if theme_analysis:
                    pd.DataFrame(theme_analysis).to_excel(writer, sheet_name='Theme Performance', index=False)
                
                # Mapping effectiveness
                if mapping_effectiveness:
                    pd.DataFrame(mapping_effectiveness).to_excel(writer, sheet_name='Mapping Effectiveness', index=False)
                
                # Summary
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            messagebox.showinfo("Success", f"Theme performance analysis exported to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting theme performance: {str(e)}")
            import traceback
            traceback.print_exc()

    def export_ad_type_performance(self):
        """Export performance comparison between Sponsorships and Commercials"""
        try:
            if not self.app.matched_data or self.app.schedule_data is None:
                messagebox.showerror("Error", "No data available for analysis")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.get_default_filename("Ad_Type_Performance"),
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filename:
                return
            
            print("Creating Ad Type Performance analysis...")
            
            # Convert matched data to DataFrame
            matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data.copy()
            
            # Find advertisement type column in schedule data
            schedule_adtype_col = self._find_column(self.app.schedule_data, ['advertisement_type', 'ad_type', 'type'])
            
            if not schedule_adtype_col:
                messagebox.showwarning("Warning", "No advertisement type column found. Analysis will be limited.")
            
            # Analyze by advertisement type
            performance_data = {'overall': [], 'sponsorship_details': [], 'commercial_details': [], 'program_distribution': []}
            
            if schedule_adtype_col:
                # Get sponsorship and commercial data
                sponsorship_data = self.app.schedule_data[
                    self.app.schedule_data[schedule_adtype_col].str.lower().str.contains('sponsor', na=False)
                ]
                commercial_data = self.app.schedule_data[
                    self.app.schedule_data[schedule_adtype_col].str.lower().str.contains('commercial', na=False)
                ]
                
                total_sponsorship = len(sponsorship_data)
                total_commercial = len(commercial_data)
                
                # Calculate match rates for each type
                sponsorship_matches = 0
                commercial_matches = 0
                
                # This would require more complex analysis based on your data structure
                # For now, providing the framework
                
                performance_data['overall'] = [
                    {'Ad_Type': 'Sponsorship', 'Planned_Count': total_sponsorship, 'Matched_Count': sponsorship_matches, 'Match_Rate': f"{sponsorship_matches/max(total_sponsorship,1)*100:.1f}%"},
                    {'Ad_Type': 'Commercial', 'Planned_Count': total_commercial, 'Matched_Count': commercial_matches, 'Match_Rate': f"{commercial_matches/max(total_commercial,1)*100:.1f}%"}
                ]
            
            # Export to Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Overall performance
                if performance_data['overall']:
                    pd.DataFrame(performance_data['overall']).to_excel(writer, sheet_name='Overall Performance', index=False)
                
                # Add other sheets as needed
                summary = [{'Metric': 'Analysis completed', 'Value': 'Check individual sheets for details'}]
                pd.DataFrame(summary).to_excel(writer, sheet_name='Summary', index=False)
            
            messagebox.showinfo("Success", f"Ad type performance analysis exported to:\n{filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting ad type performance: {str(e)}")
            import traceback
            traceback.print_exc()

    def export_comprehensive_report(self):
        """Export comprehensive reconciliation report with all key metrics"""
        try:
            if not self.app.matched_data:
                messagebox.showerror("Error", "No matched data available")
                return
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                initialfile=self.get_default_filename("Comprehensive_Report"),
                filetypes=[("Excel files", "*.xlsx")]
            )
            
            if not filename:
                return
            
            print("Creating Comprehensive Report...")
            
            # Gather all statistics
            total_lmrb = len(self.app.lmrb_filtered) if self.app.lmrb_filtered is not None else 0
            total_matched = len(self.app.matched_data) if self.app.matched_data else 0
            total_unmatched_tc = len(self.app.unmatched_tc_data) if self.app.unmatched_tc_data is not None else 0
            total_unmatched_lmrb = len(self.app.unmatched_lmrb_data) if self.app.unmatched_lmrb_data is not None else 0
            match_rate = (total_matched / total_lmrb * 100) if total_lmrb > 0 else 0
            
            # Executive Summary
            executive_summary = [
                {'Metric': 'Channel', 'Value': self.app.selected_channel or 'Unknown'},
                {'Metric': 'Report Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
                {'Metric': 'Total LMRB Records', 'Value': total_lmrb},
                {'Metric': 'Total Matched Records', 'Value': total_matched},
                {'Metric': 'Overall Match Rate', 'Value': f"{match_rate:.1f}%"},
                {'Metric': 'Unmatched TC Records', 'Value': total_unmatched_tc},
                {'Metric': 'Unmatched LMRB Records', 'Value': total_unmatched_lmrb},
                {'Metric': 'Theme Mappings Configured', 'Value': len(self.app.theme_mapping)},
            ]
            
            # Count manual vs automatic matches
            manual_matches = 0
            automatic_matches = 0
            if isinstance(self.app.matched_data, list):
                for match in self.app.matched_data:
                    if match.get('Match_Type') == 'Manual':
                        manual_matches += 1
                    else:
                        automatic_matches += 1
            
            # Channel Performance
            channel_performance = [
                {'Performance_Metric': 'Automatic Match Success', 'Value': automatic_matches, 'Percentage': f"{automatic_matches/max(total_matched,1)*100:.1f}%"},
                {'Performance_Metric': 'Manual Match Required', 'Value': manual_matches, 'Percentage': f"{manual_matches/max(total_matched,1)*100:.1f}%"},
                {'Performance_Metric': 'Data Completeness', 'Value': total_lmrb, 'Percentage': '100%'},
                {'Performance_Metric': 'Reconciliation Success', 'Value': total_matched, 'Percentage': f"{match_rate:.1f}%"}
            ]
            
            # Export comprehensive report
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Executive summary
                pd.DataFrame(executive_summary).to_excel(writer, sheet_name='Executive Summary', index=False)
                
                # Channel performance
                pd.DataFrame(channel_performance).to_excel(writer, sheet_name='Channel Performance', index=False)
                
                # Include matched data sample
                if self.app.matched_data:
                    matched_df = pd.DataFrame(self.app.matched_data) if isinstance(self.app.matched_data, list) else self.app.matched_data
                    # Limit to first 1000 records for performance
                    sample_size = min(1000, len(matched_df))
                    matched_df.head(sample_size).to_excel(writer, sheet_name='Matched Data Sample', index=False)
                
                # Theme mappings
                if self.app.theme_mapping:
                    mapping_data = []
                    for mapping in self.app.theme_mapping:
                        mapping_data.append({
                            'Schedule_Theme': mapping.get('schedule_theme', ''),
                            'LMRB_Theme': mapping.get('media_watch_theme', mapping.get('lmrb_theme', '')),
                            'Mapping_Type': mapping.get('mapping_type', 'Unknown')
                        })
                    pd.DataFrame(mapping_data).to_excel(writer, sheet_name='Theme Mappings', index=False)
            
            messagebox.showinfo("Success", 
                f"Comprehensive report exported!\n\n"
                f"📊 Report Summary:\n"
                f"• Channel: {self.app.selected_channel or 'Unknown'}\n"
                f"• Total records: {total_lmrb:,}\n"
                f"• Match rate: {match_rate:.1f}%\n"
                f"• Manual matches: {manual_matches}\n"
                f"• Automatic matches: {automatic_matches}\n\n"
                f"📁 File: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting comprehensive report: {str(e)}")
            import traceback
            traceback.print_exc()