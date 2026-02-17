def highlight_matched_rows_excel(media_watch_filtered, matched_mw_original, highlight_color="#FFFF00"):
    """
    Highlight rows in media_watch_filtered that match with rows in matched_mw_original.
    Works with both automatic and manual matches.
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    import re

    # Convert to DataFrame if not already
    if not isinstance(media_watch_filtered, pd.DataFrame):
        media_watch_filtered = pd.DataFrame(media_watch_filtered)

    if not isinstance(matched_mw_original, pd.DataFrame):
        matched_mw_original = pd.DataFrame(matched_mw_original)

    print(f"DEBUG: Media Watch filtered: {len(media_watch_filtered)} rows")
    print(f"DEBUG: Matched data: {len(matched_mw_original)} rows")

    # Find key columns that uniquely identify a row for exact matching
    def find_key_columns(df):
        columns = {}
        columns['theme'] = next((c for c in df.columns if re.search(r'theme|product detail', c.lower())), None)
        columns['date'] = next((c for c in df.columns if 'date' in c.lower()), None)
        columns['time'] = next((c for c in df.columns if re.search(r'time|air', c.lower()) and not 'prog' in c.lower()), None)
        columns['program'] = next((c for c in df.columns if 'program' in c.lower()), None)
        columns['duration'] = next((c for c in df.columns if 'dur' in c.lower()), None)
        return columns

    mw_cols = find_key_columns(media_watch_filtered)
    match_cols = find_key_columns(matched_mw_original)

    print(f"DEBUG: Media Watch key columns: {mw_cols}")
    print(f"DEBUG: Matched data key columns: {match_cols}")

    # Create a set of tuples with normalized values from matched data for exact comparison
    matched_signatures = set()
    for _, match_row in matched_mw_original.iterrows():
        signature = (
            str(match_row.get(match_cols['theme'], '')).lower().strip() if match_cols['theme'] else '',
            str(match_row.get(match_cols['date'], '')).lower().strip() if match_cols['date'] else '',
            str(match_row.get(match_cols['time'], '')).lower().strip() if match_cols['time'] else '',
            str(match_row.get(match_cols['program'], '')).lower().strip() if match_cols['program'] else '',
            str(match_row.get(match_cols['duration'], '')).lower().strip() if match_cols['duration'] else ''
        )
        matched_signatures.add(signature)

    print(f"DEBUG: Created {len(matched_signatures)} unique matched signatures")

    # Track which rows to highlight
    rows_to_highlight = set()

    # Check each row in media_watch against matched signatures
    for i, mw_row in media_watch_filtered.iterrows():
        mw_signature = (
            str(mw_row.get(mw_cols['theme'], '')).lower().strip() if mw_cols['theme'] else '',
            str(mw_row.get(mw_cols['date'], '')).lower().strip() if mw_cols['date'] else '',
            str(mw_row.get(mw_cols['time'], '')).lower().strip() if mw_cols['time'] else '',
            str(mw_row.get(mw_cols['program'], '')).lower().strip() if mw_cols['program'] else '',
            str(mw_row.get(mw_cols['duration'], '')).lower().strip() if mw_cols['duration'] else ''
        )
        
        if mw_signature in matched_signatures:
            rows_to_highlight.add(i)

    print(f"DEBUG: Found {len(rows_to_highlight)} rows to highlight")

    # Create workbook and continue with highlighting as before
    # [rest of the existing code...]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered LMRB Data"

    # Add headers
    for col_idx, column in enumerate(media_watch_filtered.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Create highlight fill
    highlight_fill = PatternFill(
        start_color=highlight_color.lstrip('#'),
        end_color=highlight_color.lstrip('#'),
        fill_type="solid"
    )

    # Add data and highlight matched rows
    for row_idx, (i, row) in enumerate(media_watch_filtered.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Highlight if this row should be highlighted
            if i in rows_to_highlight:
                cell.fill = highlight_fill

    # Adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(media_watch_filtered.columns[col_idx-1])) + 2
        
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
        
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    return wb