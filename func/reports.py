import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter



def generate_comprehensive_summary_report(matched_df, original_media_watch, original_schedule):
    """Generate a detailed reconciliation summary report with durations and calculations."""
    if matched_df.empty:
        return None

    # FIND DURATION COLUMN NAMES IN ALL DATASETS
    duration_col_matched = None
    duration_col_schedule = None

    # Check for duration column in matched data
    for col_name in ['Dur', 'Duration', 'DURATION']:
        if col_name in matched_df.columns:
            duration_col_matched = col_name
            break

    # Check for duration column in schedule data
    if original_schedule is not None:
        for col_name in ['Dur', 'Duration', 'DURATION']:
            if col_name in original_schedule.columns:
                duration_col_schedule = col_name
                break

    # First, make sure we have Schedule_Theme for proper grouping
    if 'Schedule_Theme' not in matched_df.columns:
        if 'Media_Watch_Theme' in matched_df.columns:
            matched_df['Schedule_Theme'] = matched_df['Media_Watch_Theme']

    # Group by Schedule theme and duration if available
    group_by_columns = []
    if 'Schedule_Theme' in matched_df.columns:
        group_by_columns.append('Schedule_Theme')

    if duration_col_matched:
        group_by_columns.append(duration_col_matched)
        # Standardize to 'Dur' for output
        if duration_col_matched != 'Dur':
            matched_df['Dur'] = matched_df[duration_col_matched]
            group_by_columns.remove(duration_col_matched)
            group_by_columns.append('Dur')

    # If no groupable columns, use Media_Watch_Theme
    if not group_by_columns and 'Media_Watch_Theme' in matched_df.columns:
        group_by_columns = ['Media_Watch_Theme']
        if 'Theme_Duration' in matched_df.columns:
            matched_df['Dur'] = matched_df['Theme_Duration']
            group_by_columns.append('Dur')

    # Basic summary with aired count
    if group_by_columns:
        summary = matched_df.groupby(group_by_columns).size().reset_index(name='Aired_Count')
    else:
        # Fallback if no groupable columns
        summary = pd.DataFrame({
            'Total': ['All Records'],
            'Aired_Count': [len(matched_df)]
        })

    # If schedule data is available, get scheduled counts
    if original_schedule is not None and 'Schedule_Theme' in summary.columns:
        if duration_col_schedule:
            # Standardize schedule duration column
            schedule_copy = original_schedule.copy()
            if duration_col_schedule != 'Dur':
                schedule_copy['Dur'] = schedule_copy[duration_col_schedule]
            
            # Group schedule by theme and duration
            schedule_summary = schedule_copy.groupby(['Advt_Theme', 'Dur']).size().reset_index(name='Planned_Count')
            schedule_summary.rename(columns={'Advt_Theme': 'Schedule_Theme'}, inplace=True)
            
            # Merge with summary
            summary = pd.merge(summary, schedule_summary, on=['Schedule_Theme', 'Dur'], how='left')
        else:
            # Group by theme only if no duration
            schedule_summary = original_schedule.groupby('Advt_Theme').size().reset_index(name='Planned_Count')
            schedule_summary.rename(columns={'Advt_Theme': 'Schedule_Theme'}, inplace=True)
            
            # Merge with summary
            summary = pd.merge(summary, schedule_summary, on='Schedule_Theme', how='left')

    # Fill NaN values and calculate metrics
    if 'Planned_Count' in summary.columns:
        summary['Planned_Count'] = summary['Planned_Count'].fillna(0).astype(int)
        summary['Missed_Count'] = summary['Planned_Count'] - summary['Aired_Count']
        summary['Missed_Count'] = summary['Missed_Count'].apply(lambda x: max(0, x))
        
        # Add extra spots count
        summary['Extra_Count'] = summary['Aired_Count'] - summary['Planned_Count']
        summary['Extra_Count'] = summary['Extra_Count'].apply(lambda x: max(0, x))
        
        # Calculate compliance percentage
        summary['Compliance_Rate'] = (summary['Aired_Count'] / summary['Planned_Count'] * 100).fillna(0)
        summary['Compliance_Rate'] = summary['Compliance_Rate'].clip(upper=100)

    # Calculate 30-second equivalent durations
    if 'Dur' in summary.columns:
        if 'Planned_Count' in summary.columns:
            summary['Planned_Duration_Secs'] = summary['Planned_Count'] * summary['Dur']
            summary['Planned_30s_Equiv'] = (summary['Planned_Duration_Secs'] / 30).round(1)
            
        summary['Aired_Duration_Secs'] = summary['Aired_Count'] * summary['Dur']
        summary['Aired_30s_Equiv'] = (summary['Aired_Duration_Secs'] / 30).round(1)
        
        if 'Missed_Count' in summary.columns:
            summary['Missed_Duration_Secs'] = summary['Missed_Count'] * summary['Dur']
            summary['Missed_30s_Equiv'] = (summary['Missed_Duration_Secs'] / 30).round(1)
        
        if 'Extra_Count' in summary.columns:
            summary['Extra_Duration_Secs'] = summary['Extra_Count'] * summary['Dur']
            summary['Extra_30s_Equiv'] = (summary['Extra_Duration_Secs'] / 30).round(1)

    # Add match status breakdown
    if 'Match_Status' in matched_df.columns:
        try:
            status_counts = matched_df.groupby(group_by_columns + ['Match_Status']).size().reset_index(name='Status_Count')
            
            # Create pivot to spread status counts across columns
            status_pivot = status_counts.pivot_table(
                index=group_by_columns,
                columns='Match_Status',
                values='Status_Count',
                fill_value=0
            ).reset_index()
            
            # Merge with summary
            if set(group_by_columns).issubset(set(summary.columns)):
                summary = pd.merge(summary, status_pivot, on=group_by_columns, how='left')
            
            # Fill NAs with 0
            for col in status_pivot.columns:
                if col not in group_by_columns and col in summary.columns:
                    summary[col] = summary[col].fillna(0).astype(int)
        except Exception:
            pass  # Skip status breakdown if pivot fails

    # Calculate totals for numeric columns
    numeric_cols = summary.select_dtypes(include=['int', 'float']).columns.tolist()
    summary_totals = summary[numeric_cols].sum().to_frame().T

    # Add placeholder values for non-numeric columns
    for col in summary.columns:
        if col not in numeric_cols:
            summary_totals[col] = "Total"

    # Reorder totals row columns to match summary
    summary_totals = summary_totals[summary.columns]

    # Append totals row
    summary = pd.concat([summary, summary_totals], ignore_index=True)

    return summary


def create_tc_vs_lmrb_excel_report(matched_df, program_mismatched_df, unmatched_df, time_matched_program_mismatched_df, media_watch_filtered, filename="tc_vs_lmrb_report.xlsx"):
    """Create a comprehensive Excel report with multiple sheets for TC vs LMRB matching."""
    wb = Workbook()

    # Create Summary Sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    # Add title and header
    summary_sheet.merge_cells('A1:H1')
    title_cell = summary_sheet['A1']
    title_cell.value = "TC vs LMRB Reconciliation Report"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')

    # Add metadata
    summary_sheet['A3'] = "Report Date:"
    summary_sheet['B3'] = pd.Timestamp.now().strftime("%Y-%m-%d")

    # Add match statistics
    summary_sheet['A5'] = "Match Statistics"
    summary_sheet['A5'].font = Font(bold=True)

    summary_sheet['A6'] = "Total LMRB Records:"
    summary_sheet['B6'] = len(media_watch_filtered) if media_watch_filtered is not None else 0

    summary_sheet['A7'] = "Full Matches Found:"
    summary_sheet['B7'] = len(matched_df) if matched_df is not None else 0

    summary_sheet['A8'] = "Program Mismatches:"
    summary_sheet['B8'] = len(program_mismatched_df) if program_mismatched_df is not None else 0

    summary_sheet['A9'] = "Time Matched / Program Mismatched:"
    summary_sheet['B9'] = len(time_matched_program_mismatched_df) if time_matched_program_mismatched_df is not None else 0

    # Calculate match rate
    if media_watch_filtered is not None and len(media_watch_filtered) > 0:
        match_rate = (len(matched_df) / len(media_watch_filtered)) * 100 if matched_df is not None else 0
        summary_sheet['A10'] = "Match Rate:"
        summary_sheet['B10'] = f"{match_rate:.2f}%"

    # Create Full Matches Sheet if available
    if matched_df is not None and not matched_df.empty:
        full_matches_sheet = wb.create_sheet("Full Matches")
        
        # Add headers
        for col_idx, column in enumerate(matched_df.columns, 1):
            cell = full_matches_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(matched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = full_matches_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Create Program Mismatches Sheet if available
    if program_mismatched_df is not None and not program_mismatched_df.empty:
        program_mismatch_sheet = wb.create_sheet("Program Mismatches")
        
        # Add headers
        for col_idx, column in enumerate(program_mismatched_df.columns, 1):
            cell = program_mismatch_sheet.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row_idx, row in enumerate(program_mismatched_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = program_mismatch_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths for all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Check header length
            if col_idx <= len(ws[1]):
                max_length = len(str(ws[1][col_idx].value))
            
            # Check values in each row
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            ws.column_dimensions[column_letter].width = max_length + 2

    return wb